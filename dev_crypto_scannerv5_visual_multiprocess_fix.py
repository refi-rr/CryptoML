"""
Crypto Futures Scanner Pro - Ultimate Edition
Enhanced with: Parallel Processing, Database, Alerts, Backtesting, Portfolio Tracking
Author: AI Assistant
Version: 2.0
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
from datetime import datetime, timedelta
import time
from ta.momentum import RSIIndicator, StochRSIIndicator
from ta.trend import MACD, EMAIndicator, SMAIndicator
from ta.volume import VolumeWeightedAveragePrice
from ta.volatility import BollingerBands, AverageTrueRange
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px
from concurrent.futures import ThreadPoolExecutor, as_completed
from concurrent.futures import ProcessPoolExecutor
import sqlite3
import futures_metrics_wrapper as fm
import concurrent.futures
import requests
import streamlit.components.v1 as components
from liquidation_wrappers import get_liquidation_data
from datetime import datetime
import pytz
import json
import warnings
warnings.filterwarnings('ignore')

# === Multiprocessing Safety Guard ===
import multiprocessing

if multiprocessing.get_start_method(allow_none=True) is None:
    try:
        multiprocessing.set_start_method("fork")
    except RuntimeError:
        pass

# ====================================


# ==================== CONFIGURATION ====================
st.set_page_config(
    page_title="Crypto Futures Scanner Development - Ultimate", 
    layout="wide", 
    page_icon="🚀",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.8rem; 
        color: #1f77b4; 
        text-align: center; 
        font-weight: bold;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
        margin-bottom: 1rem;
    }
    .signal-long {
        background: linear-gradient(135deg, #28a745, #20c997);
        color: white; 
        padding: 15px; 
        border-radius: 10px; 
        font-weight: bold; 
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .signal-short {
        background: linear-gradient(135deg, #dc3545, #fd7e14);
        color: white; 
        padding: 15px; 
        border-radius: 10px; 
        font-weight: bold; 
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .signal-neutral {
        background: linear-gradient(135deg, #6c757d, #adb5bd);
        color: white; 
        padding: 15px; 
        border-radius: 10px; 
        font-weight: bold; 
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px;
        border-radius: 10px;
        color: white;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    }
</style>
""", unsafe_allow_html=True)

# ==================== SESSION STATE ====================
if 'last_refresh' not in st.session_state:
    st.session_state.last_refresh = datetime.now()
if 'refresh_interval' not in st.session_state:
    st.session_state.refresh_interval = None
if 'telegram_enabled' not in st.session_state:
    st.session_state.telegram_enabled = False
if 'scan_results' not in st.session_state:
    st.session_state.scan_results = []
if 'db_initialized' not in st.session_state:
    st.session_state.db_initialized = False

# ==================== DATABASE SETUP ====================
DB_NAME = 'crypto_scanner_ultimate.db'

def init_database():
    """Initialize SQLite database with all tables"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Signals table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS signals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            symbol TEXT NOT NULL,
            exchange TEXT,
            signal TEXT,
            strength REAL,
            price REAL,
            entry REAL,
            tp1 REAL,
            tp2 REAL,
            sl REAL,
            risk_reward REAL,
            status TEXT DEFAULT 'ACTIVE',
            notes TEXT
        )
    ''')
    
    # Performance tracking table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS performance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            signal_id INTEGER,
            entry_time DATETIME,
            exit_time DATETIME,
            entry_price REAL,
            exit_price REAL,
            pnl_percentage REAL,
            pnl_amount REAL,
            hit_tp1 BOOLEAN DEFAULT 0,
            hit_tp2 BOOLEAN DEFAULT 0,
            hit_sl BOOLEAN DEFAULT 0,
            exit_reason TEXT,
            FOREIGN KEY (signal_id) REFERENCES signals(id)
        )
    ''')
    
    # Scan history table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scan_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            total_scanned INTEGER,
            signals_found INTEGER,
            avg_strength REAL,
            exchange TEXT,
            scan_duration REAL
        )
    ''')
    
    # Watchlist table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS watchlist (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            symbol TEXT UNIQUE NOT NULL,
            added_date DATETIME DEFAULT CURRENT_TIMESTAMP,
            notes TEXT
        )
    ''')
    
    conn.commit()
    conn.close()
    return True

def save_signal_to_db(result, exchange):
    """Save trading signal to database"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        plan = result.get('trading_plan', {})
        cursor.execute('''
            INSERT INTO signals 
            (symbol, exchange, signal, strength, price, entry, tp1, tp2, sl, risk_reward, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            result['symbol'],
            exchange,
            result['consensus'],
            result['strength'],
            result['price'],
            plan.get('entry'),
            plan.get('tp1'),
            plan.get('tp2'),
            plan.get('sl'),
            plan.get('risk_reward_tp1'),
            json.dumps(result['timeframes']['4h'].get('reasons', []))
        ))
        
        conn.commit()
        signal_id = cursor.lastrowid
        conn.close()
        return signal_id
    except Exception as e:
        st.error(f"Database error: {e}")
        return None

def save_scan_history(total_scanned, signals_found, avg_strength, exchange, duration):
    """Save scan history"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO scan_history 
            (total_scanned, signals_found, avg_strength, exchange, scan_duration)
            VALUES (?, ?, ?, ?, ?)
        ''', (total_scanned, signals_found, avg_strength, exchange, duration))
        
        conn.commit()
        conn.close()
    except Exception as e:
        st.error(f"Scan history error: {e}")

def get_active_signals():
    """Get active signals from database"""
    try:
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql('''
            SELECT * FROM signals 
            WHERE status = 'ACTIVE'
            ORDER BY timestamp DESC
        ''', conn)
        conn.close()
        return df
    except:
        return pd.DataFrame()

def update_signal_status(signal_id, status, exit_price=None, pnl=None):
    """Update signal status"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE signals 
            SET status = ?
            WHERE id = ?
        ''', (status, signal_id))
        
        if exit_price and pnl:
            cursor.execute('''
                INSERT INTO performance
                (signal_id, exit_time, exit_price, pnl_percentage)
                VALUES (?, ?, ?, ?)
            ''', (signal_id, datetime.now(), exit_price, pnl))
        
        conn.commit()
        conn.close()
    except Exception as e:
        st.error(f"Update error: {e}")

# ==================== API CONFIGURATION ====================
API_SOURCES = {
    'binance': {
        'name': 'Binance Futures',
        'base_url': 'https://fapi.binance.com',
        'top_symbols_endpoint': '/fapi/v1/ticker/24hr',
        'klines_endpoint': '/fapi/v1/klines'
    },
    'bybit': {
        'name': 'Bybit',
        'base_url': 'https://api.bybit.com',
        'top_symbols_endpoint': '/v5/market/tickers',
        'klines_endpoint': '/v5/market/kline'
    },
    'gateio': {
        'name': 'Gate.io',
        'base_url': 'https://api.gateio.ws/api/v4',
        'top_symbols_endpoint': '/futures/usdt/contracts',
        'klines_endpoint': '/futures/usdt/candlesticks'
    }
}

TIMEFRAME_MAPPING = {
    'binance': {'1h': '1h', '4h': '4h', '12h': '12h'},
    'bybit': {'1h': '60', '4h': '240', '12h': '720'},
    'gateio': {'1h': '1h', '4h': '4h', '12h': '12h'}
}

# ==================== RATE LIMITER ====================
class RateLimiter:
    """Rate limiter to prevent API throttling"""
    def __init__(self, max_calls=10, period=60):
        self.max_calls = max_calls
        self.period = period
        self.calls = []
    
    def wait_if_needed(self):
        now = time.time()
        self.calls = [c for c in self.calls if now - c < self.period]
        
        if len(self.calls) >= self.max_calls:
            sleep_time = self.period - (now - self.calls[0])
            if sleep_time > 0:
                time.sleep(sleep_time)
                self.calls = []
        
        self.calls.append(now)

rate_limiter = RateLimiter(max_calls=15, period=40)

# ==================== DATA FETCHING ====================
@st.cache_data(ttl=200)
def get_binance_top_symbols(limit=40):
    """Fetch top trading pairs from Binance Futures"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['binance']['base_url']}{API_SOURCES['binance']['top_symbols_endpoint']}"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        usdt_pairs = [item for item in data if item['symbol'].endswith('USDT')]
        sorted_pairs = sorted(usdt_pairs, key=lambda x: float(x.get('quoteVolume', 0)), reverse=True)
        return sorted_pairs[:limit]
    except Exception as e:
        return []

@st.cache_data(ttl=200)
def get_bybit_top_symbols(limit=40):
    """Fetch top trading pairs from Bybit"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['bybit']['base_url']}{API_SOURCES['bybit']['top_symbols_endpoint']}"
        params = {'category': 'linear'}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get('retCode') == 0:
            symbols = data['result']['list']
            usdt_pairs = [s for s in symbols if s['symbol'].endswith('USDT')]
            sorted_pairs = sorted(usdt_pairs, key=lambda x: float(x.get('turnover24h', 0)), reverse=True)
            return sorted_pairs[:limit]
        return []
    except Exception as e:
        return []

@st.cache_data(ttl=200)
def get_gateio_top_symbols(limit=40):
    """Fetch top trading pairs from Gate.io"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['gateio']['base_url']}{API_SOURCES['gateio']['top_symbols_endpoint']}"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        active_contracts = [c for c in data if c.get('trade_size', 0) > 0]
        sorted_contracts = sorted(active_contracts, key=lambda x: float(x.get('trade_size', 0)), reverse=True)
        return sorted_contracts[:limit]
    except Exception as e:
        return []

def get_top_symbols(source='binance', limit=40):
    """Get top symbols with automatic fallback"""
    sources_priority = ['binance', 'bybit', 'gateio']
    if source in sources_priority:
        sources_priority.remove(source)
        sources_priority.insert(0, source)
    
    for src in sources_priority:
        try:
            if src == 'binance':
                symbols = get_binance_top_symbols(limit)
            elif src == 'bybit':
                symbols = get_bybit_top_symbols(limit)
            elif src == 'gateio':
                symbols = get_gateio_top_symbols(limit)
            else:
                continue
            if symbols:
                return src, symbols
        except:
            continue
    return None, []

@st.cache_data(ttl=150)
def get_binance_klines(symbol, interval, limit=600):
    """Fetch klines from Binance Futures"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['binance']['base_url']}{API_SOURCES['binance']['klines_endpoint']}"
        params = {'symbol': symbol, 'interval': interval, 'limit': limit}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        df = pd.DataFrame(data, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume',
            'close_time', 'quote_volume', 'trades', 'taker_buy_base', 'taker_buy_quote', 'ignore'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        for col in ['open', 'high', 'low', 'close', 'volume']:
            df[col] = df[col].astype(float)
        return df
    except Exception as e:
        return None

@st.cache_data(ttl=180)
def get_bybit_klines(symbol, interval, limit=600):
    """Fetch klines from Bybit"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['bybit']['base_url']}{API_SOURCES['bybit']['klines_endpoint']}"
        params = {'category': 'linear', 'symbol': symbol, 'interval': interval, 'limit': limit}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get('retCode') == 0:
            klines = data['result']['list']
            df = pd.DataFrame(klines, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'turnover'])
            df['timestamp'] = pd.to_datetime(df['timestamp'].astype(float), unit='ms')
            for col in ['open', 'high', 'low', 'close', 'volume']:
                df[col] = df[col].astype(float)
            df = df.sort_values('timestamp').reset_index(drop=True)
            return df
        return None
    except Exception as e:
        return None

@st.cache_data(ttl=180)
def get_gateio_klines(symbol, interval, limit=600):
    """Fetch klines from Gate.io"""
    try:
        rate_limiter.wait_if_needed()
        contract = symbol.replace('USDT', '_USDT')
        url = f"{API_SOURCES['gateio']['base_url']}{API_SOURCES['gateio']['klines_endpoint']}"
        params = {'contract': contract, 'interval': interval, 'limit': limit}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        if isinstance(data, list):
            df = pd.DataFrame(data, columns=['timestamp', 'volume', 'close', 'high', 'low', 'open'])
            df['timestamp'] = pd.to_datetime(df['timestamp'].astype(float), unit='s')
            for col in ['open', 'high', 'low', 'close', 'volume']:
                df[col] = df[col].astype(float)
            df = df.sort_values('timestamp').reset_index(drop=True)
            return df
        return None
    except Exception as e:
        return None

def get_klines(source, symbol, timeframe, limit=600):
    """Get klines with automatic fallback"""
    sources_priority = ['binance', 'bybit', 'gateio']
    if source in sources_priority:
        sources_priority.remove(source)
        sources_priority.insert(0, source)
    
    for src in sources_priority:
        try:
            if src not in TIMEFRAME_MAPPING:
                continue
            interval = TIMEFRAME_MAPPING[src][timeframe]
            if src == 'binance':
                df = get_binance_klines(symbol, interval, limit)
            elif src == 'bybit':
                df = get_bybit_klines(symbol, interval, limit)
            elif src == 'gateio':
                df = get_gateio_klines(symbol, interval, limit)
            else:
                continue
            if df is not None and len(df) >= 200:
                return df
        except:
            continue
    return None

# ==================== TECHNICAL INDICATORS ====================
def calculate_indicators(df):
    """Calculate comprehensive technical indicators"""
    if df is None or len(df) < 200:
        return None
    
    df = df.copy()
    close = df['close']
    high = df['high']
    low = df['low']
    volume = df['volume']
    
    # RSI
    rsi = RSIIndicator(close=close, window=14)
    df['rsi'] = rsi.rsi()
    
    # Stochastic RSI
    stoch_rsi = StochRSIIndicator(close=close, window=14, smooth1=3, smooth2=3)
    df['stoch_rsi_k'] = stoch_rsi.stochrsi_k() * 100
    df['stoch_rsi_d'] = stoch_rsi.stochrsi_d() * 100
    
    # EMAs
    df['ema_9'] = EMAIndicator(close=close, window=9).ema_indicator()
    df['ema_21'] = EMAIndicator(close=close, window=21).ema_indicator()
    df['ema_50'] = EMAIndicator(close=close, window=50).ema_indicator()
    df['ema_200'] = EMAIndicator(close=close, window=200).ema_indicator()
    
    # SMAs
    df['sma_20'] = SMAIndicator(close=close, window=20).sma_indicator()
    df['sma_50'] = SMAIndicator(close=close, window=50).sma_indicator()
    
    # MACD
    macd = MACD(close=close, window_slow=26, window_fast=12, window_sign=9)
    df['macd'] = macd.macd()
    df['macd_signal'] = macd.macd_signal()
    df['macd_diff'] = macd.macd_diff()
    
    # Bollinger Bands
    bb = BollingerBands(close=close, window=20, window_dev=2)
    df['bb_upper'] = bb.bollinger_hband()
    df['bb_middle'] = bb.bollinger_mavg()
    df['bb_lower'] = bb.bollinger_lband()
    df['bb_width'] = (df['bb_upper'] - df['bb_lower']) / df['bb_middle']
    
    # ATR
    atr = AverageTrueRange(high=high, low=low, close=close, window=14)
    df['atr'] = atr.average_true_range()
    
    # Volume
    df['volume_sma'] = volume.rolling(window=20).mean()
    df['volume_ratio'] = volume / df['volume_sma']
    
    # VWAP
    df['vwap'] = (volume * (high + low + close) / 3).cumsum() / volume.cumsum()
    
    # ADX
    df['adx'] = calculate_adx(df, 14)
    
    # Momentum
    df['momentum'] = close.pct_change(periods=10) * 100
    
    # Pivot Points
    df = calculate_pivot_points(df)
    
    return df

def calculate_adx(df, period=14):
    """Calculate ADX"""
    high = df['high']
    low = df['low']
    close = df['close']
    
    plus_dm = high.diff()
    minus_dm = low.diff()
    plus_dm[plus_dm < 0] = 0
    minus_dm[minus_dm > 0] = 0
    
    tr = pd.concat([high - low, abs(high - close.shift()), abs(low - close.shift())], axis=1).max(axis=1)
    atr = tr.rolling(window=period).mean()
    plus_di = 100 * (plus_dm.rolling(window=period).mean() / atr)
    minus_di = 100 * (abs(minus_dm).rolling(window=period).mean() / atr)
    dx = 100 * abs(plus_di - minus_di) / (plus_di + minus_di)
    adx = dx.rolling(window=period).mean()
    return adx

def calculate_pivot_points(df):
    """Calculate pivot points"""
    df['pivot'] = (df['high'] + df['low'] + df['close']) / 3
    df['r1'] = 2 * df['pivot'] - df['low']
    df['s1'] = 2 * df['pivot'] - df['high']
    df['r2'] = df['pivot'] + (df['high'] - df['low'])
    df['s2'] = df['pivot'] - (df['high'] - df['low'])
    return df

def detect_chart_patterns(df):
    """Detect chart patterns"""
    patterns = []
    if len(df) < 50:
        return patterns
    
    # Golden/Death Cross
    if len(df) >= 200:
        if df['ema_50'].iloc[-2] <= df['ema_200'].iloc[-2] and df['ema_50'].iloc[-1] > df['ema_200'].iloc[-1]:
            patterns.append('GOLDEN_CROSS')
        elif df['ema_50'].iloc[-2] >= df['ema_200'].iloc[-2] and df['ema_50'].iloc[-1] < df['ema_200'].iloc[-1]:
            patterns.append('DEATH_CROSS')
    
    # Divergence
    recent = df.tail(50)
    close = recent['close'].values
    price_trend = close[-10:] - close[-20:-10].mean()
    rsi_trend = df['rsi'].iloc[-10:].values - df['rsi'].iloc[-20:-10].mean()
    
    if price_trend.mean() < 0 and rsi_trend.mean() > 0:
        patterns.append('BULLISH_DIVERGENCE')
    elif price_trend.mean() > 0 and rsi_trend.mean() < 0:
        patterns.append('BEARISH_DIVERGENCE')
    
    # BB Breakout
    if df['close'].iloc[-1] > df['bb_upper'].iloc[-2]:
        patterns.append('UPPER_BB_BREAKOUT')
    elif df['close'].iloc[-1] < df['bb_lower'].iloc[-2]:
        patterns.append('LOWER_BB_BREAKOUT')
    
    return patterns

# ==================== SIGNAL ANALYSIS ====================
def analyze_signal(df, timeframe):
    """Advanced signal analysis"""
    if df is None or len(df) < 200:
        return {
            'signal': 'INSUFFICIENT_DATA',
            'strength': 0,
            'score': 0,
            'reasons': [],
            'indicators': {}
        }
    
    latest = df.iloc[-1]
    prev = df.iloc[-2]
    
    bullish_signals = 0
    bearish_signals = 0
    reasons = []
    max_score = 10
    
    # RSI
    rsi_val = latest['rsi']
    if rsi_val < 30:
        bullish_signals += 1.5
        reasons.append(f"RSI oversold: {rsi_val:.1f}")
    elif rsi_val > 70:
        bearish_signals += 1.5
        reasons.append(f"RSI overbought: {rsi_val:.1f}")
    elif 30 <= rsi_val <= 40:
        bullish_signals += 0.5
        reasons.append(f"RSI near oversold: {rsi_val:.1f}")
    elif 60 <= rsi_val <= 70:
        bearish_signals += 0.5
        reasons.append(f"RSI near overbought: {rsi_val:.1f}")
    
    # Stoch RSI
    stoch_k = latest['stoch_rsi_k']
    stoch_d = latest['stoch_rsi_d']
    if stoch_k < 20 and stoch_k > stoch_d:
        bullish_signals += 1.5
        reasons.append(f"Stoch RSI bullish: K={stoch_k:.1f}")
    elif stoch_k > 80 and stoch_k < stoch_d:
        bearish_signals += 1.5
        reasons.append(f"Stoch RSI bearish: K={stoch_k:.1f}")
    
    # MACD
    macd_diff = latest['macd_diff']
    macd_cross_up = prev['macd_diff'] <= 0 and macd_diff > 0
    macd_cross_down = prev['macd_diff'] >= 0 and macd_diff < 0
    
    if macd_cross_up:
        bullish_signals += 1.5
        reasons.append("MACD bullish crossover")
    elif macd_cross_down:
        bearish_signals += 1.5
        reasons.append("MACD bearish crossover")
    elif macd_diff > 0:
        bullish_signals += 0.5
        reasons.append("MACD positive")
    elif macd_diff < 0:
        bearish_signals += 0.5
        reasons.append("MACD negative")
    
    # EMA Trend
    ema_alignment_bull = (latest['ema_9'] > latest['ema_21'] > latest['ema_50'])
    ema_alignment_bear = (latest['ema_9'] < latest['ema_21'] < latest['ema_50'])
    
    if ema_alignment_bull:
        bullish_signals += 2
        reasons.append("EMA bullish alignment")
    elif ema_alignment_bear:
        bearish_signals += 2
        reasons.append("EMA bearish alignment")
    
    # Golden/Death Cross
    patterns = detect_chart_patterns(df)
    if 'GOLDEN_CROSS' in patterns:
        bullish_signals += 2
        reasons.append("⭐ Golden Cross")
    elif 'DEATH_CROSS' in patterns:
        bearish_signals += 2
        reasons.append("⭐ Death Cross")
    
    # Volume
    vol_ratio = latest['volume_ratio']
    if vol_ratio > 1.5:
        if bullish_signals > bearish_signals:
            bullish_signals += 1
            reasons.append(f"High volume: {vol_ratio:.1f}x")
        elif bearish_signals > bullish_signals:
            bearish_signals += 1
            reasons.append(f"High volume: {vol_ratio:.1f}x")
    
    # ADX
    adx_val = latest['adx']
    if adx_val > 25:
        reasons.append(f"Strong trend: ADX={adx_val:.1f}")
        if bullish_signals > bearish_signals:
            bullish_signals += 0.5
        elif bearish_signals > bullish_signals:
            bearish_signals += 0.5
    
    # Divergence
    if 'BULLISH_DIVERGENCE' in patterns:
        bullish_signals += 1
        reasons.append("📈 Bullish divergence")
    elif 'BEARISH_DIVERGENCE' in patterns:
        bearish_signals += 1
        reasons.append("📉 Bearish divergence")
    
    # Determine signal
    if bullish_signals > bearish_signals and bullish_signals >= 3:
        signal_type = 'LONG'
        signal_score = bullish_signals
    elif bearish_signals > bullish_signals and bearish_signals >= 3:
        signal_type = 'SHORT'
        signal_score = bearish_signals
    else:
        signal_type = 'NEUTRAL'
        signal_score = 0
    
    strength = min((signal_score / max_score) * 100, 100)
    
    return {
        'signal': signal_type,
        'strength': round(strength, 1),
        'score': round(signal_score, 2),
        'max_score': max_score,
        'reasons': reasons[:5],
        'indicators': {
            'rsi': round(float(rsi_val), 2),
            'stoch_rsi': round(float(stoch_k), 2),
            'macd_diff': round(float(macd_diff), 8),
            'volume_ratio': round(float(vol_ratio), 2),
            'adx': round(float(adx_val), 2) if pd.notna(adx_val) else 0,
            'ema_trend': 'BULL' if ema_alignment_bull else 'BEAR' if ema_alignment_bear else 'NEUTRAL'
        },
        'patterns': patterns
    }

def calculate_support_resistance(df, lookback=100):
    """Calculate support and resistance levels"""
    if df is None or len(df) < lookback:
        return None, None
    
    recent = df.tail(lookback)
    swing_highs = []
    swing_lows = []
    
    for i in range(2, len(recent) - 2):
        if (recent['high'].iloc[i] > recent['high'].iloc[i-1] and 
            recent['high'].iloc[i] > recent['high'].iloc[i-2] and
            recent['high'].iloc[i] > recent['high'].iloc[i+1] and 
            recent['high'].iloc[i] > recent['high'].iloc[i+2]):
            swing_highs.append(recent['high'].iloc[i])
        
        if (recent['low'].iloc[i] < recent['low'].iloc[i-1] and 
            recent['low'].iloc[i] < recent['low'].iloc[i-2] and
            recent['low'].iloc[i] < recent['low'].iloc[i+1] and 
            recent['low'].iloc[i] < recent['low'].iloc[i+2]):
            swing_lows.append(recent['low'].iloc[i])
    
    pivot_support = recent['s1'].iloc[-1]
    pivot_resistance = recent['r1'].iloc[-1]
    recent_high = recent['high'].tail(20).max()
    recent_low = recent['low'].tail(20).min()
    current_price = df['close'].iloc[-1]
    
    # Resistance
    potential_resistances = []
    if swing_highs:
        potential_resistances.extend([h for h in swing_highs if h > current_price])
    potential_resistances.append(pivot_resistance)
    potential_resistances.append(recent_high)
    valid_resistances = [r for r in potential_resistances if current_price < r < current_price * 1.2]
    resistance = min(valid_resistances) if valid_resistances else recent_high
    
    # Support
    potential_supports = []
    if swing_lows:
        potential_supports.extend([l for l in swing_lows if l < current_price])
    potential_supports.append(pivot_support)
    potential_supports.append(recent_low)
    valid_supports = [s for s in potential_supports if current_price * 0.8 < s < current_price]
    support = max(valid_supports) if valid_supports else recent_low
    
    return support, resistance

def generate_entry_tp_sl(df, signal_type, current_price):
    """Generate entry, TP, SL with S/R validation"""
    if df is None or signal_type not in ['LONG', 'SHORT']:
        return None
    
    latest = df.iloc[-1]
    atr = latest['atr']
    support, resistance = calculate_support_resistance(df)
    
    if support and resistance:
        sr_distance = resistance - support
        price_to_support_pct = abs(current_price - support) / current_price * 100
        price_to_resistance_pct = abs(resistance - current_price) / current_price * 100
    else:
        sr_distance = atr * 3
        price_to_support_pct = 5
        price_to_resistance_pct = 5
    
    if signal_type == 'LONG':
        if price_to_support_pct > 3:
            entry = (current_price + support) / 2
        else:
            entry = current_price * 1.002
        if support:
            entry = max(entry, support * 1.01)
        sl = support * 0.995 if support else entry - (atr * 1.5)
        risk = entry - sl
        tp1 = entry + (risk * 2)
        tp2 = entry + (risk * 3.5)
        if resistance and tp1 > resistance * 1.05:
            tp1 = resistance * 0.98
            tp2 = resistance * 1.03
    else:
        if price_to_resistance_pct > 3:
            entry = (current_price + resistance) / 2
        else:
            entry = current_price * 0.998
        if resistance:
            entry = min(entry, resistance * 0.99)
        sl = resistance * 1.005 if resistance else entry + (atr * 1.5)
        risk = sl - entry
        tp1 = entry - (risk * 2)
        tp2 = entry - (risk * 3.5)
        if support and tp1 < support * 0.95:
            tp1 = support * 1.02
            tp2 = support * 0.97
    
    avg_candle_move = df['close'].pct_change().abs().tail(20).mean() * 100
    if avg_candle_move > 0:
        hours_to_entry = abs((entry - current_price) / current_price * 100) / avg_candle_move
        hours_to_tp1 = abs((tp1 - current_price) / current_price * 100) / avg_candle_move
    else:
        hours_to_entry = 0
        hours_to_tp1 = 0
    
    risk_amount = abs(entry - sl)
    reward_tp1 = abs(tp1 - entry)
    reward_tp2 = abs(tp2 - entry)
    
    entry_note = ""
    if signal_type == 'LONG':
        if support and entry < support:
            entry_note = "⚠️ Entry below support"
        elif support and (entry - support) / support * 100 > 5:
            entry_note = "⚠️ Entry far from support - Wait for pullback"
    else:
        if resistance and entry > resistance:
            entry_note = "⚠️ Entry above resistance"
        elif resistance and (resistance - entry) / entry * 100 > 5:
            entry_note = "⚠️ Entry far from resistance - Wait for rally"
    
    return {
        'entry': round(entry, 8),
        'tp1': round(tp1, 8),
        'tp2': round(tp2, 8),
        'sl': round(sl, 8),
        'support': round(support, 8) if support else None,
        'resistance': round(resistance, 8) if resistance else None,
        'risk_reward_tp1': round(reward_tp1 / risk_amount, 2) if risk_amount > 0 else 0,
        'risk_reward_tp2': round(reward_tp2 / risk_amount, 2) if risk_amount > 0 else 0,
        'risk_percentage': round((risk_amount / entry) * 100, 2),
        'est_hours_to_entry': round(hours_to_entry, 1),
        'est_hours_to_tp1': round(hours_to_tp1, 1),
        'entry_note': entry_note,
        'entry_to_support_pct': round(price_to_support_pct, 2) if signal_type == 'LONG' else None,
        'entry_to_resistance_pct': round(price_to_resistance_pct, 2) if signal_type == 'SHORT' else None
    }

# ==================== MULTI-TIMEFRAME ANALYSIS ====================
def analyze_single_coin(source, symbol_data):
    """Analyze single coin across multiple timeframes"""
    try:
        if source == 'binance':
            symbol = symbol_data['symbol']
            name = symbol.replace('USDT', '')
            current_price = float(symbol_data['lastPrice'])
        elif source == 'bybit':
            symbol = symbol_data['symbol']
            name = symbol.replace('USDT', '')
            current_price = float(symbol_data['lastPrice'])
        else:
            symbol = symbol_data['name'].replace('_', '')
            name = symbol.replace('USDT', '')
            current_price = float(symbol_data.get('last_price', 0))
        
        timeframes = ['1h', '4h', '12h']
        results = {}
        
        for tf in timeframes:
            df = get_klines(source, symbol, tf, limit=500)
            if df is not None and len(df) >= 200:
                df = calculate_indicators(df)
                if df is not None:
                    results[tf] = analyze_signal(df, tf)
                    results[f'{tf}_df'] = df
                else:
                    results[tf] = {'signal': 'ERROR', 'strength': 0, 'reasons': [], 'indicators': {}}
            else:
                results[tf] = {'signal': 'INSUFFICIENT_DATA', 'strength': 0, 'reasons': [], 'indicators': {}}
        
        signals = []
        weights = {'12h': 3, '4h': 2, '1h': 1}
        weighted_long = 0
        weighted_short = 0
        
        for tf in timeframes:
            signal = results[tf]['signal']
            if signal == 'LONG':
                weighted_long += weights[tf]
                signals.append('LONG')
            elif signal == 'SHORT':
                weighted_short += weights[tf]
                signals.append('SHORT')
        
        if weighted_long > weighted_short and weighted_long >= 3:
            consensus = 'LONG'
        elif weighted_short > weighted_long and weighted_short >= 3:
            consensus = 'SHORT'
        else:
            consensus = 'NEUTRAL'
        
        total_strength = 0
        total_weight = 0
        for tf in timeframes:
            if results[tf]['signal'] not in ['ERROR', 'INSUFFICIENT_DATA']:
                total_strength += results[tf]['strength'] * weights[tf]
                total_weight += weights[tf]
        
        avg_strength = total_strength / total_weight if total_weight > 0 else 0
        
        trading_plan = None
        if consensus in ['LONG', 'SHORT']:
            main_df = results.get('4h_df')
            if main_df is not None:
                trading_plan = generate_entry_tp_sl(main_df, consensus, current_price)
        
        return {
            'symbol': symbol,
            'name': name,
            'price': current_price,
            'consensus': consensus,
            'strength': round(avg_strength, 1),
            'timeframes': results,
            'trading_plan': trading_plan
        }
    except Exception as e:
        return None

def parallel_scan(symbols, source, max_workers=5):
    """Parallel scanning for faster performance"""
    results = []
    
    #with ThreadPoolExecutor(max_workers=max_workers) as executor:
    #    future_to_symbol = {
    #        executor.submit(analyze_single_coin, source, symbol_data): symbol_data 
    #        for symbol_data in symbols
    #    }
    #    
    #    for future in as_completed(future_to_symbol):
    #        try:
    #            result = future.result(timeout=30)
    #            if result:
    #                results.append(result)
    #        except Exception as e:
    #            continue
    progress = st.progress(0)
    total = len(symbols)
    done = 0

    with ProcessPoolExecutor(max_workers=10) as executor:
        futures = [executor.submit(analyze_single_coin, source, coin) for coin in symbols]
        for future in as_completed(futures):
            done += 1
            progress.progress(done / total)
            result = future.result()
            if result:
                results.append(result)

    st.success("✅ All coins analyzed using multiprocessing!")

    return results

# ==================== TELEGRAM ALERTS ====================
def send_telegram_alert(token, chat_id, symbol, signal_data):
    """Send alert via Telegram"""
    try:
        plan = signal_data.get('trading_plan', {})
        if not plan:
            return
        
        message = f"""
🚨 *NEW SIGNAL ALERT* 🚨

💰 *{symbol}*
📊 Signal: *{signal_data['consensus']}*
💪 Strength: *{signal_data['strength']}%*

📍 Entry: ${plan['entry']:,.4f}
🎯 TP1: ${plan['tp1']:,.4f} (R:R {plan['risk_reward_tp1']})
🎯 TP2: ${plan['tp2']:,.4f} (R:R {plan['risk_reward_tp2']})
🛑 SL: ${plan['sl']:,.4f}

⚡ Risk: {plan['risk_percentage']}%
📈 Multi-TF Confirmation: ✅

⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'Markdown'
        }
        
        response = requests.post(url, data=data, timeout=10)
        return response.status_code == 200
    except Exception as e:
        return False

# ==================== VISUALIZATION ====================
def create_chart(df, symbol, trading_plan=None):
    """Create advanced trading chart"""
    if df is None or len(df) < 50:
        return None
    
    fig = make_subplots(
        rows=5, cols=1, 
        shared_xaxes=True, 
        vertical_spacing=0.02,
        row_heights=[0.4, 0.15, 0.15, 0.15, 0.15],
        subplot_titles=(f'{symbol} Price', 'RSI', 'Stoch RSI', 'MACD', 'Volume')
    )
    
    # Candlestick
    fig.add_trace(go.Candlestick(
        x=df['timestamp'], 
        open=df['open'], 
        high=df['high'],
        low=df['low'], 
        close=df['close'], 
        name='Price'
    ), row=1, col=1)
    
    # EMAs
    colors = {'ema_9': 'purple', 'ema_21': 'orange', 'ema_50': 'blue', 'ema_200': 'red'}
    for ema, color in colors.items():
        fig.add_trace(go.Scatter(
            x=df['timestamp'], 
            y=df[ema], 
            name=ema.upper(),
            line=dict(color=color, width=1)
        ), row=1, col=1)
    
    # Bollinger Bands
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['bb_upper'], 
        name='BB Upper',
        line=dict(color='gray', width=1, dash='dash')
    ), row=1, col=1)
    
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['bb_lower'], 
        name='BB Lower',
        line=dict(color='gray', width=1, dash='dash'), 
        fill='tonexty',
        fillcolor='rgba(128,128,128,0.1)'
    ), row=1, col=1)
    
    # Trading Plan Lines
    if trading_plan:
        for level, color, name in [
            (trading_plan['entry'], 'yellow', 'Entry'),
            (trading_plan['tp1'], 'green', 'TP1'), 
            (trading_plan['tp2'], 'lightgreen', 'TP2'),
            (trading_plan['sl'], 'red', 'SL')
        ]:
            fig.add_hline(
                y=level, 
                line_dash="dash", 
                line_color=color,
                annotation_text=name, 
                row=1, col=1
            )
    
    # RSI
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['rsi'], 
        name='RSI',
        line=dict(color='purple')
    ), row=2, col=1)
    fig.add_hline(y=70, line_dash="dash", line_color="red", row=2, col=1)
    fig.add_hline(y=30, line_dash="dash", line_color="green", row=2, col=1)
    fig.add_hline(y=50, line_dash="dot", line_color="gray", row=2, col=1)
    
    # Stochastic RSI
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['stoch_rsi_k'], 
        name='Stoch K',
        line=dict(color='blue')
    ), row=3, col=1)
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['stoch_rsi_d'], 
        name='Stoch D',
        line=dict(color='orange')
    ), row=3, col=1)
    fig.add_hline(y=80, line_dash="dash", line_color="red", row=3, col=1)
    fig.add_hline(y=20, line_dash="dash", line_color="green", row=3, col=1)
    
    # MACD
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['macd'], 
        name='MACD',
        line=dict(color='blue')
    ), row=4, col=1)
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['macd_signal'], 
        name='Signal',
        line=dict(color='orange')
    ), row=4, col=1)
    fig.add_trace(go.Bar(
        x=df['timestamp'], 
        y=df['macd_diff'], 
        name='Histogram',
        marker_color=['green' if val > 0 else 'red' for val in df['macd_diff']]
    ), row=4, col=1)
    
    # Volume
    fig.add_trace(go.Bar(
        x=df['timestamp'], 
        y=df['volume'], 
        name='Volume',
        marker_color=['green' if df['close'].iloc[i] > df['open'].iloc[i] else 'red' for i in range(len(df))]
    ), row=5, col=1)
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['volume_sma'], 
        name='Vol SMA',
        line=dict(color='orange', width=2)
    ), row=5, col=1)
    
    fig.update_layout(
        height=1200, 
        showlegend=True, 
        xaxis_rangeslider_visible=False, 
        hovermode='x unified',
        template='plotly_dark'
    )
    fig.update_xaxes(title_text="Time", row=5, col=1)
    
    return fig

def create_market_heatmap(results):
    """Create market heatmap visualization"""
    data = []
    for r in results:
        tf_4h = r['timeframes'].get('4h', {})
        indicators = tf_4h.get('indicators', {})
        
        data.append({
            'symbol': r['name'],
            'strength': r['strength'],
            'signal': r['consensus'],
            'volume': indicators.get('volume_ratio', 1)
        })
    
    if not data:
        return None
    
    df = pd.DataFrame(data)
    
    fig = px.treemap(
        df,
        path=['signal', 'symbol'],
        values='strength',
        color='volume',
        color_continuous_scale='RdYlGn',
        title='Market Overview Heatmap'
    )
    
    fig.update_layout(height=600)
    
    return fig

def create_performance_chart():
    """Create performance tracking chart"""
    try:
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql('''
            SELECT 
                DATE(entry_time) as date,
                SUM(pnl_percentage) as daily_pnl,
                COUNT(*) as trades
            FROM performance
            WHERE entry_time >= DATE('now', '-30 days')
            GROUP BY DATE(entry_time)
            ORDER BY date
        ''', conn)
        conn.close()
        
        if len(df) == 0:
            return None
        
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            x=df['date'],
            y=df['daily_pnl'],
            name='Daily PnL %',
            marker_color=['green' if x > 0 else 'red' for x in df['daily_pnl']]
        ))
        
        fig.add_trace(go.Scatter(
            x=df['date'],
            y=df['daily_pnl'].cumsum(),
            name='Cumulative PnL %',
            yaxis='y2',
            line=dict(color='orange', width=3)
        ))
        
        fig.update_layout(
            title='Performance Last 30 Days',
            yaxis=dict(title='Daily PnL %'),
            yaxis2=dict(title='Cumulative PnL %', overlaying='y', side='right'),
            hovermode='x unified',
            height=400
        )
        
        return fig
    except:
        return None

# ==================== PORTFOLIO TRACKER ====================
def display_portfolio_tracker():
    """Display active positions and PnL"""
    st.markdown("### 💼 Active Portfolio")
    
    active_signals = get_active_signals()
    
    if len(active_signals) == 0:
        st.info("📭 No active positions")
        return
    
    total_pnl = 0
    positions_data = []
    
    for idx, row in active_signals.iterrows():
        symbol = row['symbol']
        entry = row['entry']
        signal_type = row['signal']
        
        # Get current price (simplified - use last known price)
        try:
            # Try to fetch current price
            if row['exchange'] == 'binance':
                current_price = entry  # Placeholder
            else:
                current_price = entry
        except:
            current_price = entry
        
        # Calculate PnL
        if signal_type == 'LONG':
            pnl_pct = ((current_price - entry) / entry) * 100
        else:
            pnl_pct = ((entry - current_price) / entry) * 100
        
        total_pnl += pnl_pct
        
        positions_data.append({
            'ID': row['id'],
            'Symbol': symbol,
            'Type': signal_type,
            'Entry': f"${entry:.4f}",
            'Current': f"${current_price:.4f}",
            'PnL%': f"{pnl_pct:+.2f}%",
            'TP1': f"${row['tp1']:.4f}",
            'SL': f"${row['sl']:.4f}",
            'Time': row['timestamp']
        })
    
    # Display metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Active Positions", len(active_signals))
    with col2:
        st.metric("Total PnL", f"{total_pnl:+.2f}%", delta=f"{total_pnl:.2f}%")
    with col3:
        avg_pnl = total_pnl / len(active_signals) if len(active_signals) > 0 else 0
        st.metric("Avg PnL per Trade", f"{avg_pnl:+.2f}%")
    
    # Display positions table
    df_positions = pd.DataFrame(positions_data)
    st.dataframe(df_positions, use_container_width=True, hide_index=True)
    
    # Close position buttons
    st.markdown("#### 🔄 Manage Positions")
    cols = st.columns(len(active_signals))
    for idx, (col, row) in enumerate(zip(cols, active_signals.iterrows())):
        with col:
            if st.button(f"Close {row[1]['symbol']}", key=f"close_{row[1]['id']}"):
                update_signal_status(row[1]['id'], 'CLOSED')
                st.success(f"✅ Closed {row[1]['symbol']}")
                st.rerun()

# ==================== EXPORT FUNCTIONS ====================
def export_to_excel(results):
    """Export scan results to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary['A1'] = "Crypto Futures Scanner - Ultimate Edition"
        ws_summary['A1'].font = Font(size=16, bold=True, color="1F77B4")
        ws_summary['A3'] = f"Scan Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A4'] = f"Total Analyzed: {len(results)}"
        ws_summary['A5'] = f"Long Signals: {len([r for r in results if r['consensus'] == 'LONG'])}"
        ws_summary['A6'] = f"Short Signals: {len([r for r in results if r['consensus'] == 'SHORT'])}"
        
        # Signals Sheet
        ws_signals = wb.create_sheet("Signals")
        headers = ['Symbol', 'Signal', 'Strength%', 'Price', 'Entry', 'TP1', 'TP2', 'SL', 'R:R', 'Risk%']
        ws_signals.append(headers)
        
        # Style headers
        for cell in ws_signals[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for result in results:
            plan = result.get('trading_plan', {})
            ws_signals.append([
                result['symbol'],
                result['consensus'],
                result['strength'],
                result['price'],
                plan.get('entry'),
                plan.get('tp1'),
                plan.get('tp2'),
                plan.get('sl'),
                plan.get('risk_reward_tp1'),
                plan.get('risk_percentage')
            ])
        
        filename = f"crypto_scanner_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename
    except Exception as e:
        st.error(f"Export error: {e}")
        return None

# ==================== MAIN APP ====================


import pandas as pd
import numpy as np
import requests
import streamlit as st
from datetime import datetime, timedelta

# ==================== FUNDING RATE ANALYSIS ====================
# @st.cache_data(ttl=300)
def get_funding_rate_binance(symbol):
     """Get current and historical funding rate from Binance"""
     try:
         url = "https://fapi.binance.com/fapi/v1/fundingRate"
         params = {'symbol': symbol, 'limit': 24}  # Last 24 funding periods (3 days)
         response = requests.get(url, params=params, timeout=10)
         response.raise_for_status()
         data = response.json()
         
         if not data:
             return None
#         
#         # Convert to DataFrame
         df_funding = pd.DataFrame(data)
         df_funding['fundingRate'] = df_funding['fundingRate'].astype(float)
         df_funding['fundingTime'] = pd.to_datetime(df_funding['fundingTime'], unit='ms')
#         
#         # Current funding rate
         latest_funding = df_funding['fundingRate'].iloc[-1]
#         
#         # Average funding rate (last 24 periods)
         avg_funding = df_funding['fundingRate'].mean()
#         
#         # Funding rate trend
         recent_avg = df_funding['fundingRate'].tail(8).mean()  # Last 8 periods
         older_avg = df_funding['fundingRate'].head(8).mean()   # First 8 periods
         funding_trend = 'INCREASING' if recent_avg > older_avg else 'DECREASING'
#         
         return {
             'current': latest_funding,
             'average': avg_funding,
             'trend': funding_trend,
             'annualized': latest_funding * 365 * 3,  # Annualized rate (3 times per day)
             'history': df_funding
         }
     except Exception as e:
         return None
 
 # @st.cache_data(ttl=300)
def get_funding_rate_bybit(symbol):
     """Get funding rate from Bybit"""
     try:
         url = "https://api.bybit.com/v5/market/funding/history"
         params = {
             'category': 'linear',
             'symbol': symbol,
             'limit': 50
         }
         response = requests.get(url, params=params, timeout=10)
         response.raise_for_status()
         data = response.json()
         
         if data.get('retCode') != 0:
             return None
         
         funding_list = data['result']['list']
         if not funding_list:
             return None
         
         rates = [float(item['fundingRate']) for item in funding_list]
         latest_funding = rates[0]
         avg_funding = sum(rates) / len(rates)
         
         return {
             'current': latest_funding,
             'average': avg_funding,
             'annualized': latest_funding * 365 * 3
         }
     except Exception as e:
         return None
 
# Removed deprecated commented function

def calculate_order_flow(df):
    """
    Calculate Cumulative Volume Delta (CVD)
    Tracks buying vs selling pressure
    """
    if df is None or len(df) < 20:
        return None
    
    df = df.copy()
    
    # For Binance data with taker buy volume
    if 'taker_buy_base' in df.columns:
        df['buy_volume'] = df['taker_buy_base'].astype(float)
        df['sell_volume'] = df['volume'] - df['buy_volume']
    else:
        # Estimate buy/sell volume from price movement
        df['buy_volume'] = np.where(
            df['close'] > df['open'],
            df['volume'] * 0.6,  # Assume 60% buying on green candles
            df['volume'] * 0.4   # 40% buying on red candles
        )
        df['sell_volume'] = df['volume'] - df['buy_volume']
    
    # Calculate delta (buy volume - sell volume)
    df['delta'] = df['buy_volume'] - df['sell_volume']
    
    # Cumulative Volume Delta
    df['cvd'] = df['delta'].cumsum()
    
    # Normalize CVD
    df['cvd_normalized'] = (df['cvd'] - df['cvd'].min()) / (df['cvd'].max() - df['cvd'].min())
    
    return df

def analyze_order_flow(df):
    """
    Analyze order flow for divergences and trend
    """
    if df is None or 'cvd' not in df.columns:
        return {
            'status': 'NO_DATA',
            'strength': 0,
            'signal': 'NEUTRAL',
            'reasons': []
        }
    
    try:
        # Recent periods
        recent_20 = df.tail(20)
        older_20 = df.tail(40).head(20)
        
        # Price trend
        recent_price_avg = recent_20['close'].mean()
        older_price_avg = older_20['close'].mean()
        price_trend = 'UP' if recent_price_avg > older_price_avg else 'DOWN'
        
        # CVD trend
        recent_cvd_avg = recent_20['cvd'].mean()
        older_cvd_avg = older_20['cvd'].mean()
        cvd_trend = 'UP' if recent_cvd_avg > older_cvd_avg else 'DOWN'
        
        # Buy/Sell pressure
        recent_delta = recent_20['delta'].sum()
        buying_pressure = recent_delta > 0
        
        reasons = []
        signal_strength = 0
        
        # DIVERGENCE ANALYSIS
        if price_trend == 'UP' and cvd_trend == 'DOWN':
            # Bearish divergence
            status = 'BEARISH_DIVERGENCE'
            signal = 'SHORT'
            signal_strength = 2.5
            reasons.append("🔴 Bearish Divergence: Price up but CVD down")
            reasons.append("⚠️ Buying pressure weakening - reversal likely")
            
        elif price_trend == 'DOWN' and cvd_trend == 'UP':
            # Bullish divergence
            status = 'BULLISH_DIVERGENCE'
            signal = 'LONG'
            signal_strength = 2.5
            reasons.append("🟢 Bullish Divergence: Price down but CVD up")
            reasons.append("⚠️ Selling pressure weakening - reversal likely")
            
        elif price_trend == 'UP' and cvd_trend == 'UP':
            # Confirmation
            status = 'BULLISH_CONFIRMATION'
            signal = 'LONG'
            signal_strength = 2.0
            reasons.append("🟢 Bullish Confirmation: Price and CVD rising")
            reasons.append("💪 Strong buying pressure")
            
        elif price_trend == 'DOWN' and cvd_trend == 'DOWN':
            # Confirmation
            status = 'BEARISH_CONFIRMATION'
            signal = 'SHORT'
            signal_strength = 2.0
            reasons.append("🔴 Bearish Confirmation: Price and CVD falling")
            reasons.append("💪 Strong selling pressure")
            
        else:
            status = 'NEUTRAL'
            signal = 'NEUTRAL'
            signal_strength = 0
            reasons.append("Order flow neutral")
        
        # Add delta information
        if abs(recent_delta) > df['volume'].tail(20).mean() * 0.3:
            if recent_delta > 0:
                reasons.append(f"💹 Strong buying: +{abs(recent_delta):.0f} delta")
            else:
                reasons.append(f"📉 Strong selling: -{abs(recent_delta):.0f} delta")
        
        return {
            'status': status,
            'signal': signal,
            'strength': signal_strength,
            'recent_delta': recent_delta,
            'buying_pressure': buying_pressure,
            'reasons': reasons,
            'cvd_data': df[['timestamp', 'cvd', 'delta']].tail(50)
        }
        
    except Exception as e:
        return {
            'status': 'ERROR',
            'strength': 0,
            'signal': 'NEUTRAL',
            'reasons': [f'Analysis error: {str(e)}']
        }

# ==================== INTEGRATED ANALYSIS ====================

# ========== WRAPPER ADAPTERS FOR FUNDING & OI ==========
def fetch_funding_and_oi(exchange, symbol):
    """
    Unified fetcher that calls futures_metrics_wrappers (fm) for funding and open interest
    Returns (funding_data, df_oi) where funding_data is dict-like or None and df_oi is DataFrame or None.
    """
    try:
        # Binance symbols are like BTCUSDT; OKX uses instId like BTC-USDT-SWAP
        if exchange == 'binance':
            # funding: list of dicts => convert to expected structure
            try:
                fr = fm.get_funding_rate_binance(symbol, limit=24)
                if fr:
                    import pandas as _pd
                    df_fr = _pd.DataFrame([{'fundingRate': item['fundingRate'], 'fundingTime': item['time'], 'symbol': item.get('symbol', symbol)} for item in fr])
                    df_fr['fundingTime'] = _pd.to_datetime(df_fr['fundingTime'], unit='ms')
                    funding_data = {
                        'current': df_fr['fundingRate'].iloc[-1],
                        'average': df_fr['fundingRate'].mean(),
                        'trend': 'INCREASING' if df_fr['fundingRate'].tail(8).mean() > df_fr['fundingRate'].head(8).mean() else 'DECREASING',
                        'annualized': df_fr['fundingRate'].iloc[-1] * 365 * 3,
                        'history': df_fr
                    }
                else:
                    funding_data = None
            except Exception:
                funding_data = None

            try:
                oi_list = fm.get_open_interest_binance(symbol, period='1h', limit=48)
                import pandas as _pd
                if oi_list:
                    df_oi = _pd.DataFrame([{'timestamp': item['time'], 'sumOpenInterest': item['openInterest'] if 'openInterest' in item else item.get('openInterest', item.get('openInterest', item.get('open_interest', None))), 'sumOpenInterestValue': item.get('value', None)} for item in oi_list])
                    # Try to coerce existing keys
                    if 'timestamp' in df_oi.columns:
                        df_oi['timestamp'] = _pd.to_datetime(df_oi['timestamp'], unit='ms')
                    df_oi = df_oi.replace({None: _pd.NA}).dropna(axis=1, how='all')
                else:
                    df_oi = None
            except Exception:
                df_oi = None

        elif exchange == 'bybit':
            try:
                fr = fm.get_funding_rate_bybit(symbol, limit=24)
                if fr:
                    # fm returns list of dicts with 'fundingRate' & 'time' or similar
                    import pandas as _pd
                    df_fr = _pd.DataFrame(fr)
                    if 'time' in df_fr.columns:
                        df_fr['fundingTime'] = _pd.to_datetime(df_fr['time'], unit='ms')
                    funding_data = {
                        'current': df_fr['fundingRate'].iloc[0] if len(df_fr)>0 else None,
                        'average': df_fr['fundingRate'].mean() if len(df_fr)>0 else None,
                        'annualized': (df_fr['fundingRate'].iloc[0] * 365 * 3) if len(df_fr)>0 else None
                    }
                else:
                    funding_data = None
            except Exception:
                funding_data = None

            try:
                oi = fm.get_open_interest_bybit(symbol, interval='1h', limit=48)
                import pandas as _pd
                if oi:
                    df_oi = _pd.DataFrame(oi)
                    if 'timestamp' in df_oi.columns:
                        df_oi['timestamp'] = _pd.to_datetime(df_oi['timestamp'], unit='ms')
                else:
                    df_oi = None
            except Exception:
                df_oi = None

        elif exchange == 'okx':
            # For OKX, symbol should be instId like "BTC-USDT-SWAP"
            try:
                fr = fm.get_funding_rate_okx(symbol, limit=24)
                if fr:
                    import pandas as _pd
                    df_fr = _pd.DataFrame(fr)
                    df_fr['fundingTime'] = _pd.to_datetime(df_fr['time'], unit='ms', errors='coerce') if 'time' in df_fr.columns else None
                    funding_data = {
                        'current': df_fr['fundingRate'].iloc[0] if len(df_fr)>0 else None,
                        'average': df_fr['fundingRate'].mean() if len(df_fr)>0 else None,
                        'annualized': (df_fr['fundingRate'].iloc[0] * 365 * 3) if len(df_fr)>0 else None,
                        'history': df_fr
                    }
                else:
                    funding_data = None
            except Exception:
                funding_data = None

            try:
                oi = fm.get_open_interest_okx(symbol)
                import pandas as _pd
                if oi:
                    # fm.get_open_interest_okx returns a dict
                    df_oi = _pd.DataFrame([oi])
                    df_oi['timestamp'] = _pd.to_datetime(df_oi.get('ts', None), unit='ms', errors='coerce')
                else:
                    df_oi = None
            except Exception:
                df_oi = None

        else:
            funding_data = None
            df_oi = None

        return funding_data, df_oi
    except Exception as e:
        return None, None

# ========== END WRAPPER ADAPTERS ==========


def enhanced_futures_analysis(symbol, df_price, base_signal, exchange='binance', debug=False):
    """
    Unified enhanced analysis using wrapper fetcher (fetch_funding_and_oi).
    Returns enhanced_results with keys: funding, open_interest, order_flow, final_signal, final_strength
    """
    enhanced_results = {
        'base_signal': base_signal,
        'funding': None,
        'open_interest': None,
        'order_flow': None,
        'final_signal': base_signal.get('signal', 'NEUTRAL'),
        'final_strength': base_signal.get('strength', 0),
        'total_score': base_signal.get('score', 0),
        'all_reasons': base_signal.get('reasons', []).copy()
    }

    try:
        funding_raw, oi_raw = fetch_funding_and_oi(exchange, symbol)
        if debug:
            st.text(f"[DEBUG] [{symbol}] fetch_funding_and_oi -> funding_raw type: {type(funding_raw)}, oi_raw type: {type(oi_raw)}")
    except Exception as e:
        funding_raw, oi_raw = None, None
        if debug:
            st.text(f"[DEBUG] [{symbol}] fetch error: {e}")

    # Analyze Funding
    funding_list = None
    if isinstance(funding_raw, dict) and 'history' in funding_raw:
        try:
            funding_list = funding_raw['history'].to_dict(orient='records')
        except Exception:
            funding_list = funding_raw.get('history')
    elif isinstance(funding_raw, list):
        funding_list = funding_raw

    if funding_list:
        try:
            funding_analysis = analyze_funding_rate(funding_list, base_signal.get('signal'))
            enhanced_results['funding'] = {
                'current': funding_analysis.get('signal') and (funding_list[-1].get('fundingRate') if isinstance(funding_list, list) else funding_raw.get('current')),
                'average': funding_analysis.get('strength') and (pd.DataFrame(funding_list)['fundingRate'].mean() if isinstance(funding_list, list) else funding_raw.get('average')),
                'analysis': funding_analysis
            }
            enhanced_results['total_score'] += funding_analysis.get('strength', 0)
            enhanced_results['all_reasons'].extend(funding_analysis.get('reasons', []))
            if debug:
                st.text(f"[DEBUG] [{symbol}] funding analysis -> {funding_analysis}")
        except Exception as e:
            if debug:
                st.text(f"[DEBUG] [{symbol}] funding analysis error: {e}")

    # Analyze OI
    if oi_raw is not None:
        try:
            if isinstance(oi_raw, dict):
                df_oi = pd.DataFrame([oi_raw])
            else:
                df_oi = pd.DataFrame(oi_raw)
            oi_analysis = analyze_open_interest(df_oi)
            enhanced_results['open_interest'] = {'data': df_oi, 'analysis': oi_analysis}
            enhanced_results['total_score'] += oi_analysis.get('strength', 0)
            enhanced_results['all_reasons'].extend(oi_analysis.get('reasons', []))
            if debug:
                st.text(f"[DEBUG] [{symbol}] OI analysis -> {oi_analysis}")
        except Exception as e:
            if debug:
                st.text(f"[DEBUG] [{symbol}] OI analysis error: {e}")

    # Order Flow (CVD)
    try:
        df_flow = calculate_order_flow(df_price)
        if df_flow is not None:
            flow_analysis = analyze_order_flow(df_flow)
            enhanced_results['order_flow'] = flow_analysis
            enhanced_results['total_score'] += flow_analysis.get('strength', 0)
            enhanced_results['all_reasons'].extend(flow_analysis.get('reasons', []))
            if debug:
                st.text(f"[DEBUG] [{symbol}] Order flow -> {flow_analysis.get('status')}")
    except Exception as e:
        if debug:
            st.text(f"[DEBUG] [{symbol}] order flow error: {e}")

    # Voting logic
    votes = {'LONG': 0, 'SHORT': 0, 'NEUTRAL': 0}
    base = base_signal.get('signal', 'NEUTRAL')
    votes[base] += 3
    if enhanced_results['funding'] and enhanced_results['funding'].get('analysis'):
        votes[enhanced_results['funding']['analysis'].get('signal', 'NEUTRAL')] += 2
    if enhanced_results['open_interest'] and enhanced_results['open_interest'].get('analysis'):
        votes[enhanced_results['open_interest']['analysis'].get('signal', 'NEUTRAL')] += 2
    if enhanced_results['order_flow']:
        votes[enhanced_results['order_flow'].get('signal', 'NEUTRAL')] += 2

    final = max(votes, key=votes.get)
    enhanced_results['final_signal'] = final
    enhanced_results['final_strength'] = min((enhanced_results['total_score'] / (base_signal.get('max_score', 10) + 10)) * 100, 100)

    if debug:
        st.text(f"[DEBUG] [{symbol}] votes={votes} final={final} final_strength={enhanced_results['final_strength']:.2f}")

    return enhanced_results

def create_funding_chart(funding_data):
    """Create funding rate history chart"""
    if not funding_data or 'history' not in funding_data:
        return None
    
    import plotly.graph_objects as go
    
    df_funding = funding_data['history']
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=df_funding['fundingTime'],
        y=df_funding['fundingRate'] * 100,
        mode='lines+markers',
        name='Funding Rate %',
        line=dict(color='blue', width=2),
        fill='tozeroy'
    ))
    
    # Add threshold lines
    fig.add_hline(y=0.01 * 100, line_dash="dash", line_color="red", 
                  annotation_text="Extreme Long Bias")
    fig.add_hline(y=-0.01 * 100, line_dash="dash", line_color="green", 
                  annotation_text="Extreme Short Bias")
    fig.add_hline(y=0, line_dash="solid", line_color="gray")
    
    fig.update_layout(
        title='Funding Rate History',
        xaxis_title='Time',
        yaxis_title='Funding Rate %',
        height=300,
        hovermode='x unified'
    )
    
    return fig

def create_oi_chart(df_oi):
    """Create Open Interest chart"""
    if df_oi is None:
        return None
    
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    
    oi_column = 'sumOpenInterest' if 'sumOpenInterest' in df_oi.columns else 'openInterest'
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=df_oi['timestamp'],
        y=df_oi[oi_column],
        mode='lines',
        name='Open Interest',
        line=dict(color='purple', width=2),
        fill='tozeroy'
    ))
    
    fig.update_layout(
        title='Open Interest History',
        xaxis_title='Time',
        yaxis_title='Open Interest',
        height=300,
        hovermode='x unified'
    )
    
    return fig

def create_cvd_chart(cvd_data):
    """Create CVD (Cumulative Volume Delta) chart"""
    if cvd_data is None:
        return None
    
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    
    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.05,
        row_heights=[0.7, 0.3],
        subplot_titles=('Cumulative Volume Delta', 'Delta')
    )
    
    # CVD line
    fig.add_trace(go.Scatter(
        x=cvd_data['timestamp'],
        y=cvd_data['cvd'],
        mode='lines',
        name='CVD',
        line=dict(color='orange', width=2)
    ), row=1, col=1)
    
    # Delta bars
    colors = ['green' if x > 0 else 'red' for x in cvd_data['delta']]
    fig.add_trace(go.Bar(
        x=cvd_data['timestamp'],
        y=cvd_data['delta'],
        name='Delta',
        marker_color=colors
    ), row=2, col=1)
    
    fig.update_layout(
        height=400,
        hovermode='x unified',
        showlegend=True
    )
    
    return fig

# ==================== DISPLAY HELPERS FOR STREAMLIT ====================
def display_enhanced_analysis(enhanced_results, symbol):
    """Display enhanced analysis in Streamlit"""
    
    st.markdown("### 🚀 Enhanced Futures Analysis")
    
    # Summary metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        signal_color = "🟢" if enhanced_results['final_signal'] == 'LONG' else "🔴" if enhanced_results['final_signal'] == 'SHORT' else "⚪"
        st.metric(
            "Final Signal",
            f"{signal_color} {enhanced_results['final_signal']}",
            f"{enhanced_results['final_strength']:.1f}%"
        )
    
    with col2:
        if enhanced_results['funding']:
            funding_rate = enhanced_results['funding']['current_rate'] * 100
            st.metric(
                "Funding Rate",
                f"{funding_rate:.3f}%",
                f"APR: {enhanced_results['funding']['annualized']*100:.1f}%"
            )
    
    with col3:
        if enhanced_results['open_interest']:
            oi_change = enhanced_results['open_interest']['oi_change_pct']
            st.metric(
                "OI Change",
                f"{oi_change:+.1f}%",
                enhanced_results['open_interest']['status']
            )
    
    with col4:
        if enhanced_results['order_flow']:
            flow_status = enhanced_results['order_flow']['status']
            st.metric(
                "Order Flow",
                flow_status.replace('_', ' '),
                f"Strength: {enhanced_results['order_flow']['strength']:.1f}"
            )
    
    st.markdown("---")
    
    # Detailed breakdown
    tabs = st.tabs(["📊 Summary", "💰 Funding Rate", "📈 Open Interest", "🔄 Order Flow"])
    
    with tabs[0]:
        st.markdown("#### 🎯 All Signals & Reasons")
        
        # Display all reasons
        for idx, reason in enumerate(enhanced_results['all_reasons'], 1):
            st.markdown(f"{idx}. {reason}")
        
        # Score breakdown
        st.markdown("---")
        st.markdown("#### 📊 Score Breakdown")
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"**Base Technical Score:** {enhanced_results['base_signal'].get('score', 0):.1f}")
            if enhanced_results['funding']:
                st.markdown(f"**Funding Rate Score:** +{enhanced_results['funding']['strength']:.1f}")
            if enhanced_results['open_interest']:
                st.markdown(f"**Open Interest Score:** +{enhanced_results['open_interest']['strength']:.1f}")
            if enhanced_results['order_flow']:
                st.markdown(f"**Order Flow Score:** +{enhanced_results['order_flow']['strength']:.1f}")
        
        with col2:
            st.markdown(f"**Total Score:** {enhanced_results['total_score']:.1f}")
            st.markdown(f"**Final Strength:** {enhanced_results['final_strength']:.1f}%")
            
            # Signal quality indicator
            if enhanced_results['final_strength'] >= 70:
                st.success("✅ HIGH QUALITY SIGNAL")
            elif enhanced_results['final_strength'] >= 50:
                st.info("ℹ️ MODERATE QUALITY SIGNAL")
            else:
                st.warning("⚠️ LOW QUALITY SIGNAL")
    
    with tabs[1]:
        if enhanced_results['funding']:
            funding = enhanced_results['funding']
            
            st.markdown("#### 💰 Funding Rate Analysis")
            
            # Metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Current Rate", f"{funding['current_rate']*100:.3f}%")
            with col2:
                st.metric("Average Rate", f"{funding.get('average', 0)*100:.3f}%")
            with col3:
                st.metric("Annualized", f"{funding['annualized']*100:.1f}%")
            
            # Interpretation
            st.markdown("##### 📖 What This Means:")
            for reason in funding['reasons']:
                st.markdown(f"- {reason}")
            
            st.markdown("---")
            st.markdown("##### 📚 Funding Rate Guide:")
            st.info("""
            **Positive Funding (Longs pay Shorts):**
            - 0-0.01%: Normal bullish sentiment
            - 0.01-0.05%: High bullish sentiment (caution)
            - >0.05%: Extreme bullish - reversal likely
            
            **Negative Funding (Shorts pay Longs):**
            - 0 to -0.01%: Normal bearish sentiment
            - -0.01 to -0.05%: High bearish sentiment (caution)
            - <-0.05%: Extreme bearish - reversal likely
            
            💡 **Trading Tip:** Extreme funding often precedes sharp reversals!
            """)
            
            # Chart if data available
            if 'history' in funding:
                from datetime import datetime
                st.markdown("##### 📈 Funding Rate History")
                chart = create_funding_chart(funding)
                if chart:
                    st.plotly_chart(chart, use_container_width=True)
        else:
            st.warning("⚠️ Funding rate data not available for this symbol/exchange")
    
    with tabs[2]:
        if enhanced_results['open_interest']:
            oi = enhanced_results['open_interest']
            
            st.markdown("#### 📈 Open Interest Analysis")
            
            # Status display
            status_color = "🟢" if "UPTREND" in oi['status'] or "BULLISH" in oi['status'] else \
                          "🔴" if "DOWNTREND" in oi['status'] or "BEARISH" in oi['status'] else "🟡"
            
            st.markdown(f"### {status_color} Status: **{oi['status'].replace('_', ' ')}**")
            
            # Metrics
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("OI Change", f"{oi['oi_change_pct']:+.2f}%")
            with col2:
                st.metric("Price Change", f"{oi['price_change_pct']:+.2f}%")
            with col3:
                st.metric("Trend", oi['oi_trend'])
            
            # Interpretation
            st.markdown("---")
            st.markdown("##### 📖 What This Means:")
            for reason in oi['reasons']:
                st.markdown(f"- {reason}")
            
            st.markdown("---")
            st.markdown("##### 📚 Open Interest Guide:")
            st.info("""
            **4 Key Scenarios:**
            
            1. 🟢 **Rising OI + Rising Price** = Strong Uptrend
               - New longs entering the market
               - Bullish momentum building
            
            2. 🔴 **Rising OI + Falling Price** = Strong Downtrend
               - New shorts entering the market
               - Bearish momentum building
            
            3. 🟡 **Falling OI + Rising Price** = Weak Rally
               - Short covering (forced buying)
               - May not sustain without new buyers
            
            4. 🟡 **Falling OI + Falling Price** = Weak Selloff
               - Long liquidation (forced selling)
               - May find support soon
            
            💡 **Trading Tip:** Rising OI confirms trend strength!
            """)
            
            # Chart if data available
            # Note: You need to pass df_oi to display function or store it
            st.markdown("##### 📊 Open Interest Chart")
            st.info("Chart will be displayed when OI data is available")
        else:
            st.warning("⚠️ Open Interest data not available for this symbol/exchange")
    
    with tabs[3]:
        if enhanced_results['order_flow']:
            flow = enhanced_results['order_flow']
            
            st.markdown("#### 🔄 Order Flow Analysis (CVD)")
            
            # Status
            status_color = "🟢" if "BULLISH" in flow['status'] else \
                          "🔴" if "BEARISH" in flow['status'] else "⚪"
            
            st.markdown(f"### {status_color} Status: **{flow['status'].replace('_', ' ')}**")
            
            # Metrics
            col1, col2 = st.columns(2)
            with col1:
                delta = flow.get('recent_delta', 0)
                delta_color = "🟢" if delta > 0 else "🔴"
                st.metric("Recent Delta", f"{delta_color} {delta:,.0f}")
            
            with col2:
                pressure = "BUYING" if flow.get('buying_pressure', False) else "SELLING"
                pressure_color = "🟢" if pressure == "BUYING" else "🔴"
                st.metric("Pressure", f"{pressure_color} {pressure}")
            
            # Interpretation
            st.markdown("---")
            st.markdown("##### 📖 What This Means:")
            for reason in flow['reasons']:
                st.markdown(f"- {reason}")
            
            st.markdown("---")
            st.markdown("##### 📚 Order Flow Guide:")
            st.info("""
            **Cumulative Volume Delta (CVD):**
            - Tracks buying vs selling volume over time
            - Rising CVD = More buying than selling
            - Falling CVD = More selling than buying
            
            **Key Patterns:**
            
            1. 🔴 **Bearish Divergence:**
               - Price making higher highs
               - CVD making lower highs
               - ⚠️ Buying pressure weakening - reversal coming
            
            2. 🟢 **Bullish Divergence:**
               - Price making lower lows
               - CVD making higher lows
               - ⚠️ Selling pressure weakening - reversal coming
            
            3. ✅ **Confirmation:**
               - Price and CVD moving together
               - Strong directional conviction
            
            💡 **Trading Tip:** Divergences are early reversal warnings!
            """)
            
            # CVD Chart
            if 'cvd_data' in flow:
                st.markdown("##### 📈 CVD Chart")
                chart = create_cvd_chart(flow['cvd_data'])
                if chart:
                    st.plotly_chart(chart, use_container_width=True)
        else:
            st.warning("⚠️ Order flow data not available")

# ==================== USAGE EXAMPLE ====================

# ==================== INTEGRATION WRAPPER ====================

# Plot charts by default (expanded)
try:
    funding_raw = enhanced.get('funding_raw') or enhanced.get('funding') or None
    oi_raw = enhanced.get('open_interest_raw') or enhanced.get('open_interest') or None
    # some structures store analysis under 'analysis' key; try to extract raw lists
    if isinstance(funding_raw, dict) and 'history' in funding_raw:
        funding_raw = funding_raw['history']
    # funding chart
    fig_fr = plot_funding_chart(funding_raw, symbol)
    if fig_fr:
        st.plotly_chart(fig_fr, use_container_width=True)
    # OI chart
    fig_oi = plot_open_interest_chart(oi_raw, symbol)
    if fig_oi:
        st.plotly_chart(fig_oi, use_container_width=True)
    # funding flip
    flip = detect_funding_flip(funding_raw)
    if flip:
        st.markdown(f"**⚠️ Funding Flip:** {flip['desc']}")
except Exception as _e:
    if 'show_enhanced_debug' in globals() and show_enhanced_debug:
        st.text(f"[DEBUG] plot error: {_e}")

def integrate_with_existing_scanner(symbol, df_price, base_timeframes_analysis, exchange='binance'):
    """
    Wrapper function to integrate with your existing multi-timeframe scanner
    
    Args:
        symbol: Trading pair (e.g., 'BTCUSDT')
        df_price: Price dataframe with OHLCV
        base_timeframes_analysis: Your existing timeframe analysis results
        exchange: 'binance' or 'bybit'
    
    Returns:
        Enhanced analysis with funding, OI, and order flow
    """
    
    # Get consensus from base analysis
    base_consensus = base_timeframes_analysis.get('consensus', 'NEUTRAL')
    base_strength = base_timeframes_analysis.get('avg_strength', 0)
    
    # Create base signal format
    base_signal = {
        'signal': base_consensus,
        'strength': base_strength,
        'score': base_strength / 10,  # Convert % to score
        'max_score': 10,
        'reasons': []
    }
    
    # Add reasons from timeframes
    for tf in ['12h', '4h', '1h']:
        tf_data = base_timeframes_analysis.get('timeframes', {}).get(tf, {})
        if tf_data.get('reasons'):
            base_signal['reasons'].extend(tf_data['reasons'][:2])
    
    # Run enhanced analysis
    enhanced = enhanced_futures_analysis(symbol, df_price, base_signal, exchange)
    
    # Merge results back
    base_timeframes_analysis['enhanced'] = enhanced
    base_timeframes_analysis['final_signal'] = enhanced['final_signal']
    base_timeframes_analysis['final_strength'] = enhanced['final_strength']
    
    return base_timeframes_analysis

#=========================

@st.cache_data(ttl=60)
def get_coinglass_liquidations(symbol, timeframe='1h', limit=None):
    """
    Fetch liquidation data from Coinglass (Free API)
    Best alternative for liquidation data
    """
    try:
        # Remove USDT suffix
        base_symbol = symbol.replace('USDT', '').replace('BUSD', '')
        
        url = "https://open-api.coinglass.com/public/v2/liquidation_history"
        params = {
            'symbol': base_symbol,
            'time_type': timeframe,  # '1h', '4h', '12h', '24h'
        }
        
        response = requests.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
        
        if data.get('success') and data.get('data'):
            liquidations = []
            
            for item in data['data']:
                liquidations.append({
                    'time': pd.to_datetime(item.get('createTime', 0), unit='ms'),
                    'buy_vol': float(item.get('buyVolUsd', 0)),  # Short liquidations
                    'sell_vol': float(item.get('sellVolUsd', 0)),  # Long liquidations
                    'total_vol': float(item.get('volUsd', 0))
                })
            
            df = pd.DataFrame(liquidations)
            
            # Add side labels for compatibility
            df['side'] = df.apply(lambda x: 'BUY' if x['buy_vol'] > x['sell_vol'] else 'SELL', axis=1)
            df['value_usd'] = df['total_vol']
            df['side_label'] = df['side'].map({
                'BUY': 'Short Liquidation',
                'SELL': 'Long Liquidation'
            })
            
            return df
        
        return None
    
    except Exception as e:
        st.warning(f"Coinglass API error: {e}")
        return None


@st.cache_data(ttl=60)
def get_coinglass_liquidation_heatmap(symbol):
    """
    Get liquidation heatmap from Coinglass
    Shows liquidation levels and clustering
    """
    try:
        base_symbol = symbol.replace('USDT', '')
        
        url = "https://open-api.coinglass.com/public/v2/liquidation_chart"
        params = {
            'symbol': base_symbol,
            'time_type': '24h'
        }
        
        response = requests.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
        
        if data.get('success') and data.get('data'):
            # Parse price levels and liquidation volumes
            levels = []
            for item in data['data']:
                levels.append({
                    'price': float(item.get('price', 0)),
                    'liq_volume': float(item.get('volUsd', 0)),
                    'side': item.get('side', 'unknown')
                })
            
            return pd.DataFrame(levels)
        
        return None
    
    except Exception as e:
        st.warning(f"Heatmap error: {e}")
        return None

# ==================== LONG/SHORT RATIO FUNCTIONS ====================

@st.cache_data(ttl=60)
def get_binance_long_short_ratio(symbol, period='5m', limit=100):
    """
    Fetch Long/Short Ratio from Binance
    Shows market sentiment
    """
    try:
        url = "https://fapi.binance.com/futures/data/globalLongShortAccountRatio"
        params = {
            'symbol': symbol,
            'period': period,
            'limit': limit
        }
        
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if not data:
            return None
        
        df = pd.DataFrame(data)
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        df['longShortRatio'] = df['longShortRatio'].astype(float)
        df['longAccount'] = df['longAccount'].astype(float)
        df['shortAccount'] = df['shortAccount'].astype(float)
        
        return df
    
    except Exception as e:
        st.error(f"Error fetching Long/Short ratio: {e}")
        return None

@st.cache_data(ttl=60)
def get_binance_taker_buysell_volume(symbol, period='5m', limit=100):
    """
    Fetch Taker Buy/Sell Volume Ratio
    Shows actual trading activity
    """
    try:
        url = "https://fapi.binance.com/futures/data/takerlongshortRatio"
        params = {
            'symbol': symbol,
            'period': period,
            'limit': limit
        }
        
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        
        if not data:
            return None
        
        df = pd.DataFrame(data)
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        df['buySellRatio'] = df['buySellRatio'].astype(float)
        df['buyVol'] = df['buyVol'].astype(float)
        df['sellVol'] = df['sellVol'].astype(float)
        
        return df
    
    except Exception as e:
        st.error(f"Error fetching Taker volume: {e}")
        return None

# ==================== ORDER BOOK FUNCTIONS ====================

@st.cache_data(ttl=30)
def get_orderbook(symbol, exchange='binance', limit=1000):
    """
    Fetch order book with depth
    """
    try:
        if exchange == 'binance':
            url = "https://fapi.binance.com/fapi/v1/depth"
            params = {
                'symbol': symbol,
                'limit': limit
            }
            
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            # Parse bids and asks
            bids = pd.DataFrame(data['bids'], columns=['price', 'quantity'])
            asks = pd.DataFrame(data['asks'], columns=['price', 'quantity'])
            
            bids['price'] = bids['price'].astype(float)
            bids['quantity'] = bids['quantity'].astype(float)
            bids['total'] = bids['price'] * bids['quantity']
            bids['side'] = 'BID'
            
            asks['price'] = asks['price'].astype(float)
            asks['quantity'] = asks['quantity'].astype(float)
            asks['total'] = asks['price'] * asks['quantity']
            asks['side'] = 'ASK'
            
            return bids, asks
        
        elif exchange == 'bybit':
            url = "https://api.bybit.com/v5/market/orderbook"
            params = {
                'category': 'linear',
                'symbol': symbol,
                'limit': limit
            }
            
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()
            
            if data.get('retCode') == 0:
                result = data['result']
                
                bids = pd.DataFrame(result['b'], columns=['price', 'quantity'])
                asks = pd.DataFrame(result['a'], columns=['price', 'quantity'])
                
                bids['price'] = bids['price'].astype(float)
                bids['quantity'] = bids['quantity'].astype(float)
                bids['total'] = bids['price'] * bids['quantity']
                bids['side'] = 'BID'
                
                asks['price'] = asks['price'].astype(float)
                asks['quantity'] = asks['quantity'].astype(float)
                asks['total'] = asks['price'] * asks['quantity']
                asks['side'] = 'ASK'
                
                return bids, asks
        
        return None, None
    
    except Exception as e:
        st.error(f"Error fetching order book: {e}")
        return None, None

def analyze_whale_orders(bids, asks, threshold_percentile=95):
    """
    Identify whale orders (large orders)
    """
    if bids is None or asks is None:
        return None, None
    
    # Calculate thresholds for whale orders
    bid_threshold = bids['total'].quantile(threshold_percentile / 100)
    ask_threshold = asks['total'].quantile(threshold_percentile / 100)
    
    whale_bids = bids[bids['total'] >= bid_threshold].copy()
    whale_asks = asks[asks['total'] >= ask_threshold].copy()
    
    whale_bids['type'] = 'WHALE BID'
    whale_asks['type'] = 'WHALE ASK'
    
    return whale_bids, whale_asks

# ==================== VISUALIZATION FUNCTIONS ====================

def create_liquidation_chart(df_liq):
    """
    Create liquidation visualization
    """
    if df_liq is None or len(df_liq) == 0:
        return None
    
    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.05,
        subplot_titles=('Liquidation Volume Over Time', 'Cumulative Liquidation'),
        row_heights=[0.6, 0.4]
    )
    
    # Separate long and short liquidations
    long_liqs = df_liq[df_liq['side'] == 'SELL']  # SELL = Long Liquidation
    short_liqs = df_liq[df_liq['side'] == 'BUY']   # BUY = Short Liquidation
    
    # Row 1: Bar chart of liquidations
    fig.add_trace(
        go.Bar(
            x=long_liqs['time'],
            y=long_liqs['value_usd'],
            name='Long Liquidations',
            marker_color='red',
            opacity=0.7
        ),
        row=1, col=1
    )
    
    fig.add_trace(
        go.Bar(
            x=short_liqs['time'],
            y=short_liqs['value_usd'],
            name='Short Liquidations',
            marker_color='green',
            opacity=0.7
        ),
        row=1, col=1
    )
    
    # Row 2: Cumulative liquidations
    df_cumulative = df_liq.sort_values('time').copy()
    df_cumulative['long_cumsum'] = df_cumulative[df_cumulative['side'] == 'SELL']['value_usd'].cumsum()
    df_cumulative['short_cumsum'] = df_cumulative[df_cumulative['side'] == 'BUY']['value_usd'].cumsum()
    
    fig.add_trace(
        go.Scatter(
            x=df_cumulative['time'],
            y=df_cumulative['long_cumsum'],
            name='Cumulative Longs',
            line=dict(color='red', width=2),
            fill='tozeroy'
        ),
        row=2, col=1
    )
    
    fig.add_trace(
        go.Scatter(
            x=df_cumulative['time'],
            y=df_cumulative['short_cumsum'],
            name='Cumulative Shorts',
            line=dict(color='green', width=2),
            fill='tozeroy'
        ),
        row=2, col=1
    )
    
    fig.update_layout(
        height=600,
        hovermode='x unified',
        showlegend=True
    )
    
    fig.update_xaxes(title_text="Time", row=2, col=1)
    fig.update_yaxes(title_text="USD Value", row=1, col=1)
    fig.update_yaxes(title_text="Cumulative USD", row=2, col=1)
    
    return fig

def create_long_short_chart(df_ratio, df_volume):
    """
    Create Long/Short ratio and volume chart
    """
    fig = make_subplots(
        rows=2, cols=1,
        shared_xaxes=True,
        vertical_spacing=0.08,
        subplot_titles=('Long/Short Account Ratio', 'Taker Buy/Sell Volume Ratio'),
        row_heights=[0.5, 0.5]
    )
    
    # Row 1: Long/Short Account Ratio
    if df_ratio is not None and len(df_ratio) > 0:
        fig.add_trace(
            go.Scatter(
                x=df_ratio['timestamp'],
                y=df_ratio['longShortRatio'],
                mode='lines+markers',
                name='L/S Ratio',
                line=dict(color='blue', width=2),
                fill='tozeroy'
            ),
            row=1, col=1
        )
        
        # Add reference line at 1.0
        fig.add_hline(
            y=1.0,
            line_dash="dash",
            line_color="gray",
            annotation_text="Neutral",
            row=1, col=1
        )
    
    # Row 2: Taker Buy/Sell Ratio
    if df_volume is not None and len(df_volume) > 0:
        fig.add_trace(
            go.Scatter(
                x=df_volume['timestamp'],
                y=df_volume['buySellRatio'],
                mode='lines+markers',
                name='Buy/Sell Ratio',
                line=dict(color='orange', width=2),
                fill='tozeroy'
            ),
            row=2, col=1
        )
        
        # Add reference line at 1.0
        fig.add_hline(
            y=1.0,
            line_dash="dash",
            line_color="gray",
            annotation_text="Neutral",
            row=2, col=1
        )
    
    fig.update_layout(
        height=500,
        hovermode='x unified',
        showlegend=True
    )
    
    fig.update_xaxes(title_text="Time", row=2, col=1)
    fig.update_yaxes(title_text="Ratio", row=1, col=1)
    fig.update_yaxes(title_text="Ratio", row=2, col=1)
    
    return fig

def create_orderbook_chart(bids, asks, whale_bids, whale_asks):
    """
    Create order book visualization with whale orders highlighted
    """
    if bids is None or asks is None:
        return None
    
    fig = go.Figure()
    
    # Calculate cumulative sums
    bids_sorted = bids.sort_values('price', ascending=False).copy()
    asks_sorted = asks.sort_values('price', ascending=True).copy()
    
    bids_sorted['cumulative'] = bids_sorted['quantity'].cumsum()
    asks_sorted['cumulative'] = asks_sorted['quantity'].cumsum()
    
    # Plot regular orders
    fig.add_trace(
        go.Scatter(
            x=bids_sorted['price'],
            y=bids_sorted['cumulative'],
            name='Bids (Buy Orders)',
            fill='tozeroy',
            line=dict(color='green', width=2),
            mode='lines'
        )
    )
    
    fig.add_trace(
        go.Scatter(
            x=asks_sorted['price'],
            y=asks_sorted['cumulative'],
            name='Asks (Sell Orders)',
            fill='tozeroy',
            line=dict(color='red', width=2),
            mode='lines'
        )
    )
    
    # Highlight whale orders
    if whale_bids is not None and len(whale_bids) > 0:
        fig.add_trace(
            go.Scatter(
                x=whale_bids['price'],
                y=whale_bids['quantity'],
                name='Whale Bids',
                mode='markers',
                marker=dict(
                    size=15,
                    color='darkgreen',
                    symbol='diamond',
                    line=dict(color='white', width=2)
                )
            )
        )
    
    if whale_asks is not None and len(whale_asks) > 0:
        fig.add_trace(
            go.Scatter(
                x=whale_asks['price'],
                y=whale_asks['quantity'],
                name='Whale Asks',
                mode='markers',
                marker=dict(
                    size=15,
                    color='darkred',
                    symbol='diamond',
                    line=dict(color='white', width=2)
                )
            )
        )
    
    # Calculate spread
    best_bid = bids['price'].max()
    best_ask = asks['price'].min()
    mid_price = (best_bid + best_ask) / 2
    
    fig.add_vline(
        x=mid_price,
        line_dash="dash",
        line_color="yellow",
        annotation_text=f"Mid: ${mid_price:,.2f}"
    )
    
    fig.update_layout(
        title='Order Book Depth with Whale Orders',
        xaxis_title='Price (USD)',
        yaxis_title='Cumulative Quantity',
        height=500,
        hovermode='x unified'
    )
    
    return fig
# ===== Enhanced Futures Analysis integrated =====

def main():
    # Initialize database
    if not st.session_state.db_initialized:
        if init_database():
            st.session_state.db_initialized = True
    
    # Header
    st.markdown('<p class="main-header">🚀 Crypto Futures Scanner Development (v5) </p>', unsafe_allow_html=True)
    
    # Top bar
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        st.markdown(f"**🕐 Time:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    with col2:
        st.markdown(f"**🔄 Last Scan:** {st.session_state.last_refresh.strftime('%H:%M:%S')}")
    with col3:
        if st.button("🔄 Refresh", use_container_width=True):
            st.session_state.last_refresh = datetime.now()
            st.cache_data.clear()
            st.rerun()
    
    st.markdown("---")
    
    # Create tabs
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["📊 Scanner", "💼 Portfolio", "📈 Performance", "⚙️ Settings", "📊Single Analyzer", "📊 Live Price", "Sentiment Economic"])
    
    # ==================== TAB 1: SCANNER ====================
    with tab1:
        
# ==================== DEBUG TOGGLE ====================
        with st.sidebar:
            show_enhanced_debug = st.checkbox("🧩 Show Enhanced Debug Log", value=False)
            """
            Crypto Futures Scanner Pro - Ultimate Edition v5.0
            Features:

                🔄 Multi-Exchange Support (Binance, Bybit, Gate.io)
                ⚡ Parallel Processing for Fast Scanning
                📊 Advanced Technical Analysis (15+ Indicators)
                💼 Portfolio Tracking
                📱 Telegram Alerts
                📈 Performance Analytics
                💾 Database Storage
                📥 Excel Export
                ⚠️ Disclaimer: For educational purposes only. Always DYOR and manage risk properly.

            """
            def simple_trading_sessions():
                """Versi sederhana trading session monitor"""
                
                st.header("🕒 Trading Sessions (WIB Time)")
        
                # Current time
                wib = pytz.timezone('Asia/Jakarta')
                current_time = datetime.now(wib)
                current_hour = current_time.hour
                
                st.metric("Current Time (WIB)", current_time.strftime("%H:%M:%S"))
                
                # CORRECTED session times
                sessions = [
                    {"name": "Sydney", "open": 5, "close": 14, "active": False},
                    {"name": "Tokyo", "open": 7, "close": 16, "active": False},  # CORRECTED ⚡
                    {"name": "London", "open": 14, "close": 23, "active": False},
                    {"name": "New York", "open": 19, "close": 4, "active": False}
                ]
                
                # Check active sessions
                for session in sessions:
                    if session['name'] == 'New York':  # Cross midnight
                        session['active'] = current_hour >= session['open'] or current_hour < session['close']
                    else:  # Normal sessions
                        session['active'] = session['open'] <= current_hour < session['close']
                
                # Display sessions
                cols = st.columns(4)
                for idx, session in enumerate(sessions):
                    with cols[idx]:
                        if session['active']:
                            st.success(f"🟢 {session['name']}")
                            st.write(f"⏰ {session['open']:02d}:00-{session['close']:02d}:00")
                            st.write("**ACTIVE**")
                        else:
                            st.error(f"🔴 {session['name']}")
                            st.write(f"⏰ {session['open']:02d}:00-{session['close']:02d}:00")
                            st.write("CLOSED")
                
                # Market status based on corrected times
                st.markdown("---")
                active_count = sum(1 for s in sessions if s['active'])
                if active_count >= 2:
                    st.success("🚀 HIGH VOLATILITY - Multiple sessions active")
                elif active_count == 1:
                    st.info("📈 MARKET OPEN - One session active")
                else:
                    st.warning("💤 MARKETS CLOSED - No active sessions")

        # Gunakan yang simple
        simple_trading_sessions()

        #=================================#

        st.header("⚙️ Scanner Settings") 
        st.subheader("🔡 Data Source")
        api_source = st.selectbox(
            "Exchange:",
            options=['binance', 'bybit', 'gateio'],
            format_func=lambda x: API_SOURCES[x]['name'],
            help="Auto fallback enabled"
        )
            
        scan_limit = st.slider("Coins to Scan", 10, 50, 25, 5)
            
        st.markdown("---")
        st.subheader("🎯 Filters")
        min_strength = st.slider("Min Strength (%)", 0, 100, 40, 5)
        show_only_signals = st.checkbox("Show Only Signals", value=True)
        min_volume = st.number_input("Min Volume Ratio", 0.0, 5.0, 1.2, 0.1)
            
        signal_types = st.multiselect(
            "Signal Types",
            options=['LONG', 'SHORT', 'NEUTRAL'],
            default=['LONG', 'SHORT']
        )
            
        min_rr = st.slider("Min Risk/Reward", 1.0, 5.0, 2.0, 0.5)
            
        st.markdown("---")
        st.subheader("🚀 Performance")
        parallel_workers = st.slider("Parallel Workers", 1, 10, 5)
        st.info(f"⚡ Scanning with {parallel_workers} threads")
        
        # Initialize source variable in session state if not exists
        if 'current_exchange' not in st.session_state:
            st.session_state.current_exchange = api_source
        
        # Start scan
        if st.button("🎯 Start Scan", type="primary", use_container_width=True):
            scan_start_time = time.time()
            
            with st.spinner(f"🔍 Fetching top {scan_limit} coins..."):
                source, symbols = get_top_symbols(api_source, scan_limit)
            
            if not symbols:
                st.error("❌ Failed to fetch data from all exchanges")
                        
            # Store current exchange in session state
            st.session_state.current_exchange = source if source else api_source
            
            if source != api_source:
                st.warning(f"⚠️ {API_SOURCES[api_source]['name']} unavailable. Using {API_SOURCES[source]['name']}")
            else:
                st.success(f"✅ Connected to {API_SOURCES[source]['name']}")
            
            # Parallel scanning
            with st.spinner(f"⚡ Analyzing {scan_limit} coins with {parallel_workers} workers..."):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                results = parallel_scan(symbols, source, max_workers=parallel_workers)
                
                for i in range(len(results)):
                    progress = (i + 1) / len(results)
                    progress_bar.progress(progress)
                    status_text.text(f"Analyzed {i+1}/{len(results)} coins")
                
                progress_bar.empty()
                status_text.empty()
            
            scan_duration = time.time() - scan_start_time
            
            # Filter results
            filtered_results = []
            for r in results:
                if show_only_signals and r['consensus'] not in signal_types:
                    continue
                if r['strength'] < min_strength:
                    continue
                
                # Check volume
                volume_ok = False
                for tf in ['1h', '4h', '12h']:
                    tf_data = r['timeframes'].get(tf, {})
                    indicators = tf_data.get('indicators', {})
                    if indicators.get('volume_ratio', 0) >= min_volume:
                        volume_ok = True
                        break
                if not volume_ok:
                    continue
                
                # Check risk/reward
                plan = r.get('trading_plan')
                if plan and plan.get('risk_reward_tp1', 0) < min_rr:
                    continue
                
                filtered_results.append(r)
            
            # Sort by strength
            filtered_results = sorted(filtered_results, key=lambda x: x['strength'], reverse=True)
            
            # Save to session state
            st.session_state.scan_results = filtered_results
            
            # Save to database
            avg_strength = sum([r['strength'] for r in filtered_results]) / len(filtered_results) if filtered_results else 0
            save_scan_history(len(results), len(filtered_results), avg_strength, source, scan_duration)
            
            # Display summary
            st.success(f"✅ Scan completed in {scan_duration:.1f}s")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Scanned", len(results))
            with col2:
                st.metric("Signals Found", len(filtered_results))
            with col3:
                long_count = len([r for r in filtered_results if r['consensus'] == 'LONG'])
                st.metric("Long Signals", long_count)
            with col4:
                short_count = len([r for r in filtered_results if r['consensus'] == 'SHORT'])
                st.metric("Short Signals", short_count)
        
        # Display results
        if st.session_state.scan_results:
            results = st.session_state.scan_results
            
            # Get current exchange from session state
            current_exchange = st.session_state.get('current_exchange', 'binance')
            
            st.markdown(f"### 📊 Found {len(results)} Trading Opportunities")
            
            # Market Heatmap
            with st.expander("🗺️ Market Heatmap", expanded=False):
                try:
                    heatmap = create_market_heatmap_safe(results)
                    
                    if heatmap:
                        st.plotly_chart(heatmap, use_container_width=True)
                    else:
                        # Fallback 1: Grid view
                        st.info("📊 Using grid view (heatmap unavailable)")
                        grid_result = create_market_grid(results)
                        
                        if not grid_result:
                            # Fallback 2: Simple table
                            st.info("📋 Using table view")
                            create_simple_signal_table(results)
                
                except Exception as e:
                    st.error(f"⚠️ Visualization error: {e}")
                    
                    # Emergency fallback
                    try:
                        st.info("📊 Showing alternative view...")
                        create_market_grid(results)
                    except:
                        st.info("📋 Showing simple table...")
                        create_simple_signal_table(results)


            # Export button
            col1, col2 = st.columns([1, 5])
            with col1:
                if st.button("📥 Export Excel"):
                    filename = export_to_excel(results)
                    if filename:
                        st.success(f"✅ Exported to {filename}")
            
            st.markdown("---")
            
            # Display each signal
            for result in results:
                signal_emoji = "🟢" if result['consensus'] == 'LONG' else "🔴" if result['consensus'] == 'SHORT' else "⚪"
                
                with st.expander(
                    f"{signal_emoji} **{result['name']}** ({result['symbol']}) | "
                    f"Signal: **{result['consensus']}** | Strength: **{result['strength']}%**",
                    expanded=False
                ):
                    col1, col2, col3 = st.columns([2, 2, 3])
                    
                    with col1:
                        st.markdown(f"### ${result['price']:,.8f}")
                        signal_class = f"signal-{result['consensus'].lower()}"
                        st.markdown(f'<div class="{signal_class}">🎯 {result["consensus"]}</div>', unsafe_allow_html=True)
                        st.metric("Strength", f"{result['strength']}%")
                    
                    with col2:
                        st.markdown("**📊 Multi-Timeframe:**")
                        for tf in ['12h', '4h', '1h']:
                            tf_data = result['timeframes'].get(tf, {})
                            signal = tf_data.get('signal', 'N/A')
                            strength = tf_data.get('strength', 0)
                            emoji = "🟢" if signal == "LONG" else "🔴" if signal == "SHORT" else "⚪"
                            st.markdown(f"{emoji} **{tf.upper()}:** {signal} ({strength:.1f}%)")
                    
                    with col3:
                        if result['trading_plan']:
                            plan = result['trading_plan']
                            st.markdown("**💼 Trading Plan:**")
                            if plan.get('entry_note'):
                                st.warning(plan['entry_note'])
                            st.markdown(f"📍 **Entry:** ${plan['entry']:,.8f}")
                            st.markdown(f"🎯 **TP1:** ${plan['tp1']:,.8f} (R:R **{plan['risk_reward_tp1']}**)")
                            st.markdown(f"🎯 **TP2:** ${plan['tp2']:,.8f} (R:R **{plan['risk_reward_tp2']}**)")
                            st.markdown(f"🛑 **SL:** ${plan['sl']:,.8f} (Risk: {plan['risk_percentage']:.2f}%)")
                            
                            if plan['support'] and plan['resistance']:
                                st.markdown("---")
                                st.markdown(f"📉 **Support:** ${plan['support']:,.8f}")
                                st.markdown(f"📈 **Resistance:** ${plan['resistance']:,.8f}")
                            
                            # Save to DB button - Use current_exchange variable
                            if st.button(f"💾 Save to Portfolio", key=f"save_{result['symbol']}"):
                                signal_id = save_signal_to_db(result, current_exchange)
                                if signal_id:
                                    st.success(f"✅ Saved {result['symbol']} to portfolio!")
                                    
                                    # Send Telegram alert if enabled
                                    if st.session_state.telegram_enabled:
                                        telegram_token = st.session_state.get('telegram_token')
                                        telegram_chat_id = st.session_state.get('telegram_chat_id')
                                        if telegram_token and telegram_chat_id:
                                            if send_telegram_alert(
                                                telegram_token,
                                                telegram_chat_id,
                                                result['symbol'],
                                                result
                                            ):
                                                st.success("📱 Telegram alert sent!")
                    
                    # Technical details
                    st.markdown("---")
                    st.markdown("**📊 Technical Indicators:**")
                    cols = st.columns(3)
                    
                    for idx, tf in enumerate(['12h', '4h', '1h']):
                        tf_data = result['timeframes'].get(tf, {})
                        indicators = tf_data.get('indicators', {})
                        reasons = tf_data.get('reasons', [])
                        patterns = tf_data.get('patterns', [])
                        
                        if tf_data.get('signal') not in ['ERROR', 'INSUFFICIENT_DATA']:
                            with cols[idx]:
                                st.markdown(f"**⏰ {tf.upper()}**")
                                
                                rsi = indicators.get('rsi', 0)
                                rsi_emoji = "🟢" if rsi < 30 else "🔴" if rsi > 70 else "🟡" if rsi < 40 else "🟠" if rsi > 60 else "⚪"
                                st.markdown(f"{rsi_emoji} RSI: **{rsi:.1f}**")
                                
                                stoch = indicators.get('stoch_rsi', 0)
                                stoch_emoji = "🟢" if stoch < 20 else "🔴" if stoch > 80 else "⚪"
                                st.markdown(f"{stoch_emoji} Stoch: **{stoch:.1f}**")
                                
                                macd = indicators.get('macd_diff', 0)
                                macd_emoji = "🟢" if macd > 0 else "🔴"
                                st.markdown(f"{macd_emoji} MACD: **{macd:.8f}**")
                                
                                vol = indicators.get('volume_ratio', 0)
                                vol_emoji = "🟢" if vol > 1.5 else "🟡" if vol > 1.2 else "⚪"
                                st.markdown(f"{vol_emoji} Volume: **{vol:.2f}x**")
                                
                                adx = indicators.get('adx', 0)
                                adx_emoji = "💪" if adx > 25 else "⚪"
                                st.markdown(f"{adx_emoji} ADX: **{adx:.1f}**")
                                
                                ema_trend = indicators.get('ema_trend', 'NEUTRAL')
                                trend_emoji = "📈" if ema_trend == 'BULL' else "📉" if ema_trend == 'BEAR' else "➡️"
                                st.markdown(f"{trend_emoji} Trend: **{ema_trend}**")
                                
                                if patterns:
                                    st.markdown("**🎭 Patterns:**")
                                    for pattern in patterns[:2]:
                                        if 'GOLDEN' in pattern:
                                            st.markdown("⭐ Golden Cross")
                                        elif 'DEATH' in pattern:
                                            st.markdown("💀 Death Cross")
                                        elif 'BULLISH' in pattern:
                                            st.markdown("📈 Bull Div")
                                        elif 'BEARISH' in pattern:
                                            st.markdown("📉 Bear Div")
                                
                                if reasons:
                                    st.markdown("**💡 Signals:**")
                                    for reason in reasons[:3]:
                                        st.markdown(f"• {reason}")
                    
                    # Chart
                    main_df = result['timeframes'].get('4h_df')
                    if main_df is not None and len(main_df) >= 50:
                        st.markdown("---")
                        st.markdown("**📈 Chart Analysis (4H)**")
                        chart = create_chart(main_df, result['name'], result['trading_plan'])
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
        
                        # ========== NEW: ENHANCED ANALYSIS DISPLAY ==========
                        st.markdown("---")
                        if result.get('enhanced'):
                            try:
                                display_enhanced_analysis(result['enhanced'], result['symbol'])
                            except Exception as e:
                                st.warning(f"Enhanced analysis display error: {e}")
                        else:
                            st.info("💡 Enhanced futures analysis not available for this coin")
        else:
            st.info("👆 Click 'Start Scan' to begin analyzing the market")
    
    # ==================== TAB 2: PORTFOLIO ====================
    with tab2:
        display_portfolio_tracker()
        
        st.markdown("---")
        st.markdown("### 📜 Recent Signals")
        
        try:
            conn = sqlite3.connect(DB_NAME)
            recent_signals = pd.read_sql('''
                SELECT 
                    timestamp,
                    symbol,
                    signal,
                    strength,
                    entry,
                    tp1,
                    tp2,
                    sl,
                    status
                FROM signals
                ORDER BY timestamp DESC
                LIMIT 20
            ''', conn)
            conn.close()
            
            if len(recent_signals) > 0:
                st.dataframe(recent_signals, use_container_width=True, hide_index=True)
            else:
                st.info("No signals history yet")
        except Exception as e:
            st.error(f"Error loading signals: {e}")
    
    # ==================== TAB 3: PERFORMANCE ====================
    with tab3:
        st.markdown("### 📈 Performance Analytics")
        
        # Performance chart
        perf_chart = create_performance_chart()
        if perf_chart:
            st.plotly_chart(perf_chart, use_container_width=True)
        else:
            st.info("No performance data yet. Complete some trades to see analytics.")
        
        # Statistics
        try:
            conn = sqlite3.connect(DB_NAME)
            
            # Overall stats
            stats = pd.read_sql('''
                SELECT 
                    COUNT(*) as total_trades,
                    SUM(CASE WHEN pnl_percentage > 0 THEN 1 ELSE 0 END) as winning_trades,
                    SUM(CASE WHEN pnl_percentage < 0 THEN 1 ELSE 0 END) as losing_trades,
                    AVG(pnl_percentage) as avg_pnl,
                    SUM(pnl_percentage) as total_pnl,
                    MAX(pnl_percentage) as best_trade,
                    MIN(pnl_percentage) as worst_trade
                FROM performance
            ''', conn)
            
            if len(stats) > 0 and stats['total_trades'].iloc[0] > 0:
                st.markdown("#### 📊 Overall Statistics")
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Trades", int(stats['total_trades'].iloc[0]))
                
                with col2:
                    win_rate = (stats['winning_trades'].iloc[0] / stats['total_trades'].iloc[0]) * 100
                    st.metric("Win Rate", f"{win_rate:.1f}%")
                
                with col3:
                    st.metric("Total PnL", f"{stats['total_pnl'].iloc[0]:+.2f}%")
                
                with col4:
                    st.metric("Avg PnL", f"{stats['avg_pnl'].iloc[0]:+.2f}%")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Best Trade", f"{stats['best_trade'].iloc[0]:+.2f}%", delta="🏆")
                with col2:
                    st.metric("Worst Trade", f"{stats['worst_trade'].iloc[0]:+.2f}%", delta="📉")
            
            # Scan history
            st.markdown("---")
            st.markdown("#### 📜 Scan History")
            
            scan_history = pd.read_sql('''
                SELECT 
                    timestamp,
                    total_scanned,
                    signals_found,
                    avg_strength,
                    exchange,
                    ROUND(scan_duration, 2) as duration_sec
                FROM scan_history
                ORDER BY timestamp DESC
                LIMIT 10
            ''', conn)
            
            conn.close()
            
            if len(scan_history) > 0:
                st.dataframe(scan_history, use_container_width=True, hide_index=True)
            
        except Exception as e:
            st.error(f"Error loading performance data: {e}")
    
    # ==================== TAB 4: SETTINGS ====================
    with tab4:
        st.markdown("### ⚙️ Application Settings")
        
        # Telegram Settings
        st.markdown("#### 📱 Telegram Alerts")
        
        col1, col2 = st.columns([1, 3])
        with col1:
            telegram_enabled = st.checkbox("Enable Telegram", value=st.session_state.telegram_enabled)
        
        if telegram_enabled:
            with col2:
                telegram_token = st.text_input("Bot Token", type="password", placeholder="123456:ABC-DEF1234ghIkl-zyx57W2v1u123ew11")
                telegram_chat_id = st.text_input("Chat ID", placeholder="-1001234567890")
                
                if st.button("🧪 Test Connection"):
                    if telegram_token and telegram_chat_id:
                        test_msg = f"🧪 Test Alert\n\nConnection successful!\n⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        url = f"https://api.telegram.org/bot{telegram_token}/sendMessage"
                        data = {'chat_id': telegram_chat_id, 'text': test_msg}
                        
                        try:
                            response = requests.post(url, data=data, timeout=10)
                            if response.status_code == 200:
                                st.success("✅ Telegram connection successful!")
                                st.session_state.telegram_enabled = True
                                st.session_state.telegram_token = telegram_token
                                st.session_state.telegram_chat_id = telegram_chat_id
                            else:
                                st.error("❌ Connection failed. Check your token and chat ID.")
                        except Exception as e:
                            st.error(f"❌ Error: {e}")
                    else:
                        st.warning("⚠️ Please enter both Bot Token and Chat ID")
        
        st.markdown("---")
        
        # Auto Refresh
        st.markdown("#### 🔄 Auto Refresh")
        refresh_option = st.radio(
            "Interval:",
            options=["Disabled", "Every 5 Minutes", "Every 15 Minutes", "Every 1 Hour"],
            index=0
        )
        
        if refresh_option == "Every 5 Minutes":
            st.session_state.refresh_interval = 5
        elif refresh_option == "Every 15 Minutes":
            st.session_state.refresh_interval = 15
        elif refresh_option == "Every 1 Hour":
            st.session_state.refresh_interval = 60
        else:
            st.session_state.refresh_interval = None
        
        if st.session_state.refresh_interval:
            time_until = st.session_state.refresh_interval - ((datetime.now() - st.session_state.last_refresh).total_seconds() / 60)
            if time_until > 0:
                st.info(f"⏱️ Next refresh in: {time_until:.1f} minutes")
        
        st.markdown("---")
        
        # Database Management
        st.markdown("#### 🗄️ Database Management")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🗑️ Clear Scan History"):
                try:
                    conn = sqlite3.connect(DB_NAME)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM scan_history")
                    conn.commit()
                    conn.close()
                    st.success("✅ Scan history cleared")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        with col2:
            if st.button("🗑️ Clear Closed Signals"):
                try:
                    conn = sqlite3.connect(DB_NAME)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM signals WHERE status = 'CLOSED'")
                    conn.commit()
                    conn.close()
                    st.success("✅ Closed signals cleared")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        with col3:
            if st.button("⚠️ Reset Database"):
                try:
                    conn = sqlite3.connect(DB_NAME)
                    cursor = conn.cursor()
                    cursor.execute("DROP TABLE IF EXISTS signals")
                    cursor.execute("DROP TABLE IF EXISTS performance")
                    cursor.execute("DROP TABLE IF EXISTS scan_history")
                    cursor.execute("DROP TABLE IF EXISTS watchlist")
                    conn.commit()
                    conn.close()
                    init_database()
                    st.success("✅ Database reset")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        st.markdown("---")
        
        # About
        st.markdown("#### ℹ️ About")
        st.info("""
        **Crypto Futures Scanner Pro - Ultimate Edition v5.0**
        
        Features:
        - 🔄 Multi-Exchange Support (Binance, Bybit, Gate.io)
        - ⚡ Parallel Processing for Fast Scanning
        - 📊 Advanced Technical Analysis (15+ Indicators)
        - 💼 Portfolio Tracking
        - 📱 Telegram Alerts
        - 📈 Performance Analytics
        - 💾 Database Storage
        - 📥 Excel Export
        
        ⚠️ **Disclaimer:** For educational purposes only. Always DYOR and manage risk properly.
        """)
    
    # ==================== MARKET VISUALIZATION FUNCTIONS (SAFE) ====================

def create_market_grid(results):
    """
    Simple grid view as fallback when heatmap fails
    """
    if not results or len(results) == 0:
        st.info("No data available")
        return None
    
    # Group by signal type
    long_signals = [r for r in results if r.get('consensus') == 'LONG']
    short_signals = [r for r in results if r.get('consensus') == 'SHORT']
    neutral_signals = [r for r in results if r.get('consensus') == 'NEUTRAL']
    
    # Sort by strength
    long_signals = sorted(long_signals, key=lambda x: x.get('strength', 0), reverse=True)
    short_signals = sorted(short_signals, key=lambda x: x.get('strength', 0), reverse=True)
    neutral_signals = sorted(neutral_signals, key=lambda x: x.get('strength', 0), reverse=True)
    
    # Display in columns
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("### 🟢 LONG Signals")
        if long_signals:
            for r in long_signals[:15]:
                symbol = r.get('name', r.get('symbol', 'N/A'))
                strength = r.get('strength', 0)
                price = r.get('price', 0)
                st.markdown(f"**{symbol}**: {strength:.1f}% | ${price:,.4f}")
        else:
            st.info("No long signals")
    
    with col2:
        st.markdown("### 🔴 SHORT Signals")
        if short_signals:
            for r in short_signals[:15]:
                symbol = r.get('name', r.get('symbol', 'N/A'))
                strength = r.get('strength', 0)
                price = r.get('price', 0)
                st.markdown(f"**{symbol}**: {strength:.1f}% | ${price:,.4f}")
        else:
            st.info("No short signals")
    
    with col3:
        st.markdown("### ⚪ NEUTRAL")
        if neutral_signals:
            for r in neutral_signals[:15]:
                symbol = r.get('name', r.get('symbol', 'N/A'))
                strength = r.get('strength', 0)
                price = r.get('price', 0)
                st.markdown(f"**{symbol}**: {strength:.1f}% | ${price:,.4f}")
        else:
            st.info("No neutral signals")
    
    return True


def create_market_heatmap_safe(results):
    """
    Ultra-safe heatmap with multiple fallback visualizations
    """
    if not results or len(results) == 0:
        st.info("No data available for heatmap")
        return None
    
    # Prepare data with extensive validation
    data = []
    for r in results:
        try:
            tf_4h = r.get('timeframes', {}).get('4h', {})
            indicators = tf_4h.get('indicators', {}) if isinstance(tf_4h, dict) else {}
            
            symbol = str(r.get('name') or r.get('symbol') or 'Unknown')
            strength = float(r.get('strength', 0) or 0)
            consensus = str(r.get('consensus', 'NEUTRAL'))
            volume = float(indicators.get('volume_ratio', 1) or 1)
            
            # Only include valid entries
            if strength > 0 and volume > 0 and symbol != 'Unknown':
                data.append({
                    'symbol': symbol,
                    'strength': strength,
                    'consensus': consensus,
                    'volume': volume
                })
        except (TypeError, ValueError, KeyError) as e:
            continue
    
    if len(data) == 0:
        st.warning("⚠️ No valid data for heatmap (all strengths are zero or invalid)")
        return None
    
    df = pd.DataFrame(data)
    
    # Data validation and cleaning
    df['strength'] = pd.to_numeric(df['strength'], errors='coerce').fillna(0)
    df['volume'] = pd.to_numeric(df['volume'], errors='coerce').fillna(1)
    
    # Remove zero values
    df = df[df['strength'] > 0].copy()
    
    if len(df) == 0:
        st.warning("⚠️ All strength values are zero - cannot create heatmap")
        return None
    
    # Ensure minimum values for visualization
    df['strength'] = df['strength'].clip(lower=0.1)
    df['volume'] = df['volume'].clip(lower=0.1)
    
    # Method 1: Try Treemap
    try:
        fig = px.treemap(
            df,
            path=['consensus', 'symbol'],
            values='strength',
            color='volume',
            color_continuous_scale='RdYlGn',
            title='Market Overview Heatmap',
            hover_data={
                'strength': ':.1f',
                'volume': ':.2f'
            }
        )
        fig.update_layout(
            height=600,
            margin=dict(t=50, l=25, r=25, b=25)
        )
        return fig
    except Exception as e1:
        st.warning(f"Treemap failed, trying alternative... ({str(e1)[:50]})")
    
    # Method 2: Try Sunburst
    try:
        fig = px.sunburst(
            df,
            path=['consensus', 'symbol'],
            values='strength',
            color='volume',
            color_continuous_scale='RdYlGn',
            title='Market Overview (Sunburst Chart)'
        )
        fig.update_layout(height=600)
        return fig
    except Exception as e2:
        st.warning(f"Sunburst failed, using bar chart... ({str(e2)[:50]})")
    
    # Method 3: Fallback to Bar Chart
    try:
        df_sorted = df.sort_values('strength', ascending=False).head(30)
        
        fig = go.Figure()
        
        colors = {
            'LONG': 'green',
            'SHORT': 'red',
            'NEUTRAL': 'gray'
        }
        
        for signal in df_sorted['consensus'].unique():
            df_signal = df_sorted[df_sorted['consensus'] == signal]
            
            fig.add_trace(go.Bar(
                y=df_signal['symbol'],
                x=df_signal['strength'],
                name=signal,
                orientation='h',
                marker_color=colors.get(signal, 'blue'),
                text=df_signal['strength'].round(1),
                textposition='auto',
                hovertemplate='<b>%{y}</b><br>Strength: %{x:.1f}%<extra></extra>'
            ))
        
        fig.update_layout(
            title='Market Overview - Top 30 by Strength',
            xaxis_title='Strength %',
            yaxis_title='Symbol',
            height=max(400, len(df_sorted) * 20),  # Dynamic height
            barmode='group',
            showlegend=True,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="right",
                x=1
            )
        )
        
        return fig
    except Exception as e3:
        st.error(f"❌ All visualization methods failed: {e3}")
        return None


def create_simple_signal_table(results):
    """
    Ultra-simple table view as last resort
    """
    if not results:
        return None
    
    table_data = []
    for r in results:
        try:
            table_data.append({
                'Symbol': r.get('name', r.get('symbol', 'N/A')),
                'Signal': r.get('consensus', 'N/A'),
                'Strength': f"{r.get('strength', 0):.1f}%",
                'Price': f"${r.get('price', 0):,.4f}"
            })
        except:
            continue
    
    if table_data:
        df = pd.DataFrame(table_data)
        st.dataframe(
            df.sort_values('Strength', ascending=False),
            use_container_width=True,
            hide_index=True
        )
        return True
    
    return None


# ==================== REPLACE create_market_heatmap ====================

def create_market_heatmap(results):
    """
    DEPRECATED: Use create_market_heatmap_safe instead
    This wrapper ensures backward compatibility
    """
    return create_market_heatmap_safe(results)

    def main():
        tab1, tab2, tab3, tab4, tab5 = st.tabs(['Scanner', 'Portfolio', 'Performance', 'Settings', 'Single Analyzer'])
    
    # ================= TAB 5 Single Analyzer ==============================
    with tab5:
        st.header("🧩 Single Coin Analyzer (Fast Mode)")
        st.caption("Analyze one specific coin instantly with the same logic as the scanner.")

        exchange_single = st.selectbox("Select Exchange", ["binance", "bybit", "gateio"], index=0, key="single_exchange")

        # Default symbols
        default_symbols = ["BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT", "XRPUSDT", "ADAUSDT", "DOGEUSDT", "MATICUSDT", "DOTUSDT", "AVAXUSDT"]
        
        # Get symbols from last scan if available
        scan_symbols = []
        if st.session_state.get("scan_results"):
            scan_symbols = sorted(set([r["symbol"] for r in st.session_state.get("scan_results", [])]))
        
        all_symbols = scan_symbols if scan_symbols else default_symbols

        # Symbol input with autocomplete
        selected_symbol = st.selectbox(
            "Select Coin",
            all_symbols,
            key="single_symbol_select"
        )
        
        # Or manual input
        manual_symbol = st.text_input("Or enter symbol manually (e.g., BTCUSDT)", key="manual_symbol_input")
        
        # Use manual input if provided
        symbol_to_analyze = manual_symbol.upper() if manual_symbol else selected_symbol

        if st.button("🔍 Analyze Coin", type="primary", use_container_width=True):
            if not symbol_to_analyze:
                st.warning("⚠️ Please select or enter a coin symbol first.")
                st.stop()

            try:
                with st.spinner(f"Fetching {symbol_to_analyze} data from {exchange_single}..."):
                    # --- Fetch current price ---
                    current_price = 0
                    symbol_data = None
                    
                    try:
                        if exchange_single == "binance":
                            # Get price
                            resp = requests.get(
                                f"https://api.binance.com/api/v3/ticker/price?symbol={symbol_to_analyze}", 
                                timeout=5
                            )
                            resp.raise_for_status()
                            current_price = float(resp.json().get("price", 0))
                            
                            # Create symbol_data in format expected by analyze_single_coin
                            symbol_data = {
                                'symbol': symbol_to_analyze,
                                'lastPrice': str(current_price)
                            }
                            
                        elif exchange_single == "bybit":
                            resp = requests.get(
                                f"https://api.bybit.com/v5/market/tickers?category=linear&symbol={symbol_to_analyze}", 
                                timeout=5
                            )
                            resp.raise_for_status()
                            data = resp.json()
                            if data.get("retCode") == 0 and data.get("result", {}).get("list"):
                                current_price = float(data["result"]["list"][0].get("lastPrice", 0))
                                symbol_data = {
                                    'symbol': symbol_to_analyze,
                                    'lastPrice': str(current_price)
                                }
                        
                        elif exchange_single == "gateio":
                            contract = symbol_to_analyze.replace('USDT', '_USDT')
                            resp = requests.get(
                                f"https://api.gateio.ws/api/v4/futures/usdt/contracts/{contract}",
                                timeout=5
                            )
                            resp.raise_for_status()
                            data = resp.json()
                            current_price = float(data.get("last_price", 0))
                            symbol_data = {
                                'name': contract,
                                'last_price': str(current_price)
                            }
                            
                    except Exception as e:
                        st.error(f"❌ Failed to fetch price from {exchange_single}: {e}")
                        st.stop()
                    
                    if not symbol_data or current_price == 0:
                        st.error(f"❌ Could not fetch data for {symbol_to_analyze} on {exchange_single}")
                        st.stop()
                    
                    st.success(f"✅ Current Price: ${current_price:,.8f}")

                # --- Run Analysis ---
                with st.spinner(f"⚡ Running multi-timeframe analysis..."):
                    # Use analyze_single_coin function
                    analysis = analyze_single_coin(exchange_single, symbol_data)
                    
                    if not analysis:
                        st.error("❌ Analysis failed. Please try again.")
                        st.stop()

                # --- Display Results ---
                st.markdown("---")
                st.subheader(f"📊 Analysis Results for {symbol_to_analyze}")
                
                # Main metrics
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    consensus = analysis.get('consensus', 'NEUTRAL')
                    signal_emoji = "🟢" if consensus == 'LONG' else "🔴" if consensus == 'SHORT' else "⚪"
                    st.metric("Signal", f"{signal_emoji} {consensus}")
                
                with col2:
                    strength = analysis.get('strength', 0)
                    st.metric("Strength", f"{strength:.1f}%")
                
                with col3:
                    st.metric("Price", f"${current_price:,.8f}")
                
                with col4:
                    # Count confirming timeframes
                    confirming_tfs = 0
                    for tf in ['1h', '4h', '12h']:
                        tf_signal = analysis.get('timeframes', {}).get(tf, {}).get('signal')
                        if tf_signal == consensus:
                            confirming_tfs += 1
                    st.metric("TF Confirm", f"{confirming_tfs}/3")
                
                # Timeframe breakdown
                st.markdown("---")
                st.markdown("#### ⏱️ Multi-Timeframe Analysis")
                
                tf_cols = st.columns(3)
                for idx, tf in enumerate(['12h', '4h', '1h']):
                    tf_data = analysis.get('timeframes', {}).get(tf, {})
                    
                    with tf_cols[idx]:
                        st.markdown(f"**{tf.upper()} Timeframe**")
                        
                        signal = tf_data.get('signal', 'N/A')
                        strength = tf_data.get('strength', 0)
                        
                        emoji = "🟢" if signal == "LONG" else "🔴" if signal == "SHORT" else "⚪"
                        st.markdown(f"{emoji} **{signal}** ({strength:.1f}%)")
                        
                        # Show top indicators
                        indicators = tf_data.get('indicators', {})
                        if indicators:
                            st.caption("**Key Indicators:**")
                            st.caption(f"RSI: {indicators.get('rsi', 0):.1f}")
                            st.caption(f"Volume: {indicators.get('volume_ratio', 0):.2f}x")
                            st.caption(f"ADX: {indicators.get('adx', 0):.1f}")
                        
                        # Show reasons
                        reasons = tf_data.get('reasons', [])
                        if reasons:
                            st.caption("**Signals:**")
                            for reason in reasons[:3]:
                                st.caption(f"• {reason}")
                
                # Trading Plan
                if analysis.get('trading_plan'):
                    st.markdown("---")
                    st.markdown("#### 💼 Trading Plan")
                    
                    plan = analysis['trading_plan']
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown(f"**📍 Entry:** ${plan['entry']:,.8f}")
                        st.markdown(f"**🎯 TP1:** ${plan['tp1']:,.8f} (R:R {plan['risk_reward_tp1']})")
                        st.markdown(f"**🎯 TP2:** ${plan['tp2']:,.8f} (R:R {plan['risk_reward_tp2']})")
                        st.markdown(f"**🛑 Stop Loss:** ${plan['sl']:,.8f}")
                    
                    with col2:
                        st.markdown(f"**⚡ Risk:** {plan['risk_percentage']:.2f}%")
                        if plan.get('support'):
                            st.markdown(f"**📉 Support:** ${plan['support']:,.8f}")
                        if plan.get('resistance'):
                            st.markdown(f"**📈 Resistance:** ${plan['resistance']:,.8f}")
                        if plan.get('entry_note'):
                            st.warning(plan['entry_note'])
                
                # Enhanced Analysis
                st.markdown("---")
                st.markdown("#### 🚀 Enhanced Futures Analysis")
                
                # Run enhanced analysis
                main_df = analysis.get('timeframes', {}).get('4h_df')
                if main_df is not None:
                    with st.spinner("Running enhanced analysis..."):
                        # Create base signal for enhanced analysis
                        base_signal = {
                            'signal': analysis.get('consensus', 'NEUTRAL'),
                            'strength': analysis.get('strength', 0),
                            'score': analysis.get('strength', 0) / 10,
                            'max_score': 10,
                            'reasons': []
                        }
                        
                        # Add reasons from 4h timeframe
                        tf_4h = analysis.get('timeframes', {}).get('4h', {})
                        if tf_4h.get('reasons'):
                            base_signal['reasons'].extend(tf_4h['reasons'][:3])
                        
                        # Run enhanced analysis
                        enhanced = enhanced_futures_analysis(
                            symbol_to_analyze, 
                            main_df, 
                            base_signal, 
                            exchange_single,
                            debug=show_enhanced_debug
                        )
                        
                        # Display enhanced results
                        if enhanced:
                            display_enhanced_analysis(enhanced, symbol_to_analyze)
                
                # Chart
                if main_df is not None and len(main_df) >= 50:
                    st.markdown("---")
                    st.markdown("#### 📈 Price Chart (4H)")
                    chart = create_chart(main_df, symbol_to_analyze, analysis.get('trading_plan'))
                    if chart:
                        st.plotly_chart(chart, use_container_width=True)
                
                # Save to portfolio option
                st.markdown("---")
                col1, col2 = st.columns([1, 3])
                with col1:
                    if st.button("💾 Save to Portfolio", use_container_width=True):
                        signal_id = save_signal_to_db(analysis, exchange_single)
                        if signal_id:
                            st.success(f"✅ Saved {symbol_to_analyze} to portfolio!")
                            
                            # Send Telegram if enabled
                            if st.session_state.telegram_enabled:
                                telegram_token = st.session_state.get('telegram_token')
                                telegram_chat_id = st.session_state.get('telegram_chat_id')
                                if telegram_token and telegram_chat_id:
                                    if send_telegram_alert(telegram_token, telegram_chat_id, symbol_to_analyze, analysis):
                                        st.success("📱 Telegram alert sent!")

            except requests.exceptions.Timeout:
                st.error("⏱️ Request timeout. Please try again.")
            except requests.exceptions.RequestException as e:
                st.error(f"❌ Network error: {e}")
            except Exception as e:
                st.error(f"❌ Unexpected error: {e}")
                if show_enhanced_debug:
                    st.exception(e)

        st.header("📊 Advanced Market Analysis: Liquidation & Order Flow")
        
        # Symbol selection
        col1, col2 = st.columns([2, 1])
        
        with col1:
            # Get available symbols
            available_symbols = []
            if st.session_state.get("scan_results"):
                available_symbols = sorted(set([r["symbol"] for r in st.session_state.scan_results]))
            
            if not available_symbols:
                available_symbols = [
                    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT", "XRPUSDT",
                    "ADAUSDT", "DOGEUSDT", "MATICUSDT", "DOTUSDT", "AVAXUSDT"
                ]
            
            selected_symbol_liq = st.selectbox(
                "Select Trading Pair",
                available_symbols,
                key="liq_symbol_select"
            )
        
        with col2:
            manual_symbol_liq = st.text_input(
                "Or enter manually",
                placeholder="BTCUSDT",
                key="liq_manual_symbol"
            )
        
        # Use manual input if provided
        symbol_liq = (manual_symbol_liq.upper() if manual_symbol_liq else selected_symbol_liq)
        
        # Exchange selection
        exchange_liq = st.selectbox(
            "Exchange",
            ["binance", "bybit", "coinglass"],
            key="liq_exchange"
        )
        
        # Analysis button
        if st.button("🔍 Analyze Liquidation & Orders", type="primary", use_container_width=True):
            
            with st.spinner(f"Fetching data for {symbol_liq}..."):
                
                # Create tabs for different analyses
                tab_liq, tab_ratio, tab_orderbook = st.tabs([
                    "💥 Liquidations",
                    "📊 Long/Short Ratio",
                    "📖 Order Book & Whales"
                ])
                
                # ===== TAB 1: LIQUIDATIONS =====
                with tab_liq:
                    st.subheader(f"💥 Liquidation Analysis - {symbol_liq}")
                    
                    # Fetch liquidation data
                    df_liq = get_liquidation_data(symbol_liq, exchange="auto", debug=True)
                    
                    if df_liq is not None and len(df_liq) > 0:
                        # Summary metrics
                        col1, col2, col3, col4 = st.columns(4)
                        
                        total_liq = df_liq['value_usd'].sum()
                        long_liq = df_liq[df_liq['side'] == 'SELL']['value_usd'].sum()
                        short_liq = df_liq[df_liq['side'] == 'BUY']['value_usd'].sum()
                        
                        with col1:
                            st.metric("Total Liquidations", f"${total_liq:,.0f}")
                        
                        with col2:
                            st.metric("Long Liquidations", f"${long_liq:,.0f}", 
                                     delta=f"{(long_liq/total_liq*100):.1f}%")
                        
                        with col3:
                            st.metric("Short Liquidations", f"${short_liq:,.0f}",
                                     delta=f"{(short_liq/total_liq*100):.1f}%")
                        
                        with col4:
                            ratio = long_liq / short_liq if short_liq > 0 else 0
                            st.metric("Long/Short Ratio", f"{ratio:.2f}")
                        
                        # Chart
                        st.markdown("---")
                        liq_chart = create_liquidation_chart(df_liq)
                        if liq_chart:
                            st.plotly_chart(liq_chart, use_container_width=True)
                        
                        # Recent liquidations table
                        st.markdown("---")
                        st.subheader("📋 Recent Liquidations")
                        
                        df_display = df_liq.sort_values('time', ascending=False).head(20).copy()
                        df_display['time'] = df_display['time'].dt.strftime('%Y-%m-%d %H:%M:%S')
                        df_display = df_display[['time', 'side_label', 'price', 'executedQty', 'value_usd']]
                        df_display.columns = ['Time', 'Type', 'Price', 'Quantity', 'Value (USD)']
                        
                        st.dataframe(df_display, use_container_width=True, hide_index=True)
                        
                        # Interpretation
                        st.markdown("---")
                        st.info(f"""
                        ### 📖 Liquidation Analysis Interpretation:
                        
                        **Current Status:**
                        - Long Liquidations: **${long_liq:,.0f}** ({(long_liq/total_liq*100):.1f}%)
                        - Short Liquidations: **${short_liq:,.0f}** ({(short_liq/total_liq*100):.1f}%)
                        
                        **What This Means:**
                        - 🔴 **High Long Liquidations**: Price dropped, longs got liquidated → Bearish pressure
                        - 🟢 **High Short Liquidations**: Price pumped, shorts got liquidated → Bullish pressure
                        - ⚖️ **Balanced Liquidations**: Market is ranging, no clear direction
                        
                        **Trading Tips:**
                        - Watch for liquidation cascades (rapid increase)
                        - High liquidations often mark local tops/bottoms
                        - Use as confirmation with technical analysis
                        """)
                    
                    else:
                        st.warning("No liquidation data available for this symbol")
                
                # ===== TAB 2: LONG/SHORT RATIO =====
                with tab_ratio:
                    st.subheader(f"📊 Long/Short Ratio Analysis - {symbol_liq}")
                    
                    # Fetch ratio data
                    df_ratio = get_binance_long_short_ratio(symbol_liq, period='5m', limit=100)
                    df_volume = get_binance_taker_buysell_volume(symbol_liq, period='5m', limit=100)
                    
                    if df_ratio is not None or df_volume is not None:
                        # Summary metrics
                        col1, col2, col3 = st.columns(3)
                        
                        if df_ratio is not None and len(df_ratio) > 0:
                            current_ratio = df_ratio['longShortRatio'].iloc[-1]
                            long_pct = df_ratio['longAccount'].iloc[-1] * 100
                            short_pct = df_ratio['shortAccount'].iloc[-1] * 100
                            
                            with col1:
                                st.metric("L/S Account Ratio", f"{current_ratio:.2f}")
                            
                            with col2:
                                st.metric("Long Accounts", f"{long_pct:.1f}%")
                            
                            with col3:
                                st.metric("Short Accounts", f"{short_pct:.1f}%")
                        
                        # Chart
                        st.markdown("---")
                        ratio_chart = create_long_short_chart(df_ratio, df_volume)
                        if ratio_chart:
                            st.plotly_chart(ratio_chart, use_container_width=True)
                        
                        # Interpretation
                        st.markdown("---")
                        if df_ratio is not None and len(df_ratio) > 0:
                            if current_ratio > 1.5:
                                sentiment = "🟢 **EXTREMELY BULLISH**"
                                warning = "⚠️ Warning: Overcrowded long positions may lead to long squeeze!"
                            elif current_ratio > 1.1:
                                sentiment = "🟢 **BULLISH**"
                                warning = "ℹ️ More longs than shorts - bullish sentiment"
                            elif current_ratio > 0.9:
                                sentiment = "⚪ **NEUTRAL**"
                                warning = "ℹ️ Balanced market - no clear bias"
                            elif current_ratio > 0.7:
                                sentiment = "🔴 **BEARISH**"
                                warning = "ℹ️ More shorts than longs - bearish sentiment"
                            else:
                                sentiment = "🔴 **EXTREMELY BEARISH**"
                                warning = "⚠️ Warning: Overcrowded short positions may lead to short squeeze!"
                            
                            st.info(f"""
                            ### 📖 Long/Short Ratio Interpretation:
                            
                            **Current Sentiment:** {sentiment}
                            
                            **Ratio:** {current_ratio:.2f}
                            - Ratio > 1: More long positions than shorts
                            - Ratio < 1: More short positions than longs
                            - Ratio = 1: Balanced market
                            
                            {warning}
                            
                            **Trading Tips:**
                            - Extreme ratios (>2 or <0.5) often precede reversals
                            - Use as contrarian indicator when combined with price action
                            - Watch for divergences between ratio and price
                            """)
                    
                    else:
                        st.warning("Long/Short ratio data not available")
                
                # ===== TAB 3: ORDER BOOK & WHALES =====
                with tab_orderbook:
                    st.subheader(f"📖 Order Book Analysis - {symbol_liq}")
                    
                    # Fetch order book
                    bids, asks = get_orderbook(symbol_liq, exchange=exchange_liq, limit=1000)
                    
                    if bids is not None and asks is not None:
                        # Analyze whale orders
                        whale_threshold = st.slider(
                            "Whale Order Threshold (Percentile)",
                            90, 99, 95,
                            help="Orders above this percentile are considered 'whale orders'"
                        )
                        
                        whale_bids, whale_asks = analyze_whale_orders(bids, asks, whale_threshold)
                        
                        # Summary metrics
                        col1, col2, col3, col4 = st.columns(4)
                        
                        best_bid = bids['price'].max()
                        best_ask = asks['price'].min()
                        spread = best_ask - best_bid
                        spread_pct = (spread / best_bid) * 100
                        
                        total_bid_value = bids['total'].sum()
                        total_ask_value = asks['total'].sum()
                        
                        with col1:
                            st.metric("Best Bid", f"${best_bid:,.2f}")
                        
                        with col2:
                            st.metric("Best Ask", f"${best_ask:,.2f}")
                        
                        with col3:
                            st.metric("Spread", f"${spread:.2f}", delta=f"{spread_pct:.3f}%")
                        
                        with col4:
                            bid_ask_ratio = total_bid_value / total_ask_value if total_ask_value > 0 else 0
                            st.metric("Bid/Ask Ratio", f"{bid_ask_ratio:.2f}")
                        
                        # Order book chart
                        st.markdown("---")
                        ob_chart = create_orderbook_chart(bids, asks, whale_bids, whale_asks)
                        if ob_chart:
                            st.plotly_chart(ob_chart, use_container_width=True)
                        
                        # Whale orders table
                        st.markdown("---")
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.subheader("🐋 Whale Bids (Buy Walls)")
                            if whale_bids is not None and len(whale_bids) > 0:
                                whale_bids_display = whale_bids.sort_values('total', ascending=False).head(10).copy()
                                whale_bids_display = whale_bids_display[['price', 'quantity', 'total']]
                                whale_bids_display.columns = ['Price', 'Quantity', 'Value (USD)']
                                st.dataframe(whale_bids_display, use_container_width=True, hide_index=True)
                            else:
                                st.info("No whale bids detected at current threshold")
                        
                        with col2:
                            st.subheader("🐋 Whale Asks (Sell Walls)")
                            if whale_asks is not None and len(whale_asks) > 0:
                                whale_asks_display = whale_asks.sort_values('total', ascending=False).head(10).copy()
                                whale_asks_display = whale_asks_display[['price', 'quantity', 'total']]
                                whale_asks_display.columns = ['Price', 'Quantity', 'Value (USD)']
                                st.dataframe(whale_asks_display, use_container_width=True, hide_index=True)
                            else:
                                st.info("No whale asks detected at current threshold")
                        
                        # Order book statistics
                        st.markdown("---")
                        st.subheader("📊 Order Book Statistics")
                        
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.metric("Total Bid Volume", f"{bids['quantity'].sum():,.2f}")
                            st.metric("Total Bid Value", f"${total_bid_value:,.0f}")
                            st.metric("Avg Bid Size", f"{bids['quantity'].mean():,.2f}")
                        
                        with col2:
                            st.metric("Total Ask Volume", f"{asks['quantity'].sum():,.2f}")
                            st.metric("Total Ask Value", f"${total_ask_value:,.0f}")
                            st.metric("Avg Ask Size", f"{asks['quantity'].mean():,.2f}")
                        
                        with col3:
                            if whale_bids is not None:
                                st.metric("Whale Bids", len(whale_bids))
                            if whale_asks is not None:
                                st.metric("Whale Asks", len(whale_asks))
                            st.metric("Bid/Ask Imbalance", f"{((total_bid_value - total_ask_value) / total_bid_value * 100):+.1f}%")
                        
                        # Interpretation
                        st.markdown("---")
                        st.info(f"""
                        ### 📖 Order Book Interpretation:
                        
                        **Bid/Ask Ratio: {bid_ask_ratio:.2f}**
                        - Ratio > 1.5: Strong buying pressure (bullish)
                        - Ratio 0.8-1.2: Balanced market (neutral)
                        - Ratio < 0.8: Strong selling pressure (bearish)
                        
                        **Whale Orders Detected:**
                        - 🐋 Whale Bids: {len(whale_bids) if whale_bids is not None else 0} (Support levels)
                        - 🐋 Whale Asks: {len(whale_asks) if whale_asks is not None else 0} (Resistance levels)
                        
                        **Trading Tips:**
                        - Large bid walls can act as support (but may be fake/spoofed)
                        - Large ask walls can act as resistance (but may be fake/spoofed)
                        - Watch for sudden appearance/disappearance of whale orders
                        - Spread < 0.1% = Very liquid market
                        - Spread > 0.5% = Low liquidity, higher slippage risk
                        
                        **Current Spread: {spread_pct:.3f}%**
                        """)
                    
                    else:
                        st.warning("Order book data not available")
        
        # Help section
        with st.expander("ℹ️ Understanding Liquidation & Order Analysis"):
            st.markdown("""
            ## 📚 Complete Guide
            
            ### 💥 Liquidations
            **What are liquidations?**
            - Forced closure of leveraged positions when margin is insufficient
            - Long liquidation = Price drops → Longs forced to sell → More downward pressure
            - Short liquidation = Price rises → Shorts forced to buy → More upward pressure
            
            **How to use:**
            - High liquidations often mark local tops/bottoms
            - Cascading liquidations create "liquidation wicks"
            - Use as reversal signals when extreme
            
            ---
            
            ### 📊 Long/Short Ratio
            **What is it?**
            - Shows percentage of traders with long vs short positions
            - Account Ratio: Number of accounts holding long vs short
            - Volume Ratio: Actual trading volume of buys vs sells
            
            **How to interpret:**
            - **Ratio > 2**: Overcrowded longs → Risk of long squeeze
            - **Ratio < 0.5**: Overcrowded shorts → Risk of short squeeze
            - **Ratio ≈ 1**: Balanced market
            
            **Trading strategy:**
            - Use as contrarian indicator at extremes
            - Combine with price action for confirmation
            - Watch for divergences (price up but ratio down = bearish)
            
            ---
            
            ### 📖 Order Book & Whale Analysis
            **What is order book?**
            - Real-time list of all buy (bid) and sell (ask) orders
            - Shows supply and demand at different price levels
            - Depth shows cumulative volume
            
            **Whale orders:**
            - Unusually large orders (top 5-10% by size)
            - Can be support/resistance levels
            - May be "spoofed" (fake orders to manipulate)
            
            **Key metrics:**
            - **Bid/Ask Ratio**: Shows market bias
            - **Spread**: Indicates liquidity
            - **Order imbalance**: Predicts short-term direction
            
            **How to trade:**
            1. Identify major support/resistance from whale orders
            2. Watch for order book imbalances (more bids = bullish)
            3. Be aware of order spoofing (orders that disappear)
            4. Use tight stops near low liquidity areas
            
            ---
            
            ### 🎯 Combined Strategy
            **Step 1**: Check liquidation data for recent activity
            - High long liquidations + oversold RSI = Potential bounce
            - High short liquidations + overbought RSI = Potential reversal
            
            **Step 2**: Analyze Long/Short ratio
            - Extreme ratio + liquidations = Strong reversal signal
            - Balanced ratio = Use other indicators
            
            **Step 3**: Check order book
            - Strong bid support + bullish liquidations = Long entry
            - Strong ask resistance + bearish liquidations = Short entry
            
            **Risk Management:**
            - Always use stop losses
            - Reduce leverage during high liquidation periods
            - Be cautious of fake whale walls (spoofing)
            - Never rely on single indicator alone
            
            ---
            
            ### ⚠️ Important Warnings
            - **Liquidation cascades** can create extreme volatility
            - **Whale orders** can be fake (spoofing)
            - **Ratios** are lagging indicators
            - **Always confirm** with technical analysis
            - **Use proper risk management** at all times
            """)
        
    # ==================== TAB 6: REAL-TIME PRICE MONITOR ====================
    with tab6:
        st.header("📊 Real-Time Price Monitor")
        st.caption("Monitor live cryptocurrency prices with auto-refresh capability")
        
        # Settings Section
        col1, col2 = st.columns([2, 1])
        
        with col1:
            # Exchange selection
            monitor_exchange = st.selectbox(
                "Select Exchange",
                ["binance", "bybit", "gateio"],
                format_func=lambda x: API_SOURCES[x]['name'],
                key="monitor_exchange"
            )
        
        with col2:
            # Auto-refresh toggle
            auto_refresh_enabled = st.checkbox("🔄 Auto Refresh", value=False, key="price_auto_refresh")
            if auto_refresh_enabled:
                refresh_interval = st.selectbox(
                    "Refresh Interval",
                    [5, 10, 30, 60],
                    format_func=lambda x: f"{x} seconds",
                    key="price_refresh_interval"
                )
        
        st.markdown("---")
        
        # Symbol Selection Methods
        tab_select, tab_manual = st.tabs(["📋 Select from List", "⌨️ Manual Entry"])
        
        with tab_select:
            # Get available symbols from recent scan or defaults
            available_symbols = []
            if st.session_state.get("scan_results"):
                available_symbols = sorted(set([r["symbol"] for r in st.session_state.scan_results]))
            
            if not available_symbols:
                # Default popular symbols
                available_symbols = [
                    "BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT", "XRPUSDT",
                    "ADAUSDT", "DOGEUSDT", "MATICUSDT", "DOTUSDT", "AVAXUSDT",
                    "LINKUSDT", "LTCUSDT", "UNIUSDT", "ATOMUSDT", "ETCUSDT"
                ]
            
            # Multi-select for monitoring multiple coins
            selected_symbols = st.multiselect(
                "Select Coins to Monitor",
                available_symbols,
                default=[available_symbols[0]] if available_symbols else [],
                key="monitor_symbols_select"
            )
        
        with tab_manual:
            manual_symbols_input = st.text_area(
                "Enter Symbol(s) - One per line",
                placeholder="BTCUSDT\nETHUSDT\nBNBUSDT",
                height=100,
                key="manual_symbols_input"
            )
            
            if manual_symbols_input:
                manual_symbols = [s.strip().upper() for s in manual_symbols_input.split('\n') if s.strip()]
                if manual_symbols:
                    st.info(f"✅ {len(manual_symbols)} symbol(s) ready to monitor")
                    if st.button("Use Manual Symbols", key="use_manual"):
                        selected_symbols = manual_symbols
                        st.success("Manual symbols loaded!")
        
        # Combine symbols
        if 'selected_symbols' not in locals():
            selected_symbols = []
        
        st.markdown("---")
        
        # Control Buttons
        col1, col2, col3 = st.columns(3)
        
        with col1:
            start_monitoring = st.button(
                "🚀 Start Monitoring",
                type="primary",
                use_container_width=True,
                key="start_monitor"
            )
        
        with col2:
            if st.button("🔄 Refresh Now", use_container_width=True, key="manual_refresh"):
                st.session_state.force_refresh = True
        
        with col3:
            if st.button("🗑️ Clear Data", use_container_width=True, key="clear_monitor"):
                if 'price_data' in st.session_state:
                    del st.session_state.price_data
                st.success("Data cleared!")
                st.rerun()
        
        # Initialize session state for price data
        if 'price_data' not in st.session_state:
            st.session_state.price_data = {}
        if 'last_price_update' not in st.session_state:
            st.session_state.last_price_update = datetime.now()
        if 'force_refresh' not in st.session_state:
            st.session_state.force_refresh = False
        
        # Function to fetch live prices
        def fetch_live_prices(exchange, symbols):
            """Fetch current prices for multiple symbols"""
            prices = {}
            
            for symbol in symbols:
                try:
                    if exchange == "binance":
                        url = f"https://fapi.binance.com/fapi/v1/ticker/24hr?symbol={symbol}"
                        response = requests.get(url, timeout=5)
                        response.raise_for_status()
                        data = response.json()
                        
                        prices[symbol] = {
                            'price': float(data.get('lastPrice', 0)),
                            'change_24h': float(data.get('priceChangePercent', 0)),
                            'high_24h': float(data.get('highPrice', 0)),
                            'low_24h': float(data.get('lowPrice', 0)),
                            'volume_24h': float(data.get('volume', 0)),
                            'quote_volume': float(data.get('quoteVolume', 0)),
                            'timestamp': datetime.now()
                        }
                    
                    elif exchange == "bybit":
                        url = f"https://api.bybit.com/v5/market/tickers?category=linear&symbol={symbol}"
                        response = requests.get(url, timeout=5)
                        response.raise_for_status()
                        data = response.json()
                        
                        if data.get('retCode') == 0 and data.get('result', {}).get('list'):
                            ticker = data['result']['list'][0]
                            prices[symbol] = {
                                'price': float(ticker.get('lastPrice', 0)),
                                'change_24h': float(ticker.get('price24hPcnt', 0)) * 100,
                                'high_24h': float(ticker.get('highPrice24h', 0)),
                                'low_24h': float(ticker.get('lowPrice24h', 0)),
                                'volume_24h': float(ticker.get('volume24h', 0)),
                                'quote_volume': float(ticker.get('turnover24h', 0)),
                                'timestamp': datetime.now()
                            }
                    
                    elif exchange == "gateio":
                        contract = symbol.replace('USDT', '_USDT')
                        url = f"https://api.gateio.ws/api/v4/futures/usdt/contracts/{contract}"
                        response = requests.get(url, timeout=5)
                        response.raise_for_status()
                        data = response.json()
                        
                        prices[symbol] = {
                            'price': float(data.get('last_price', 0)),
                            'change_24h': 0,  # Gate.io doesn't provide 24h change in this endpoint
                            'high_24h': 0,
                            'low_24h': 0,
                            'volume_24h': 0,
                            'quote_volume': 0,
                            'timestamp': datetime.now()
                        }
                    
                    time.sleep(0.1)  # Rate limiting
                    
                except Exception as e:
                    st.warning(f"Failed to fetch {symbol}: {e}")
                    prices[symbol] = None
            
            return prices
        
        # Monitoring Logic
        if start_monitoring or st.session_state.force_refresh:
            st.session_state.force_refresh = False
            
            if not selected_symbols:
                st.warning("⚠️ Please select at least one symbol to monitor")
            else:
                with st.spinner(f"Fetching live prices from {API_SOURCES[monitor_exchange]['name']}..."):
                    current_prices = fetch_live_prices(monitor_exchange, selected_symbols)
                    
                    # Update session state
                    for symbol, data in current_prices.items():
                        if data:
                            if symbol not in st.session_state.price_data:
                                st.session_state.price_data[symbol] = []
                            
                            st.session_state.price_data[symbol].append(data)
                            
                            # Keep only last 100 data points
                            if len(st.session_state.price_data[symbol]) > 100:
                                st.session_state.price_data[symbol] = st.session_state.price_data[symbol][-100:]
                    
                    st.session_state.last_price_update = datetime.now()
        
        # Display Price Data
        if st.session_state.price_data:
            st.markdown("---")
            st.subheader("💰 Current Prices")
            
            # Last update time
            time_since_update = (datetime.now() - st.session_state.last_price_update).total_seconds()
            st.caption(f"Last updated: {st.session_state.last_price_update.strftime('%H:%M:%S')} ({time_since_update:.0f}s ago)")
            
            # Create metrics grid
            cols_per_row = 3
            symbols_to_display = list(st.session_state.price_data.keys())
            
            for i in range(0, len(symbols_to_display), cols_per_row):
                cols = st.columns(cols_per_row)
                
                for j, col in enumerate(cols):
                    if i + j < len(symbols_to_display):
                        symbol = symbols_to_display[i + j]
                        data_list = st.session_state.price_data[symbol]
                        
                        if data_list and data_list[-1]:
                            latest = data_list[-1]
                            
                            with col:
                                # Price change indicator
                                change_24h = latest.get('change_24h', 0)
                                change_emoji = "🟢" if change_24h >= 0 else "🔴"
                                
                                st.metric(
                                    label=f"{change_emoji} {symbol}",
                                    value=f"${latest['price']:,.8f}",
                                    delta=f"{change_24h:+.2f}%"
                                )
                                
                                # Additional info in expander
                                with st.expander("📊 Details"):
                                    st.write(f"**24h High:** ${latest.get('high_24h', 0):,.8f}")
                                    st.write(f"**24h Low:** ${latest.get('low_24h', 0):,.8f}")
                                    st.write(f"**24h Volume:** {latest.get('volume_24h', 0):,.2f}")
                                    st.write(f"**Quote Volume:** ${latest.get('quote_volume', 0):,.2f}")
            
            # Price Charts
            st.markdown("---")
            st.subheader("📈 Price History")
            
            for symbol in symbols_to_display:
                data_list = st.session_state.price_data[symbol]
                
                if data_list and len(data_list) > 1:
                    with st.expander(f"📊 {symbol} Price Chart", expanded=False):
                        # Prepare data for chart
                        df_history = pd.DataFrame(data_list)
                        
                        # Create price line chart
                        fig = go.Figure()
                        
                        fig.add_trace(go.Scatter(
                            x=df_history['timestamp'],
                            y=df_history['price'],
                            mode='lines+markers',
                            name='Price',
                            line=dict(color='#1f77b4', width=2),
                            marker=dict(size=6)
                        ))
                        
                        # Add 24h high/low reference lines
                        if df_history['high_24h'].iloc[-1] > 0:
                            fig.add_hline(
                                y=df_history['high_24h'].iloc[-1],
                                line_dash="dash",
                                line_color="green",
                                annotation_text="24h High"
                            )
                        
                        if df_history['low_24h'].iloc[-1] > 0:
                            fig.add_hline(
                                y=df_history['low_24h'].iloc[-1],
                                line_dash="dash",
                                line_color="red",
                                annotation_text="24h Low"
                            )
                        
                        fig.update_layout(
                            title=f"{symbol} Price Movement",
                            xaxis_title="Time",
                            yaxis_title="Price (USD)",
                            hovermode='x unified',
                            height=400
                        )
                        
                        st.plotly_chart(fig, use_container_width=True)
                        
                        # Statistics
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            min_price = df_history['price'].min()
                            st.metric("Min Price", f"${min_price:,.8f}")
                        
                        with col2:
                            max_price = df_history['price'].max()
                            st.metric("Max Price", f"${max_price:,.8f}")
                        
                        with col3:
                            avg_price = df_history['price'].mean()
                            st.metric("Avg Price", f"${avg_price:,.8f}")
                        
                        with col4:
                            price_range = ((max_price - min_price) / min_price) * 100
                            st.metric("Range", f"{price_range:.2f}%")
            
            # Export Options
            st.markdown("---")
            st.subheader("💾 Export Data")
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("📥 Download CSV", use_container_width=True):
                    # Combine all symbol data
                    all_data = []
                    for symbol, data_list in st.session_state.price_data.items():
                        for data in data_list:
                            if data:
                                all_data.append({
                                    'symbol': symbol,
                                    'timestamp': data['timestamp'],
                                    'price': data['price'],
                                    'change_24h': data['change_24h'],
                                    'high_24h': data['high_24h'],
                                    'low_24h': data['low_24h'],
                                    'volume_24h': data['volume_24h']
                                })
                    
                    df_export = pd.DataFrame(all_data)
                    csv = df_export.to_csv(index=False)
                    
                    st.download_button(
                        label="💾 Download CSV File",
                        data=csv,
                        file_name=f"price_monitor_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv"
                    )
            
            with col2:
                if st.button("📊 Generate Report", use_container_width=True):
                    st.info("📄 Generating summary report...")
                    
                    # Create summary statistics
                    summary_data = []
                    for symbol, data_list in st.session_state.price_data.items():
                        if data_list:
                            prices = [d['price'] for d in data_list if d]
                            latest = data_list[-1]
                            
                            summary_data.append({
                                'Symbol': symbol,
                                'Current Price': f"${latest['price']:,.8f}",
                                '24h Change': f"{latest['change_24h']:+.2f}%",
                                'Data Points': len(data_list),
                                'Min Price': f"${min(prices):,.8f}",
                                'Max Price': f"${max(prices):,.8f}",
                                'Avg Price': f"${np.mean(prices):,.8f}"
                            })
                    
                    df_summary = pd.DataFrame(summary_data)
                    st.dataframe(df_summary, use_container_width=True, hide_index=True)
        
        else:
            st.info("👆 Click 'Start Monitoring' to begin tracking prices")
        
        # Auto-refresh mechanism
        if auto_refresh_enabled and st.session_state.price_data:
            time.sleep(refresh_interval)
            st.session_state.force_refresh = True
            st.rerun()
        
        # Help Section
        with st.expander("❓ How to Use Real-Time Monitor"):
            st.markdown("""
            ### 📖 Usage Guide:
            
            **1. Select Symbols:**
            - Choose from dropdown list (recent scans)
            - Or enter symbols manually (one per line)
            
            **2. Start Monitoring:**
            - Click "Start Monitoring" to fetch initial prices
            - Enable "Auto Refresh" for continuous updates
            
            **3. View Data:**
            - Live price cards with 24h change
            - Expandable charts showing price history
            - Statistics and analytics
            
            **4. Export:**
            - Download CSV for external analysis
            - Generate summary report
            
            **Tips:**
            - Auto-refresh updates every N seconds automatically
            - Data is stored in session (max 100 data points per symbol)
            - Use manual refresh for on-demand updates
            - Clear data to reset and start fresh
            
            **Supported Exchanges:**
            - Binance Futures (most data available)
            - Bybit
            - Gate.io
            """)

    with tab7:
        st.header("📅 Important News Sentiment - Investing.com")
        
        # HTML code untuk economic calendar
        calendar_html = """
        <div style="background-color: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <iframe 
                src="https://sslecal2.investing.com?columns=exc_flags,exc_currency,exc_importance,exc_actual,exc_forecast,exc_previous&category=_employment,_economicActivity,_inflation,_credit,_centralBanks,_confidenceIndex,_balance,_Bonds&features=datepicker,timezone&countries=5&calType=week&timeZone=27&lang=1" 
                width="100%" 
                height="800" 
                frameborder="0" 
                allowtransparency="true" 
                marginwidth="0" 
                marginheight="0">
            </iframe>
            <div class="poweredBy" style="font-family: Arial, Helvetica, sans-serif; margin-top: 10px;">
                <span style="font-size: 11px;color: #333333;text-decoration: none;">
                    Real Time Economic Calendar provided by 
                    <a href="https://www.investing.com/" rel="nofollow" target="_blank" style="font-size: 11px;color: #06529D; font-weight: bold;" class="underline_link">Investing.com</a>.
                </span>
            </div>
        </div>
        """
        
        # Tampilkan HTML di Streamlit
        st.components.v1.html(calendar_html, height=800)
        
        # Tambahan informasi
        with st.expander("ℹ️ Cara Membaca Economic Calendar"):
            st.markdown("""
            **Keterangan Kolom:**
            - **🔴 High Impact**: Event dengan dampak tinggi terhadap pasar
            - **🟡 Medium Impact**: Event dengan dampak sedang  
            - **🟢 Low Impact**: Event dengan dampak rendah
            
            **Kategori Event:**
            - **Employment**: Data ketenagakerjaan (NFP, Unemployment Rate)
            - **Inflation**: Data inflasi (CPI, PPI)
            - **Central Banks**: Keputusan bank sentral (Fed, ECB, dll)
            - **GDP**: Pertumbuhan ekonomi
            - **Retail Sales**: Penjualan ritel
            
            **Tips Trading:**
            - Hindari trading 15 menit sebelum/sesudah high impact news
            - Perhatikan actual vs forecast
            - Siapkan risk management yang ketat
            """)
        st.header("📅 Important News Sentiment - Financial Juice")

        financial_juice_html = """
        <div style="background-color: white; padding: 20px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <iframe 
                src="https://www.financialjuice.com/home" 
                width="100%" 
                height="800" 
                frameborder="0" 
                allowtransparency="true" 
                marginwidth="0" 
                marginheight="0">
            </iframe>
        </div>
        """
    
        st.components.v1.html(financial_juice_html, height=800)

        st.header("🗽 Truth Social - @realDonaldTrump")
        
        js_code = """
        <script>
        function openTruthSocial() {
            window.open('https://truthsocial.com/@realDonaldTrump', '_blank', 
                    'width=1200,height=800,scrollbars=yes');
            return false;
        }
        
        function copyToClipboard() {
            navigator.clipboard.writeText('https://truthsocial.com/@realDonaldTrump');
            alert('Profi'Profilek copied to clipboard!');
        }
        </script>
        
        <style>
        .truth-card {
            background: white;
            border-radius: 15px;
            padding: 30px;
            box-shadow: 0 10px 30px rgba(0,0,0,0.1);
            text-align: center;
            margin: 20px 0;
        }
        .truth-btn {
            background: linear-gradient(45deg, #FF6B6B, #FF8E53);
            color: white;
            border: none;
            padding: 15px 30px;
            border-radius: 25px;
            font-size: 18px;
            cursor: pointer;
            margin: 10px;
            transition: all 0.3s ease;
        }
        .truth-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 5px 15px rgba(255,107,107,0.4);
        }
        .profile-stats {
            display: flex;
            justify-content: space-around;
            margin: 20px 0;
        }
        .stat-item {
            text-align: center;
        }
        </style>
        
        <div class="truth-card">
            <h2>Donald J. Trump</h2>
            <p style="color: #666; margin-bottom: 30px;">@realDonaldTrump on Truth Social</p>
            
            <div class="profile-stats">
                <div class="stat-item">
                    <h3>5.2M+</h3>
                    <p>Followers</p>
                </div>
                <div class="stat-item">
                    <h3>3.4K+</h3>
                    <p>Posts</p>
                </div>
                <div class="stat-item">
                    <h3>2022</h3>
                    <p>Joined</p>
                </div>
            </div>
            
            <button class="truth-btn" onclick="openTruthSocial()">
                🗽 Open Truth Social Profile
            </button>
            
            <button class="truth-btn" onclick="copyToClipboard()" 
                    style="background: linear-gradient(45deg, #667eea, #764ba2);">
                📋 Copy Profile Link
            </button>
            
            <div style="margin-top: 20px; color: #888;">
                <small>Truth Social doesn't allow direct embedding for security reasons</small>
            </div>
        </div>
        """
        
        components.html(js_code, height=800)
        
        # Additional information
        with st.expander("ℹ️ Why can't I embed Truth Social?"):
            st.markdown("""
            **Platform Security Restrictions:**
            
            Truth Social, like many modern social media platforms, implements:
            
            - **X-Frame-Options**: Prevents embedding in iframes
            - **Content Security Policy (CSP)**: Restricts cross-origin requests
            - **Clickjacking Protection**: Protects user interactions
            - **Authentication Requirements**: May require logged-in sessions
            
            **Best Alternatives:**
            1. Direct links (as shown above)
            2. Official mobile apps
            3. Browser extensions (if available)
            4. Official API access (if provided)
            """)



    # Footer
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: #666;'>"
        "Made with ❤️ by AI Assistant | "
        f"Database: {DB_NAME} | "
        f"Version 5.0"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()



def create_market_heatmap_safe(results):
    import pandas as _pd
    import numpy as _np
    import plotly.express as _px
    

    
    """
    Ultra-safe version with multiple fallback visualizations
    """
    if not results:
        return None
    
    # Prepare data with extensive validation
    data = []
    for r in results:
        try:
            tf_4h = r.get('timeframes', {}).get('4h', {})
            indicators = tf_4h.get('indicators', {}) if isinstance(tf_4h, dict) else {}
            
            symbol = str(r.get('name') or r.get('symbol') or 'Unknown')
            strength = float(r.get('strength', 0) or 0)
            consensus = str(r.get('consensus', 'NEUTRAL'))
            volume = float(indicators.get('volume_ratio', 1) or 1)
            
            # Only include valid entries
            if strength > 0 and volume > 0 and symbol != 'Unknown':
                data.append({
                    'symbol': symbol,
                    'strength': strength,
                    'consensus': consensus,
                    'volume': volume
                })
        except (TypeError, ValueError, KeyError) as e:
            continue
    
    if len(data) == 0:
        st.info("No valid data available for market heatmap")
        return None
    
    df = pd.DataFrame(data)
    
    # Data validation
    df['strength'] = pd.to_numeric(df['strength'], errors='coerce').fillna(0)
    df['volume'] = pd.to_numeric(df['volume'], errors='coerce').fillna(1)
    
    # Remove zero values
    df = df[df['strength'] > 0].copy()
    
    if len(df) == 0:
        st.warning("All strength values are zero - cannot create heatmap")
        return None
    
    # Ensure minimum values for visualization
    df['strength'] = df['strength'].clip(lower=0.1)
    df['volume'] = df['volume'].clip(lower=0.1)
    
    # Try Method 1: Treemap
    try:
        fig = px.treemap(
            df,
            path=['consensus', 'symbol'],
            values='strength',
            color='volume',
            color_continuous_scale='RdYlGn',
            title='Market Overview Heatmap'
        )
        fig.update_layout(height=600)
        return fig
    except Exception as e1:
        st.warning(f"Treemap failed: {str(e1)[:100]}. Trying alternative...")
    
    # Fallback Method 2: Sunburst
    try:
        fig = px.sunburst(
            df,
            path=['consensus', 'symbol'],
            values='strength',
            color='volume',
            color_continuous_scale='RdYlGn',
            title='Market Overview (Sunburst)'
        )
        fig.update_layout(height=600)
        return fig
    except Exception as e2:
        st.warning(f"Sunburst failed: {str(e2)[:100]}. Using bar chart...")
    
    # Fallback Method 3: Horizontal Bar Chart
    try:
        df_sorted = df.sort_values('strength', ascending=False).head(30)
        
        fig = go.Figure()
        
        for signal in df_sorted['consensus'].unique():
            df_signal = df_sorted[df_sorted['consensus'] == signal]
            
            color = 'green' if signal == 'LONG' else 'red' if signal == 'SHORT' else 'gray'
            
            fig.add_trace(go.Bar(
                y=df_signal['symbol'],
                x=df_signal['strength'],
                name=signal,
                orientation='h',
                marker_color=color,
                text=df_signal['strength'].round(1),
                textposition='auto'
            ))
        
        fig.update_layout(
            title='Market Overview (Top 30 by Strength)',
            xaxis_title='Strength %',
            yaxis_title='Symbol',
            height=800,
            barmode='group',
            showlegend=True
        )
        
        return fig
    except Exception as e3:
        st.error(f"All visualization methods failed: {e3}")
        return None

    
def plot_funding_chart(funding_raw, symbol):
    # funding_raw expected list-of-dicts with keys 'fundingTime' (ms) and 'fundingRate'
    try:
        df_fr = None
        if funding_raw is None:
            return None
        import pandas as _pd
        df_fr = _pd.DataFrame(funding_raw)
        if df_fr.empty:
            return None
        # normalize time column names
        if 'fundingTime' in df_fr.columns:
            df_fr['time'] = _pd.to_datetime(df_fr['fundingTime'], unit='ms', errors='coerce')
        elif 'time' in df_fr.columns:
            df_fr['time'] = _pd.to_datetime(df_fr['time'], unit='ms', errors='coerce')
        else:
            df_fr['time'] = _pd.NaT
        df_fr['fundingRate'] = _pd.to_numeric(df_fr['fundingRate'], errors='coerce')
        df_fr = df_fr.dropna(subset=['fundingRate'])
        if df_fr.empty:
            return None
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df_fr['time'], y=df_fr['fundingRate']*100, mode='lines+markers', name='Funding Rate (%)'))
        fig.update_layout(title=f"Funding Rate Trend ({symbol})", xaxis_title="Time", yaxis_title="Funding Rate (%)", height=300)
        return fig
    except Exception as e:
        return None

def plot_open_interest_chart(oi_raw, symbol):
    try:
        import pandas as _pd
        if oi_raw is None:
            return None
        df_oi = _pd.DataFrame(oi_raw)
        if df_oi.empty:
            return None
        # detect column for timestamp
        if 'timestamp' in df_oi.columns:
            df_oi['time'] = _pd.to_datetime(df_oi['timestamp'], unit='ms', errors='coerce')
        elif 'time' in df_oi.columns:
            df_oi['time'] = _pd.to_datetime(df_oi['time'], unit='ms', errors='coerce')
        elif 'ts' in df_oi.columns:
            df_oi['time'] = _pd.to_datetime(df_oi['ts'], unit='ms', errors='coerce')
        else:
            df_oi['time'] = _pd.NaT
        # detect column for open interest
        col = None
        for c in ['openInterest','sumOpenInterest','oi','sumOpenInterestValue']:
            if c in df_oi.columns:
                col = c
                break
        if col is None:
            return None
        df_oi[col] = _pd.to_numeric(df_oi[col], errors='coerce')
        df_oi = df_oi.dropna(subset=[col])
        if df_oi.empty:
            return None
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=df_oi['time'], y=df_oi[col], mode='lines+markers', name='Open Interest'))
        fig.update_layout(title=f"Open Interest Trend ({symbol})", xaxis_title="Time", yaxis_title="Open Interest", height=300)
        return fig
    except Exception as e:
        return None

def detect_funding_flip(funding_raw):
    # simple detection: check sign change in recent funding rates
    try:
        import pandas as _pd
        if not funding_raw:
            return None
        df = _pd.DataFrame(funding_raw)
        if 'fundingRate' not in df.columns:
            return None
        fr = df['fundingRate'].astype(float).fillna(0)
        if len(fr) < 4:
            return None
        # check last vs previous median
        last = fr.tail(3).mean()
        prev = fr.head(3).mean()
        if prev > 0 and last < 0:
            return {'flip': 'POS_TO_NEG', 'desc': 'Funding flipped positive → negative (bearish flip)'}
        if prev < 0 and last > 0:
            return {'flip': 'NEG_TO_POS', 'desc': 'Funding flipped negative → positive (bullish flip)'}
        return None
    except Exception as e:
        return None



def run_scanner_multiprocess(symbols, source, max_workers=None, debug=False):
    """
    Multiprocessing scanner runner. coins is list of symbol_data dicts.
    Returns list of analysis results.
    """

    import multiprocessing as _mp
    from concurrent.futures import ProcessPoolExecutor, as_completed as _as_completed
    import time as _time
    import os as _os

    cpu_count = _mp.cpu_count()
    if max_workers is None:
        max_workers = max(2, min(8, cpu_count - 1))
    st.text(f"[System] Auto-detected {cpu_count} cores -> using {max_workers} workers")

    results = []
    total = len(symbols)
    progress = st.progress(0)
    done = 0

    def _task_wrapper(symbol_data):
        # This wrapper is executed in worker process - must be top-level callable.
        try:
            # re-import necessary functions inside process if needed
            # We'll call analyze_single_coin which should be importable
            from importlib import import_module
            mod = import_module('__main__')
            return mod.analyze_single_coin(source, symbol_data)
        except Exception as e:
            return {'symbol': symbol_data.get('symbol'), 'error': str(e)}

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_task_wrapper, c): c for c in symbols}
        for f in _as_completed(futures):
            r = f.result()
            done += 1
            progress.progress(done / total)
            if r:
                results.append(r)
                # show brief per-coin log in Streamlit (debug needs to be enabled)
                if debug:
                    try:
                        sym = r.get('symbol') or r.get('name')
                        st.text(f"[Worker] {sym} -> {r.get('consensus')} ({r.get('strength')}%)")
                    except:
                        pass
    return results
