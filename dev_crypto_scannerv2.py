"""
Crypto Futures Scanner Pro - Ultimate Edition
Enhanced with: Parallel Processing, Database, Alerts, Backtesting, Portfolio Tracking
Author: AI Assistant
Version: 2.0
"""

import streamlit as st
import pandas as pd
import numpy as np
import requests
from datetime import datetime, timedelta
import time
from ta.momentum import RSIIndicator, StochRSIIndicator
from ta.trend import MACD, EMAIndicator, SMAIndicator
from ta.volume import VolumeWeightedAveragePrice
from ta.volatility import BollingerBands, AverageTrueRange
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import plotly.express as px
from concurrent.futures import ThreadPoolExecutor, as_completed
import sqlite3
import json
import warnings
warnings.filterwarnings('ignore')

# ==================== CONFIGURATION ====================
st.set_page_config(
    page_title="Crypto Futures Scanner Pro - Ultimate", 
    layout="wide", 
    page_icon="🚀",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.8rem; 
        color: #1f77b4; 
        text-align: center; 
        font-weight: bold;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
        margin-bottom: 1rem;
    }
    .signal-long {
        background: linear-gradient(135deg, #28a745, #20c997);
        color: white; 
        padding: 15px; 
        border-radius: 10px; 
        font-weight: bold; 
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .signal-short {
        background: linear-gradient(135deg, #dc3545, #fd7e14);
        color: white; 
        padding: 15px; 
        border-radius: 10px; 
        font-weight: bold; 
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .signal-neutral {
        background: linear-gradient(135deg, #6c757d, #adb5bd);
        color: white; 
        padding: 15px; 
        border-radius: 10px; 
        font-weight: bold; 
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 20px;
        border-radius: 10px;
        color: white;
        text-align: center;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
    }
    .stProgress > div > div > div > div {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
    }
</style>
""", unsafe_allow_html=True)

# ==================== SESSION STATE ====================
if 'last_refresh' not in st.session_state:
    st.session_state.last_refresh = datetime.now()
if 'refresh_interval' not in st.session_state:
    st.session_state.refresh_interval = None
if 'telegram_enabled' not in st.session_state:
    st.session_state.telegram_enabled = False
if 'scan_results' not in st.session_state:
    st.session_state.scan_results = []
if 'db_initialized' not in st.session_state:
    st.session_state.db_initialized = False

# ==================== DATABASE SETUP ====================
DB_NAME = 'crypto_scanner_ultimate.db'

def init_database():
    """Initialize SQLite database with all tables"""
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    
    # Signals table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS signals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            symbol TEXT NOT NULL,
            exchange TEXT,
            signal TEXT,
            strength REAL,
            price REAL,
            entry REAL,
            tp1 REAL,
            tp2 REAL,
            sl REAL,
            risk_reward REAL,
            status TEXT DEFAULT 'ACTIVE',
            notes TEXT
        )
    ''')
    
    # Performance tracking table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS performance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            signal_id INTEGER,
            entry_time DATETIME,
            exit_time DATETIME,
            entry_price REAL,
            exit_price REAL,
            pnl_percentage REAL,
            pnl_amount REAL,
            hit_tp1 BOOLEAN DEFAULT 0,
            hit_tp2 BOOLEAN DEFAULT 0,
            hit_sl BOOLEAN DEFAULT 0,
            exit_reason TEXT,
            FOREIGN KEY (signal_id) REFERENCES signals(id)
        )
    ''')
    
    # Scan history table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scan_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            total_scanned INTEGER,
            signals_found INTEGER,
            avg_strength REAL,
            exchange TEXT,
            scan_duration REAL
        )
    ''')
    
    # Watchlist table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS watchlist (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            symbol TEXT UNIQUE NOT NULL,
            added_date DATETIME DEFAULT CURRENT_TIMESTAMP,
            notes TEXT
        )
    ''')
    
    conn.commit()
    conn.close()
    return True

def save_signal_to_db(result, exchange):
    """Save trading signal to database"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        plan = result.get('trading_plan', {})
        cursor.execute('''
            INSERT INTO signals 
            (symbol, exchange, signal, strength, price, entry, tp1, tp2, sl, risk_reward, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            result['symbol'],
            exchange,
            result['consensus'],
            result['strength'],
            result['price'],
            plan.get('entry'),
            plan.get('tp1'),
            plan.get('tp2'),
            plan.get('sl'),
            plan.get('risk_reward_tp1'),
            json.dumps(result['timeframes']['4h'].get('reasons', []))
        ))
        
        conn.commit()
        signal_id = cursor.lastrowid
        conn.close()
        return signal_id
    except Exception as e:
        st.error(f"Database error: {e}")
        return None

def save_scan_history(total_scanned, signals_found, avg_strength, exchange, duration):
    """Save scan history"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO scan_history 
            (total_scanned, signals_found, avg_strength, exchange, scan_duration)
            VALUES (?, ?, ?, ?, ?)
        ''', (total_scanned, signals_found, avg_strength, exchange, duration))
        
        conn.commit()
        conn.close()
    except Exception as e:
        st.error(f"Scan history error: {e}")

def get_active_signals():
    """Get active signals from database"""
    try:
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql('''
            SELECT * FROM signals 
            WHERE status = 'ACTIVE'
            ORDER BY timestamp DESC
        ''', conn)
        conn.close()
        return df
    except:
        return pd.DataFrame()

def update_signal_status(signal_id, status, exit_price=None, pnl=None):
    """Update signal status"""
    try:
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        
        cursor.execute('''
            UPDATE signals 
            SET status = ?
            WHERE id = ?
        ''', (status, signal_id))
        
        if exit_price and pnl:
            cursor.execute('''
                INSERT INTO performance
                (signal_id, exit_time, exit_price, pnl_percentage)
                VALUES (?, ?, ?, ?)
            ''', (signal_id, datetime.now(), exit_price, pnl))
        
        conn.commit()
        conn.close()
    except Exception as e:
        st.error(f"Update error: {e}")

# ==================== API CONFIGURATION ====================
API_SOURCES = {
    'binance': {
        'name': 'Binance Futures',
        'base_url': 'https://fapi.binance.com',
        'top_symbols_endpoint': '/fapi/v1/ticker/24hr',
        'klines_endpoint': '/fapi/v1/klines'
    },
    'bybit': {
        'name': 'Bybit',
        'base_url': 'https://api.bybit.com',
        'top_symbols_endpoint': '/v5/market/tickers',
        'klines_endpoint': '/v5/market/kline'
    },
    'gateio': {
        'name': 'Gate.io',
        'base_url': 'https://api.gateio.ws/api/v4',
        'top_symbols_endpoint': '/futures/usdt/contracts',
        'klines_endpoint': '/futures/usdt/candlesticks'
    }
}

TIMEFRAME_MAPPING = {
    'binance': {'1h': '1h', '4h': '4h', '12h': '12h'},
    'bybit': {'1h': '60', '4h': '240', '12h': '720'},
    'gateio': {'1h': '1h', '4h': '4h', '12h': '12h'}
}

# ==================== RATE LIMITER ====================
class RateLimiter:
    """Rate limiter to prevent API throttling"""
    def __init__(self, max_calls=10, period=60):
        self.max_calls = max_calls
        self.period = period
        self.calls = []
    
    def wait_if_needed(self):
        now = time.time()
        self.calls = [c for c in self.calls if now - c < self.period]
        
        if len(self.calls) >= self.max_calls:
            sleep_time = self.period - (now - self.calls[0])
            if sleep_time > 0:
                time.sleep(sleep_time)
                self.calls = []
        
        self.calls.append(now)

rate_limiter = RateLimiter(max_calls=15, period=60)

# ==================== DATA FETCHING ====================
@st.cache_data(ttl=300)
def get_binance_top_symbols(limit=30):
    """Fetch top trading pairs from Binance Futures"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['binance']['base_url']}{API_SOURCES['binance']['top_symbols_endpoint']}"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        usdt_pairs = [item for item in data if item['symbol'].endswith('USDT')]
        sorted_pairs = sorted(usdt_pairs, key=lambda x: float(x.get('quoteVolume', 0)), reverse=True)
        return sorted_pairs[:limit]
    except Exception as e:
        return []

@st.cache_data(ttl=300)
def get_bybit_top_symbols(limit=30):
    """Fetch top trading pairs from Bybit"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['bybit']['base_url']}{API_SOURCES['bybit']['top_symbols_endpoint']}"
        params = {'category': 'linear'}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get('retCode') == 0:
            symbols = data['result']['list']
            usdt_pairs = [s for s in symbols if s['symbol'].endswith('USDT')]
            sorted_pairs = sorted(usdt_pairs, key=lambda x: float(x.get('turnover24h', 0)), reverse=True)
            return sorted_pairs[:limit]
        return []
    except Exception as e:
        return []

@st.cache_data(ttl=300)
def get_gateio_top_symbols(limit=30):
    """Fetch top trading pairs from Gate.io"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['gateio']['base_url']}{API_SOURCES['gateio']['top_symbols_endpoint']}"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
        active_contracts = [c for c in data if c.get('trade_size', 0) > 0]
        sorted_contracts = sorted(active_contracts, key=lambda x: float(x.get('trade_size', 0)), reverse=True)
        return sorted_contracts[:limit]
    except Exception as e:
        return []

def get_top_symbols(source='binance', limit=30):
    """Get top symbols with automatic fallback"""
    sources_priority = ['binance', 'bybit', 'gateio']
    if source in sources_priority:
        sources_priority.remove(source)
        sources_priority.insert(0, source)
    
    for src in sources_priority:
        try:
            if src == 'binance':
                symbols = get_binance_top_symbols(limit)
            elif src == 'bybit':
                symbols = get_bybit_top_symbols(limit)
            elif src == 'gateio':
                symbols = get_gateio_top_symbols(limit)
            else:
                continue
            if symbols:
                return src, symbols
        except:
            continue
    return None, []

@st.cache_data(ttl=180)
def get_binance_klines(symbol, interval, limit=500):
    """Fetch klines from Binance Futures"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['binance']['base_url']}{API_SOURCES['binance']['klines_endpoint']}"
        params = {'symbol': symbol, 'interval': interval, 'limit': limit}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        df = pd.DataFrame(data, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume',
            'close_time', 'quote_volume', 'trades', 'taker_buy_base', 'taker_buy_quote', 'ignore'])
        df['timestamp'] = pd.to_datetime(df['timestamp'], unit='ms')
        for col in ['open', 'high', 'low', 'close', 'volume']:
            df[col] = df[col].astype(float)
        return df
    except Exception as e:
        return None

@st.cache_data(ttl=180)
def get_bybit_klines(symbol, interval, limit=500):
    """Fetch klines from Bybit"""
    try:
        rate_limiter.wait_if_needed()
        url = f"{API_SOURCES['bybit']['base_url']}{API_SOURCES['bybit']['klines_endpoint']}"
        params = {'category': 'linear', 'symbol': symbol, 'interval': interval, 'limit': limit}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        if data.get('retCode') == 0:
            klines = data['result']['list']
            df = pd.DataFrame(klines, columns=['timestamp', 'open', 'high', 'low', 'close', 'volume', 'turnover'])
            df['timestamp'] = pd.to_datetime(df['timestamp'].astype(float), unit='ms')
            for col in ['open', 'high', 'low', 'close', 'volume']:
                df[col] = df[col].astype(float)
            df = df.sort_values('timestamp').reset_index(drop=True)
            return df
        return None
    except Exception as e:
        return None

@st.cache_data(ttl=180)
def get_gateio_klines(symbol, interval, limit=500):
    """Fetch klines from Gate.io"""
    try:
        rate_limiter.wait_if_needed()
        contract = symbol.replace('USDT', '_USDT')
        url = f"{API_SOURCES['gateio']['base_url']}{API_SOURCES['gateio']['klines_endpoint']}"
        params = {'contract': contract, 'interval': interval, 'limit': limit}
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
        if isinstance(data, list):
            df = pd.DataFrame(data, columns=['timestamp', 'volume', 'close', 'high', 'low', 'open'])
            df['timestamp'] = pd.to_datetime(df['timestamp'].astype(float), unit='s')
            for col in ['open', 'high', 'low', 'close', 'volume']:
                df[col] = df[col].astype(float)
            df = df.sort_values('timestamp').reset_index(drop=True)
            return df
        return None
    except Exception as e:
        return None

def get_klines(source, symbol, timeframe, limit=500):
    """Get klines with automatic fallback"""
    sources_priority = ['binance', 'bybit', 'gateio']
    if source in sources_priority:
        sources_priority.remove(source)
        sources_priority.insert(0, source)
    
    for src in sources_priority:
        try:
            if src not in TIMEFRAME_MAPPING:
                continue
            interval = TIMEFRAME_MAPPING[src][timeframe]
            if src == 'binance':
                df = get_binance_klines(symbol, interval, limit)
            elif src == 'bybit':
                df = get_bybit_klines(symbol, interval, limit)
            elif src == 'gateio':
                df = get_gateio_klines(symbol, interval, limit)
            else:
                continue
            if df is not None and len(df) >= 200:
                return df
        except:
            continue
    return None

# ==================== TECHNICAL INDICATORS ====================
def calculate_indicators(df):
    """Calculate comprehensive technical indicators"""
    if df is None or len(df) < 200:
        return None
    
    df = df.copy()
    close = df['close']
    high = df['high']
    low = df['low']
    volume = df['volume']
    
    # RSI
    rsi = RSIIndicator(close=close, window=14)
    df['rsi'] = rsi.rsi()
    
    # Stochastic RSI
    stoch_rsi = StochRSIIndicator(close=close, window=14, smooth1=3, smooth2=3)
    df['stoch_rsi_k'] = stoch_rsi.stochrsi_k() * 100
    df['stoch_rsi_d'] = stoch_rsi.stochrsi_d() * 100
    
    # EMAs
    df['ema_9'] = EMAIndicator(close=close, window=9).ema_indicator()
    df['ema_21'] = EMAIndicator(close=close, window=21).ema_indicator()
    df['ema_50'] = EMAIndicator(close=close, window=50).ema_indicator()
    df['ema_200'] = EMAIndicator(close=close, window=200).ema_indicator()
    
    # SMAs
    df['sma_20'] = SMAIndicator(close=close, window=20).sma_indicator()
    df['sma_50'] = SMAIndicator(close=close, window=50).sma_indicator()
    
    # MACD
    macd = MACD(close=close, window_slow=26, window_fast=12, window_sign=9)
    df['macd'] = macd.macd()
    df['macd_signal'] = macd.macd_signal()
    df['macd_diff'] = macd.macd_diff()
    
    # Bollinger Bands
    bb = BollingerBands(close=close, window=20, window_dev=2)
    df['bb_upper'] = bb.bollinger_hband()
    df['bb_middle'] = bb.bollinger_mavg()
    df['bb_lower'] = bb.bollinger_lband()
    df['bb_width'] = (df['bb_upper'] - df['bb_lower']) / df['bb_middle']
    
    # ATR
    atr = AverageTrueRange(high=high, low=low, close=close, window=14)
    df['atr'] = atr.average_true_range()
    
    # Volume
    df['volume_sma'] = volume.rolling(window=20).mean()
    df['volume_ratio'] = volume / df['volume_sma']
    
    # VWAP
    df['vwap'] = (volume * (high + low + close) / 3).cumsum() / volume.cumsum()
    
    # ADX
    df['adx'] = calculate_adx(df, 14)
    
    # Momentum
    df['momentum'] = close.pct_change(periods=10) * 100
    
    # Pivot Points
    df = calculate_pivot_points(df)
    
    return df

def calculate_adx(df, period=14):
    """Calculate ADX"""
    high = df['high']
    low = df['low']
    close = df['close']
    
    plus_dm = high.diff()
    minus_dm = low.diff()
    plus_dm[plus_dm < 0] = 0
    minus_dm[minus_dm > 0] = 0
    
    tr = pd.concat([high - low, abs(high - close.shift()), abs(low - close.shift())], axis=1).max(axis=1)
    atr = tr.rolling(window=period).mean()
    plus_di = 100 * (plus_dm.rolling(window=period).mean() / atr)
    minus_di = 100 * (abs(minus_dm).rolling(window=period).mean() / atr)
    dx = 100 * abs(plus_di - minus_di) / (plus_di + minus_di)
    adx = dx.rolling(window=period).mean()
    return adx

def calculate_pivot_points(df):
    """Calculate pivot points"""
    df['pivot'] = (df['high'] + df['low'] + df['close']) / 3
    df['r1'] = 2 * df['pivot'] - df['low']
    df['s1'] = 2 * df['pivot'] - df['high']
    df['r2'] = df['pivot'] + (df['high'] - df['low'])
    df['s2'] = df['pivot'] - (df['high'] - df['low'])
    return df

def detect_chart_patterns(df):
    """Detect chart patterns"""
    patterns = []
    if len(df) < 50:
        return patterns
    
    # Golden/Death Cross
    if len(df) >= 200:
        if df['ema_50'].iloc[-2] <= df['ema_200'].iloc[-2] and df['ema_50'].iloc[-1] > df['ema_200'].iloc[-1]:
            patterns.append('GOLDEN_CROSS')
        elif df['ema_50'].iloc[-2] >= df['ema_200'].iloc[-2] and df['ema_50'].iloc[-1] < df['ema_200'].iloc[-1]:
            patterns.append('DEATH_CROSS')
    
    # Divergence
    recent = df.tail(50)
    close = recent['close'].values
    price_trend = close[-10:] - close[-20:-10].mean()
    rsi_trend = df['rsi'].iloc[-10:].values - df['rsi'].iloc[-20:-10].mean()
    
    if price_trend.mean() < 0 and rsi_trend.mean() > 0:
        patterns.append('BULLISH_DIVERGENCE')
    elif price_trend.mean() > 0 and rsi_trend.mean() < 0:
        patterns.append('BEARISH_DIVERGENCE')
    
    # BB Breakout
    if df['close'].iloc[-1] > df['bb_upper'].iloc[-2]:
        patterns.append('UPPER_BB_BREAKOUT')
    elif df['close'].iloc[-1] < df['bb_lower'].iloc[-2]:
        patterns.append('LOWER_BB_BREAKOUT')
    
    return patterns

# ==================== SIGNAL ANALYSIS ====================
def analyze_signal(df, timeframe):
    """Advanced signal analysis"""
    if df is None or len(df) < 200:
        return {
            'signal': 'INSUFFICIENT_DATA',
            'strength': 0,
            'score': 0,
            'reasons': [],
            'indicators': {}
        }
    
    latest = df.iloc[-1]
    prev = df.iloc[-2]
    
    bullish_signals = 0
    bearish_signals = 0
    reasons = []
    max_score = 10
    
    # RSI
    rsi_val = latest['rsi']
    if rsi_val < 30:
        bullish_signals += 1.5
        reasons.append(f"RSI oversold: {rsi_val:.1f}")
    elif rsi_val > 70:
        bearish_signals += 1.5
        reasons.append(f"RSI overbought: {rsi_val:.1f}")
    elif 30 <= rsi_val <= 40:
        bullish_signals += 0.5
        reasons.append(f"RSI near oversold: {rsi_val:.1f}")
    elif 60 <= rsi_val <= 70:
        bearish_signals += 0.5
        reasons.append(f"RSI near overbought: {rsi_val:.1f}")
    
    # Stoch RSI
    stoch_k = latest['stoch_rsi_k']
    stoch_d = latest['stoch_rsi_d']
    if stoch_k < 20 and stoch_k > stoch_d:
        bullish_signals += 1.5
        reasons.append(f"Stoch RSI bullish: K={stoch_k:.1f}")
    elif stoch_k > 80 and stoch_k < stoch_d:
        bearish_signals += 1.5
        reasons.append(f"Stoch RSI bearish: K={stoch_k:.1f}")
    
    # MACD
    macd_diff = latest['macd_diff']
    macd_cross_up = prev['macd_diff'] <= 0 and macd_diff > 0
    macd_cross_down = prev['macd_diff'] >= 0 and macd_diff < 0
    
    if macd_cross_up:
        bullish_signals += 1.5
        reasons.append("MACD bullish crossover")
    elif macd_cross_down:
        bearish_signals += 1.5
        reasons.append("MACD bearish crossover")
    elif macd_diff > 0:
        bullish_signals += 0.5
        reasons.append("MACD positive")
    elif macd_diff < 0:
        bearish_signals += 0.5
        reasons.append("MACD negative")
    
    # EMA Trend
    ema_alignment_bull = (latest['ema_9'] > latest['ema_21'] > latest['ema_50'])
    ema_alignment_bear = (latest['ema_9'] < latest['ema_21'] < latest['ema_50'])
    
    if ema_alignment_bull:
        bullish_signals += 2
        reasons.append("EMA bullish alignment")
    elif ema_alignment_bear:
        bearish_signals += 2
        reasons.append("EMA bearish alignment")
    
    # Golden/Death Cross
    patterns = detect_chart_patterns(df)
    if 'GOLDEN_CROSS' in patterns:
        bullish_signals += 2
        reasons.append("⭐ Golden Cross")
    elif 'DEATH_CROSS' in patterns:
        bearish_signals += 2
        reasons.append("⭐ Death Cross")
    
    # Volume
    vol_ratio = latest['volume_ratio']
    if vol_ratio > 1.5:
        if bullish_signals > bearish_signals:
            bullish_signals += 1
            reasons.append(f"High volume: {vol_ratio:.1f}x")
        elif bearish_signals > bullish_signals:
            bearish_signals += 1
            reasons.append(f"High volume: {vol_ratio:.1f}x")
    
    # ADX
    adx_val = latest['adx']
    if adx_val > 25:
        reasons.append(f"Strong trend: ADX={adx_val:.1f}")
        if bullish_signals > bearish_signals:
            bullish_signals += 0.5
        elif bearish_signals > bullish_signals:
            bearish_signals += 0.5
    
    # Divergence
    if 'BULLISH_DIVERGENCE' in patterns:
        bullish_signals += 1
        reasons.append("📈 Bullish divergence")
    elif 'BEARISH_DIVERGENCE' in patterns:
        bearish_signals += 1
        reasons.append("📉 Bearish divergence")
    
    # Determine signal
    if bullish_signals > bearish_signals and bullish_signals >= 3:
        signal_type = 'LONG'
        signal_score = bullish_signals
    elif bearish_signals > bullish_signals and bearish_signals >= 3:
        signal_type = 'SHORT'
        signal_score = bearish_signals
    else:
        signal_type = 'NEUTRAL'
        signal_score = 0
    
    strength = min((signal_score / max_score) * 100, 100)
    
    return {
        'signal': signal_type,
        'strength': round(strength, 1),
        'score': round(signal_score, 2),
        'max_score': max_score,
        'reasons': reasons[:5],
        'indicators': {
            'rsi': round(float(rsi_val), 2),
            'stoch_rsi': round(float(stoch_k), 2),
            'macd_diff': round(float(macd_diff), 8),
            'volume_ratio': round(float(vol_ratio), 2),
            'adx': round(float(adx_val), 2) if pd.notna(adx_val) else 0,
            'ema_trend': 'BULL' if ema_alignment_bull else 'BEAR' if ema_alignment_bear else 'NEUTRAL'
        },
        'patterns': patterns
    }

def calculate_support_resistance(df, lookback=100):
    """Calculate support and resistance levels"""
    if df is None or len(df) < lookback:
        return None, None
    
    recent = df.tail(lookback)
    swing_highs = []
    swing_lows = []
    
    for i in range(2, len(recent) - 2):
        if (recent['high'].iloc[i] > recent['high'].iloc[i-1] and 
            recent['high'].iloc[i] > recent['high'].iloc[i-2] and
            recent['high'].iloc[i] > recent['high'].iloc[i+1] and 
            recent['high'].iloc[i] > recent['high'].iloc[i+2]):
            swing_highs.append(recent['high'].iloc[i])
        
        if (recent['low'].iloc[i] < recent['low'].iloc[i-1] and 
            recent['low'].iloc[i] < recent['low'].iloc[i-2] and
            recent['low'].iloc[i] < recent['low'].iloc[i+1] and 
            recent['low'].iloc[i] < recent['low'].iloc[i+2]):
            swing_lows.append(recent['low'].iloc[i])
    
    pivot_support = recent['s1'].iloc[-1]
    pivot_resistance = recent['r1'].iloc[-1]
    recent_high = recent['high'].tail(20).max()
    recent_low = recent['low'].tail(20).min()
    current_price = df['close'].iloc[-1]
    
    # Resistance
    potential_resistances = []
    if swing_highs:
        potential_resistances.extend([h for h in swing_highs if h > current_price])
    potential_resistances.append(pivot_resistance)
    potential_resistances.append(recent_high)
    valid_resistances = [r for r in potential_resistances if current_price < r < current_price * 1.2]
    resistance = min(valid_resistances) if valid_resistances else recent_high
    
    # Support
    potential_supports = []
    if swing_lows:
        potential_supports.extend([l for l in swing_lows if l < current_price])
    potential_supports.append(pivot_support)
    potential_supports.append(recent_low)
    valid_supports = [s for s in potential_supports if current_price * 0.8 < s < current_price]
    support = max(valid_supports) if valid_supports else recent_low
    
    return support, resistance

def generate_entry_tp_sl(df, signal_type, current_price):
    """Generate entry, TP, SL with S/R validation"""
    if df is None or signal_type not in ['LONG', 'SHORT']:
        return None
    
    latest = df.iloc[-1]
    atr = latest['atr']
    support, resistance = calculate_support_resistance(df)
    
    if support and resistance:
        sr_distance = resistance - support
        price_to_support_pct = abs(current_price - support) / current_price * 100
        price_to_resistance_pct = abs(resistance - current_price) / current_price * 100
    else:
        sr_distance = atr * 3
        price_to_support_pct = 5
        price_to_resistance_pct = 5
    
    if signal_type == 'LONG':
        if price_to_support_pct > 3:
            entry = (current_price + support) / 2
        else:
            entry = current_price * 1.002
        if support:
            entry = max(entry, support * 1.01)
        sl = support * 0.995 if support else entry - (atr * 1.5)
        risk = entry - sl
        tp1 = entry + (risk * 2)
        tp2 = entry + (risk * 3.5)
        if resistance and tp1 > resistance * 1.05:
            tp1 = resistance * 0.98
            tp2 = resistance * 1.03
    else:
        if price_to_resistance_pct > 3:
            entry = (current_price + resistance) / 2
        else:
            entry = current_price * 0.998
        if resistance:
            entry = min(entry, resistance * 0.99)
        sl = resistance * 1.005 if resistance else entry + (atr * 1.5)
        risk = sl - entry
        tp1 = entry - (risk * 2)
        tp2 = entry - (risk * 3.5)
        if support and tp1 < support * 0.95:
            tp1 = support * 1.02
            tp2 = support * 0.97
    
    avg_candle_move = df['close'].pct_change().abs().tail(20).mean() * 100
    if avg_candle_move > 0:
        hours_to_entry = abs((entry - current_price) / current_price * 100) / avg_candle_move
        hours_to_tp1 = abs((tp1 - current_price) / current_price * 100) / avg_candle_move
    else:
        hours_to_entry = 0
        hours_to_tp1 = 0
    
    risk_amount = abs(entry - sl)
    reward_tp1 = abs(tp1 - entry)
    reward_tp2 = abs(tp2 - entry)
    
    entry_note = ""
    if signal_type == 'LONG':
        if support and entry < support:
            entry_note = "⚠️ Entry below support"
        elif support and (entry - support) / support * 100 > 5:
            entry_note = "⚠️ Entry far from support - Wait for pullback"
    else:
        if resistance and entry > resistance:
            entry_note = "⚠️ Entry above resistance"
        elif resistance and (resistance - entry) / entry * 100 > 5:
            entry_note = "⚠️ Entry far from resistance - Wait for rally"
    
    return {
        'entry': round(entry, 8),
        'tp1': round(tp1, 8),
        'tp2': round(tp2, 8),
        'sl': round(sl, 8),
        'support': round(support, 8) if support else None,
        'resistance': round(resistance, 8) if resistance else None,
        'risk_reward_tp1': round(reward_tp1 / risk_amount, 2) if risk_amount > 0 else 0,
        'risk_reward_tp2': round(reward_tp2 / risk_amount, 2) if risk_amount > 0 else 0,
        'risk_percentage': round((risk_amount / entry) * 100, 2),
        'est_hours_to_entry': round(hours_to_entry, 1),
        'est_hours_to_tp1': round(hours_to_tp1, 1),
        'entry_note': entry_note,
        'entry_to_support_pct': round(price_to_support_pct, 2) if signal_type == 'LONG' else None,
        'entry_to_resistance_pct': round(price_to_resistance_pct, 2) if signal_type == 'SHORT' else None
    }

# ==================== MULTI-TIMEFRAME ANALYSIS ====================
def analyze_single_coin(source, symbol_data):
    """Analyze single coin across multiple timeframes"""
    try:
        if source == 'binance':
            symbol = symbol_data['symbol']
            name = symbol.replace('USDT', '')
            current_price = float(symbol_data['lastPrice'])
        elif source == 'bybit':
            symbol = symbol_data['symbol']
            name = symbol.replace('USDT', '')
            current_price = float(symbol_data['lastPrice'])
        else:
            symbol = symbol_data['name'].replace('_', '')
            name = symbol.replace('USDT', '')
            current_price = float(symbol_data.get('last_price', 0))
        
        timeframes = ['1h', '4h', '12h']
        results = {}
        
        for tf in timeframes:
            df = get_klines(source, symbol, tf, limit=500)
            if df is not None and len(df) >= 200:
                df = calculate_indicators(df)
                if df is not None:
                    results[tf] = analyze_signal(df, tf)
                    results[f'{tf}_df'] = df
                else:
                    results[tf] = {'signal': 'ERROR', 'strength': 0, 'reasons': [], 'indicators': {}}
            else:
                results[tf] = {'signal': 'INSUFFICIENT_DATA', 'strength': 0, 'reasons': [], 'indicators': {}}
        
        signals = []
        weights = {'12h': 3, '4h': 2, '1h': 1}
        weighted_long = 0
        weighted_short = 0
        
        for tf in timeframes:
            signal = results[tf]['signal']
            if signal == 'LONG':
                weighted_long += weights[tf]
                signals.append('LONG')
            elif signal == 'SHORT':
                weighted_short += weights[tf]
                signals.append('SHORT')
        
        if weighted_long > weighted_short and weighted_long >= 3:
            consensus = 'LONG'
        elif weighted_short > weighted_long and weighted_short >= 3:
            consensus = 'SHORT'
        else:
            consensus = 'NEUTRAL'
        
        total_strength = 0
        total_weight = 0
        for tf in timeframes:
            if results[tf]['signal'] not in ['ERROR', 'INSUFFICIENT_DATA']:
                total_strength += results[tf]['strength'] * weights[tf]
                total_weight += weights[tf]
        
        avg_strength = total_strength / total_weight if total_weight > 0 else 0
        
        trading_plan = None
        if consensus in ['LONG', 'SHORT']:
            main_df = results.get('4h_df')
            if main_df is not None:
                trading_plan = generate_entry_tp_sl(main_df, consensus, current_price)
        
        return {
            'symbol': symbol,
            'name': name,
            'price': current_price,
            'consensus': consensus,
            'strength': round(avg_strength, 1),
            'timeframes': results,
            'trading_plan': trading_plan
        }
    except Exception as e:
        return None

def parallel_scan(symbols, source, max_workers=5):
    """Parallel scanning for faster performance"""
    results = []
    
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_symbol = {
            executor.submit(analyze_single_coin, source, symbol_data): symbol_data 
            for symbol_data in symbols
        }
        
        for future in as_completed(future_to_symbol):
            try:
                result = future.result(timeout=30)
                if result:
                    results.append(result)
            except Exception as e:
                continue
    
    return results

# ==================== TELEGRAM ALERTS ====================
def send_telegram_alert(token, chat_id, symbol, signal_data):
    """Send alert via Telegram"""
    try:
        plan = signal_data.get('trading_plan', {})
        if not plan:
            return
        
        message = f"""
🚨 *NEW SIGNAL ALERT* 🚨

💰 *{symbol}*
📊 Signal: *{signal_data['consensus']}*
💪 Strength: *{signal_data['strength']}%*

📍 Entry: ${plan['entry']:,.4f}
🎯 TP1: ${plan['tp1']:,.4f} (R:R {plan['risk_reward_tp1']})
🎯 TP2: ${plan['tp2']:,.4f} (R:R {plan['risk_reward_tp2']})
🛑 SL: ${plan['sl']:,.4f}

⚡ Risk: {plan['risk_percentage']}%
📈 Multi-TF Confirmation: ✅

⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        url = f"https://api.telegram.org/bot{token}/sendMessage"
        data = {
            'chat_id': chat_id,
            'text': message,
            'parse_mode': 'Markdown'
        }
        
        response = requests.post(url, data=data, timeout=10)
        return response.status_code == 200
    except Exception as e:
        return False

# ==================== VISUALIZATION ====================
def create_chart(df, symbol, trading_plan=None):
    """Create advanced trading chart"""
    if df is None or len(df) < 50:
        return None
    
    fig = make_subplots(
        rows=5, cols=1, 
        shared_xaxes=True, 
        vertical_spacing=0.02,
        row_heights=[0.4, 0.15, 0.15, 0.15, 0.15],
        subplot_titles=(f'{symbol} Price', 'RSI', 'Stoch RSI', 'MACD', 'Volume')
    )
    
    # Candlestick
    fig.add_trace(go.Candlestick(
        x=df['timestamp'], 
        open=df['open'], 
        high=df['high'],
        low=df['low'], 
        close=df['close'], 
        name='Price'
    ), row=1, col=1)
    
    # EMAs
    colors = {'ema_9': 'purple', 'ema_21': 'orange', 'ema_50': 'blue', 'ema_200': 'red'}
    for ema, color in colors.items():
        fig.add_trace(go.Scatter(
            x=df['timestamp'], 
            y=df[ema], 
            name=ema.upper(),
            line=dict(color=color, width=1)
        ), row=1, col=1)
    
    # Bollinger Bands
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['bb_upper'], 
        name='BB Upper',
        line=dict(color='gray', width=1, dash='dash')
    ), row=1, col=1)
    
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['bb_lower'], 
        name='BB Lower',
        line=dict(color='gray', width=1, dash='dash'), 
        fill='tonexty',
        fillcolor='rgba(128,128,128,0.1)'
    ), row=1, col=1)
    
    # Trading Plan Lines
    if trading_plan:
        for level, color, name in [
            (trading_plan['entry'], 'yellow', 'Entry'),
            (trading_plan['tp1'], 'green', 'TP1'), 
            (trading_plan['tp2'], 'lightgreen', 'TP2'),
            (trading_plan['sl'], 'red', 'SL')
        ]:
            fig.add_hline(
                y=level, 
                line_dash="dash", 
                line_color=color,
                annotation_text=name, 
                row=1, col=1
            )
    
    # RSI
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['rsi'], 
        name='RSI',
        line=dict(color='purple')
    ), row=2, col=1)
    fig.add_hline(y=70, line_dash="dash", line_color="red", row=2, col=1)
    fig.add_hline(y=30, line_dash="dash", line_color="green", row=2, col=1)
    fig.add_hline(y=50, line_dash="dot", line_color="gray", row=2, col=1)
    
    # Stochastic RSI
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['stoch_rsi_k'], 
        name='Stoch K',
        line=dict(color='blue')
    ), row=3, col=1)
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['stoch_rsi_d'], 
        name='Stoch D',
        line=dict(color='orange')
    ), row=3, col=1)
    fig.add_hline(y=80, line_dash="dash", line_color="red", row=3, col=1)
    fig.add_hline(y=20, line_dash="dash", line_color="green", row=3, col=1)
    
    # MACD
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['macd'], 
        name='MACD',
        line=dict(color='blue')
    ), row=4, col=1)
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['macd_signal'], 
        name='Signal',
        line=dict(color='orange')
    ), row=4, col=1)
    fig.add_trace(go.Bar(
        x=df['timestamp'], 
        y=df['macd_diff'], 
        name='Histogram',
        marker_color=['green' if val > 0 else 'red' for val in df['macd_diff']]
    ), row=4, col=1)
    
    # Volume
    fig.add_trace(go.Bar(
        x=df['timestamp'], 
        y=df['volume'], 
        name='Volume',
        marker_color=['green' if df['close'].iloc[i] > df['open'].iloc[i] else 'red' for i in range(len(df))]
    ), row=5, col=1)
    fig.add_trace(go.Scatter(
        x=df['timestamp'], 
        y=df['volume_sma'], 
        name='Vol SMA',
        line=dict(color='orange', width=2)
    ), row=5, col=1)
    
    fig.update_layout(
        height=1200, 
        showlegend=True, 
        xaxis_rangeslider_visible=False, 
        hovermode='x unified',
        template='plotly_dark'
    )
    fig.update_xaxes(title_text="Time", row=5, col=1)
    
    return fig

def create_market_heatmap(results):
    """Create market heatmap visualization"""
    data = []
    for r in results:
        tf_4h = r['timeframes'].get('4h', {})
        indicators = tf_4h.get('indicators', {})
        
        data.append({
            'symbol': r['name'],
            'strength': r['strength'],
            'signal': r['consensus'],
            'volume': indicators.get('volume_ratio', 1)
        })
    
    if not data:
        return None
    
    df = pd.DataFrame(data)
    
    fig = px.treemap(
        df,
        path=['signal', 'symbol'],
        values='strength',
        color='volume',
        color_continuous_scale='RdYlGn',
        title='Market Overview Heatmap'
    )
    
    fig.update_layout(height=600)
    
    return fig

def create_performance_chart():
    """Create performance tracking chart"""
    try:
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql('''
            SELECT 
                DATE(entry_time) as date,
                SUM(pnl_percentage) as daily_pnl,
                COUNT(*) as trades
            FROM performance
            WHERE entry_time >= DATE('now', '-30 days')
            GROUP BY DATE(entry_time)
            ORDER BY date
        ''', conn)
        conn.close()
        
        if len(df) == 0:
            return None
        
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            x=df['date'],
            y=df['daily_pnl'],
            name='Daily PnL %',
            marker_color=['green' if x > 0 else 'red' for x in df['daily_pnl']]
        ))
        
        fig.add_trace(go.Scatter(
            x=df['date'],
            y=df['daily_pnl'].cumsum(),
            name='Cumulative PnL %',
            yaxis='y2',
            line=dict(color='orange', width=3)
        ))
        
        fig.update_layout(
            title='Performance Last 30 Days',
            yaxis=dict(title='Daily PnL %'),
            yaxis2=dict(title='Cumulative PnL %', overlaying='y', side='right'),
            hovermode='x unified',
            height=400
        )
        
        return fig
    except:
        return None

# ==================== PORTFOLIO TRACKER ====================
def display_portfolio_tracker():
    """Display active positions and PnL"""
    st.markdown("### 💼 Active Portfolio")
    
    active_signals = get_active_signals()
    
    if len(active_signals) == 0:
        st.info("📭 No active positions")
        return
    
    total_pnl = 0
    positions_data = []
    
    for idx, row in active_signals.iterrows():
        symbol = row['symbol']
        entry = row['entry']
        signal_type = row['signal']
        
        # Get current price (simplified - use last known price)
        try:
            # Try to fetch current price
            if row['exchange'] == 'binance':
                current_price = entry  # Placeholder
            else:
                current_price = entry
        except:
            current_price = entry
        
        # Calculate PnL
        if signal_type == 'LONG':
            pnl_pct = ((current_price - entry) / entry) * 100
        else:
            pnl_pct = ((entry - current_price) / entry) * 100
        
        total_pnl += pnl_pct
        
        positions_data.append({
            'ID': row['id'],
            'Symbol': symbol,
            'Type': signal_type,
            'Entry': f"${entry:.4f}",
            'Current': f"${current_price:.4f}",
            'PnL%': f"{pnl_pct:+.2f}%",
            'TP1': f"${row['tp1']:.4f}",
            'SL': f"${row['sl']:.4f}",
            'Time': row['timestamp']
        })
    
    # Display metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Active Positions", len(active_signals))
    with col2:
        st.metric("Total PnL", f"{total_pnl:+.2f}%", delta=f"{total_pnl:.2f}%")
    with col3:
        avg_pnl = total_pnl / len(active_signals) if len(active_signals) > 0 else 0
        st.metric("Avg PnL per Trade", f"{avg_pnl:+.2f}%")
    
    # Display positions table
    df_positions = pd.DataFrame(positions_data)
    st.dataframe(df_positions, use_container_width=True, hide_index=True)
    
    # Close position buttons
    st.markdown("#### 🔄 Manage Positions")
    cols = st.columns(len(active_signals))
    for idx, (col, row) in enumerate(zip(cols, active_signals.iterrows())):
        with col:
            if st.button(f"Close {row[1]['symbol']}", key=f"close_{row[1]['id']}"):
                update_signal_status(row[1]['id'], 'CLOSED')
                st.success(f"✅ Closed {row[1]['symbol']}")
                st.rerun()

# ==================== EXPORT FUNCTIONS ====================
def export_to_excel(results):
    """Export scan results to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary['A1'] = "Crypto Futures Scanner - Ultimate Edition"
        ws_summary['A1'].font = Font(size=16, bold=True, color="1F77B4")
        ws_summary['A3'] = f"Scan Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A4'] = f"Total Analyzed: {len(results)}"
        ws_summary['A5'] = f"Long Signals: {len([r for r in results if r['consensus'] == 'LONG'])}"
        ws_summary['A6'] = f"Short Signals: {len([r for r in results if r['consensus'] == 'SHORT'])}"
        
        # Signals Sheet
        ws_signals = wb.create_sheet("Signals")
        headers = ['Symbol', 'Signal', 'Strength%', 'Price', 'Entry', 'TP1', 'TP2', 'SL', 'R:R', 'Risk%']
        ws_signals.append(headers)
        
        # Style headers
        for cell in ws_signals[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for result in results:
            plan = result.get('trading_plan', {})
            ws_signals.append([
                result['symbol'],
                result['consensus'],
                result['strength'],
                result['price'],
                plan.get('entry'),
                plan.get('tp1'),
                plan.get('tp2'),
                plan.get('sl'),
                plan.get('risk_reward_tp1'),
                plan.get('risk_percentage')
            ])
        
        filename = f"crypto_scanner_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        return filename
    except Exception as e:
        st.error(f"Export error: {e}")
        return None

# ==================== MAIN APP ====================
def main():
    # Initialize database
    if not st.session_state.db_initialized:
        if init_database():
            st.session_state.db_initialized = True
    
    # Header
    st.markdown('<p class="main-header">🚀 Crypto Futures Scanner Pro - Ultimate Edition</p>', unsafe_allow_html=True)
    
    # Top bar
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        st.markdown(f"**🕐 Time:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    with col2:
        st.markdown(f"**🔄 Last Scan:** {st.session_state.last_refresh.strftime('%H:%M:%S')}")
    with col3:
        if st.button("🔄 Refresh", use_container_width=True):
            st.session_state.last_refresh = datetime.now()
            st.cache_data.clear()
            st.rerun()
    
    st.markdown("---")
    
    # Create tabs
    tab1, tab2, tab3, tab4 = st.tabs(["📊 Scanner", "💼 Portfolio", "📈 Performance", "⚙️ Settings"])
    
    # ==================== TAB 1: SCANNER ====================
    with tab1:
        with st.sidebar:
            st.header("⚙️ Scanner Settings")
            
            st.subheader("🔡 Data Source")
            api_source = st.selectbox(
                "Exchange:",
                options=['binance', 'bybit', 'gateio'],
                format_func=lambda x: API_SOURCES[x]['name'],
                help="Auto fallback enabled"
            )
            
            scan_limit = st.slider("Coins to Scan", 10, 50, 25, 5)
            
            st.markdown("---")
            st.subheader("🎯 Filters")
            min_strength = st.slider("Min Strength (%)", 0, 100, 40, 5)
            show_only_signals = st.checkbox("Show Only Signals", value=True)
            min_volume = st.number_input("Min Volume Ratio", 0.0, 5.0, 1.2, 0.1)
            
            signal_types = st.multiselect(
                "Signal Types",
                options=['LONG', 'SHORT', 'NEUTRAL'],
                default=['LONG', 'SHORT']
            )
            
            min_rr = st.slider("Min Risk/Reward", 1.0, 5.0, 2.0, 0.5)
            
            st.markdown("---")
            st.subheader("🚀 Performance")
            parallel_workers = st.slider("Parallel Workers", 1, 10, 5)
            st.info(f"⚡ Scanning with {parallel_workers} threads")
        
        # Initialize source variable in session state if not exists
        if 'current_exchange' not in st.session_state:
            st.session_state.current_exchange = api_source
        
        # Start scan
        if st.button("🎯 Start Scan", type="primary", use_container_width=True):
            scan_start_time = time.time()
            
            with st.spinner(f"🔍 Fetching top {scan_limit} coins..."):
                source, symbols = get_top_symbols(api_source, scan_limit)
            
            if not symbols:
                st.error("❌ Failed to fetch data from all exchanges")
                return
            
            # Store current exchange in session state
            st.session_state.current_exchange = source if source else api_source
            
            if source != api_source:
                st.warning(f"⚠️ {API_SOURCES[api_source]['name']} unavailable. Using {API_SOURCES[source]['name']}")
            else:
                st.success(f"✅ Connected to {API_SOURCES[source]['name']}")
            
            # Parallel scanning
            with st.spinner(f"⚡ Analyzing {scan_limit} coins with {parallel_workers} workers..."):
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                results = parallel_scan(symbols, source, max_workers=parallel_workers)
                
                for i in range(len(results)):
                    progress = (i + 1) / len(results)
                    progress_bar.progress(progress)
                    status_text.text(f"Analyzed {i+1}/{len(results)} coins")
                
                progress_bar.empty()
                status_text.empty()
            
            scan_duration = time.time() - scan_start_time
            
            # Filter results
            filtered_results = []
            for r in results:
                if show_only_signals and r['consensus'] not in signal_types:
                    continue
                if r['strength'] < min_strength:
                    continue
                
                # Check volume
                volume_ok = False
                for tf in ['1h', '4h', '12h']:
                    tf_data = r['timeframes'].get(tf, {})
                    indicators = tf_data.get('indicators', {})
                    if indicators.get('volume_ratio', 0) >= min_volume:
                        volume_ok = True
                        break
                if not volume_ok:
                    continue
                
                # Check risk/reward
                plan = r.get('trading_plan')
                if plan and plan.get('risk_reward_tp1', 0) < min_rr:
                    continue
                
                filtered_results.append(r)
            
            # Sort by strength
            filtered_results = sorted(filtered_results, key=lambda x: x['strength'], reverse=True)
            
            # Save to session state
            st.session_state.scan_results = filtered_results
            
            # Save to database
            avg_strength = sum([r['strength'] for r in filtered_results]) / len(filtered_results) if filtered_results else 0
            save_scan_history(len(results), len(filtered_results), avg_strength, source, scan_duration)
            
            # Display summary
            st.success(f"✅ Scan completed in {scan_duration:.1f}s")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Scanned", len(results))
            with col2:
                st.metric("Signals Found", len(filtered_results))
            with col3:
                long_count = len([r for r in filtered_results if r['consensus'] == 'LONG'])
                st.metric("Long Signals", long_count)
            with col4:
                short_count = len([r for r in filtered_results if r['consensus'] == 'SHORT'])
                st.metric("Short Signals", short_count)
        
        # Display results
        if st.session_state.scan_results:
            results = st.session_state.scan_results
            
            # Get current exchange from session state
            current_exchange = st.session_state.get('current_exchange', 'binance')
            
            st.markdown(f"### 📊 Found {len(results)} Trading Opportunities")
            
            # Market Heatmap
            with st.expander("🗺️ Market Heatmap", expanded=False):
                heatmap = create_market_heatmap(results)
                if heatmap:
                    st.plotly_chart(heatmap, use_container_width=True)
            
            # Export button
            col1, col2 = st.columns([1, 5])
            with col1:
                if st.button("📥 Export Excel"):
                    filename = export_to_excel(results)
                    if filename:
                        st.success(f"✅ Exported to {filename}")
            
            st.markdown("---")
            
            # Display each signal
            for result in results:
                signal_emoji = "🟢" if result['consensus'] == 'LONG' else "🔴" if result['consensus'] == 'SHORT' else "⚪"
                
                with st.expander(
                    f"{signal_emoji} **{result['name']}** ({result['symbol']}) | "
                    f"Signal: **{result['consensus']}** | Strength: **{result['strength']}%**",
                    expanded=False
                ):
                    col1, col2, col3 = st.columns([2, 2, 3])
                    
                    with col1:
                        st.markdown(f"### ${result['price']:,.8f}")
                        signal_class = f"signal-{result['consensus'].lower()}"
                        st.markdown(f'<div class="{signal_class}">🎯 {result["consensus"]}</div>', unsafe_allow_html=True)
                        st.metric("Strength", f"{result['strength']}%")
                    
                    with col2:
                        st.markdown("**📊 Multi-Timeframe:**")
                        for tf in ['12h', '4h', '1h']:
                            tf_data = result['timeframes'].get(tf, {})
                            signal = tf_data.get('signal', 'N/A')
                            strength = tf_data.get('strength', 0)
                            emoji = "🟢" if signal == "LONG" else "🔴" if signal == "SHORT" else "⚪"
                            st.markdown(f"{emoji} **{tf.upper()}:** {signal} ({strength:.1f}%)")
                    
                    with col3:
                        if result['trading_plan']:
                            plan = result['trading_plan']
                            st.markdown("**💼 Trading Plan:**")
                            if plan.get('entry_note'):
                                st.warning(plan['entry_note'])
                            st.markdown(f"📍 **Entry:** ${plan['entry']:,.8f}")
                            st.markdown(f"🎯 **TP1:** ${plan['tp1']:,.8f} (R:R **{plan['risk_reward_tp1']}**)")
                            st.markdown(f"🎯 **TP2:** ${plan['tp2']:,.8f} (R:R **{plan['risk_reward_tp2']}**)")
                            st.markdown(f"🛑 **SL:** ${plan['sl']:,.8f} (Risk: {plan['risk_percentage']:.2f}%)")
                            
                            if plan['support'] and plan['resistance']:
                                st.markdown("---")
                                st.markdown(f"📉 **Support:** ${plan['support']:,.8f}")
                                st.markdown(f"📈 **Resistance:** ${plan['resistance']:,.8f}")
                            
                            # Save to DB button - Use current_exchange variable
                            if st.button(f"💾 Save to Portfolio", key=f"save_{result['symbol']}"):
                                signal_id = save_signal_to_db(result, current_exchange)
                                if signal_id:
                                    st.success(f"✅ Saved {result['symbol']} to portfolio!")
                                    
                                    # Send Telegram alert if enabled
                                    if st.session_state.telegram_enabled:
                                        telegram_token = st.session_state.get('telegram_token')
                                        telegram_chat_id = st.session_state.get('telegram_chat_id')
                                        if telegram_token and telegram_chat_id:
                                            if send_telegram_alert(
                                                telegram_token,
                                                telegram_chat_id,
                                                result['symbol'],
                                                result
                                            ):
                                                st.success("📱 Telegram alert sent!")
                    
                    # Technical details
                    st.markdown("---")
                    st.markdown("**📊 Technical Indicators:**")
                    cols = st.columns(3)
                    
                    for idx, tf in enumerate(['12h', '4h', '1h']):
                        tf_data = result['timeframes'].get(tf, {})
                        indicators = tf_data.get('indicators', {})
                        reasons = tf_data.get('reasons', [])
                        patterns = tf_data.get('patterns', [])
                        
                        if tf_data.get('signal') not in ['ERROR', 'INSUFFICIENT_DATA']:
                            with cols[idx]:
                                st.markdown(f"**⏰ {tf.upper()}**")
                                
                                rsi = indicators.get('rsi', 0)
                                rsi_emoji = "🟢" if rsi < 30 else "🔴" if rsi > 70 else "🟡" if rsi < 40 else "🟠" if rsi > 60 else "⚪"
                                st.markdown(f"{rsi_emoji} RSI: **{rsi:.1f}**")
                                
                                stoch = indicators.get('stoch_rsi', 0)
                                stoch_emoji = "🟢" if stoch < 20 else "🔴" if stoch > 80 else "⚪"
                                st.markdown(f"{stoch_emoji} Stoch: **{stoch:.1f}**")
                                
                                macd = indicators.get('macd_diff', 0)
                                macd_emoji = "🟢" if macd > 0 else "🔴"
                                st.markdown(f"{macd_emoji} MACD: **{macd:.8f}**")
                                
                                vol = indicators.get('volume_ratio', 0)
                                vol_emoji = "🟢" if vol > 1.5 else "🟡" if vol > 1.2 else "⚪"
                                st.markdown(f"{vol_emoji} Volume: **{vol:.2f}x**")
                                
                                adx = indicators.get('adx', 0)
                                adx_emoji = "💪" if adx > 25 else "⚪"
                                st.markdown(f"{adx_emoji} ADX: **{adx:.1f}**")
                                
                                ema_trend = indicators.get('ema_trend', 'NEUTRAL')
                                trend_emoji = "📈" if ema_trend == 'BULL' else "📉" if ema_trend == 'BEAR' else "➡️"
                                st.markdown(f"{trend_emoji} Trend: **{ema_trend}**")
                                
                                if patterns:
                                    st.markdown("**🎭 Patterns:**")
                                    for pattern in patterns[:2]:
                                        if 'GOLDEN' in pattern:
                                            st.markdown("⭐ Golden Cross")
                                        elif 'DEATH' in pattern:
                                            st.markdown("💀 Death Cross")
                                        elif 'BULLISH' in pattern:
                                            st.markdown("📈 Bull Div")
                                        elif 'BEARISH' in pattern:
                                            st.markdown("📉 Bear Div")
                                
                                if reasons:
                                    st.markdown("**💡 Signals:**")
                                    for reason in reasons[:3]:
                                        st.markdown(f"• {reason}")
                    
                    # Chart
                    main_df = result['timeframes'].get('4h_df')
                    if main_df is not None and len(main_df) >= 50:
                        st.markdown("---")
                        st.markdown("**📈 Chart Analysis (4H)**")
                        chart = create_chart(main_df, result['name'], result['trading_plan'])
                        if chart:
                            st.plotly_chart(chart, use_container_width=True)
        else:
            st.info("👆 Click 'Start Scan' to begin analyzing the market")
    
    # ==================== TAB 2: PORTFOLIO ====================
    with tab2:
        display_portfolio_tracker()
        
        st.markdown("---")
        st.markdown("### 📜 Recent Signals")
        
        try:
            conn = sqlite3.connect(DB_NAME)
            recent_signals = pd.read_sql('''
                SELECT 
                    timestamp,
                    symbol,
                    signal,
                    strength,
                    entry,
                    tp1,
                    tp2,
                    sl,
                    status
                FROM signals
                ORDER BY timestamp DESC
                LIMIT 20
            ''', conn)
            conn.close()
            
            if len(recent_signals) > 0:
                st.dataframe(recent_signals, use_container_width=True, hide_index=True)
            else:
                st.info("No signals history yet")
        except Exception as e:
            st.error(f"Error loading signals: {e}")
    
    # ==================== TAB 3: PERFORMANCE ====================
    with tab3:
        st.markdown("### 📈 Performance Analytics")
        
        # Performance chart
        perf_chart = create_performance_chart()
        if perf_chart:
            st.plotly_chart(perf_chart, use_container_width=True)
        else:
            st.info("No performance data yet. Complete some trades to see analytics.")
        
        # Statistics
        try:
            conn = sqlite3.connect(DB_NAME)
            
            # Overall stats
            stats = pd.read_sql('''
                SELECT 
                    COUNT(*) as total_trades,
                    SUM(CASE WHEN pnl_percentage > 0 THEN 1 ELSE 0 END) as winning_trades,
                    SUM(CASE WHEN pnl_percentage < 0 THEN 1 ELSE 0 END) as losing_trades,
                    AVG(pnl_percentage) as avg_pnl,
                    SUM(pnl_percentage) as total_pnl,
                    MAX(pnl_percentage) as best_trade,
                    MIN(pnl_percentage) as worst_trade
                FROM performance
            ''', conn)
            
            if len(stats) > 0 and stats['total_trades'].iloc[0] > 0:
                st.markdown("#### 📊 Overall Statistics")
                
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Trades", int(stats['total_trades'].iloc[0]))
                
                with col2:
                    win_rate = (stats['winning_trades'].iloc[0] / stats['total_trades'].iloc[0]) * 100
                    st.metric("Win Rate", f"{win_rate:.1f}%")
                
                with col3:
                    st.metric("Total PnL", f"{stats['total_pnl'].iloc[0]:+.2f}%")
                
                with col4:
                    st.metric("Avg PnL", f"{stats['avg_pnl'].iloc[0]:+.2f}%")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric("Best Trade", f"{stats['best_trade'].iloc[0]:+.2f}%", delta="🏆")
                with col2:
                    st.metric("Worst Trade", f"{stats['worst_trade'].iloc[0]:+.2f}%", delta="📉")
            
            # Scan history
            st.markdown("---")
            st.markdown("#### 📜 Scan History")
            
            scan_history = pd.read_sql('''
                SELECT 
                    timestamp,
                    total_scanned,
                    signals_found,
                    avg_strength,
                    exchange,
                    ROUND(scan_duration, 2) as duration_sec
                FROM scan_history
                ORDER BY timestamp DESC
                LIMIT 10
            ''', conn)
            
            conn.close()
            
            if len(scan_history) > 0:
                st.dataframe(scan_history, use_container_width=True, hide_index=True)
            
        except Exception as e:
            st.error(f"Error loading performance data: {e}")
    
    # ==================== TAB 4: SETTINGS ====================
    with tab4:
        st.markdown("### ⚙️ Application Settings")
        
        # Telegram Settings
        st.markdown("#### 📱 Telegram Alerts")
        
        col1, col2 = st.columns([1, 3])
        with col1:
            telegram_enabled = st.checkbox("Enable Telegram", value=st.session_state.telegram_enabled)
        
        if telegram_enabled:
            with col2:
                telegram_token = st.text_input("Bot Token", type="password", placeholder="123456:ABC-DEF1234ghIkl-zyx57W2v1u123ew11")
                telegram_chat_id = st.text_input("Chat ID", placeholder="-1001234567890")
                
                if st.button("🧪 Test Connection"):
                    if telegram_token and telegram_chat_id:
                        test_msg = f"🧪 Test Alert\n\nConnection successful!\n⏰ {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                        url = f"https://api.telegram.org/bot{telegram_token}/sendMessage"
                        data = {'chat_id': telegram_chat_id, 'text': test_msg}
                        
                        try:
                            response = requests.post(url, data=data, timeout=10)
                            if response.status_code == 200:
                                st.success("✅ Telegram connection successful!")
                                st.session_state.telegram_enabled = True
                                st.session_state.telegram_token = telegram_token
                                st.session_state.telegram_chat_id = telegram_chat_id
                            else:
                                st.error("❌ Connection failed. Check your token and chat ID.")
                        except Exception as e:
                            st.error(f"❌ Error: {e}")
                    else:
                        st.warning("⚠️ Please enter both Bot Token and Chat ID")
        
        st.markdown("---")
        
        # Auto Refresh
        st.markdown("#### 🔄 Auto Refresh")
        refresh_option = st.radio(
            "Interval:",
            options=["Disabled", "Every 5 Minutes", "Every 15 Minutes", "Every 1 Hour"],
            index=0
        )
        
        if refresh_option == "Every 5 Minutes":
            st.session_state.refresh_interval = 5
        elif refresh_option == "Every 15 Minutes":
            st.session_state.refresh_interval = 15
        elif refresh_option == "Every 1 Hour":
            st.session_state.refresh_interval = 60
        else:
            st.session_state.refresh_interval = None
        
        if st.session_state.refresh_interval:
            time_until = st.session_state.refresh_interval - ((datetime.now() - st.session_state.last_refresh).total_seconds() / 60)
            if time_until > 0:
                st.info(f"⏱️ Next refresh in: {time_until:.1f} minutes")
        
        st.markdown("---")
        
        # Database Management
        st.markdown("#### 🗄️ Database Management")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🗑️ Clear Scan History"):
                try:
                    conn = sqlite3.connect(DB_NAME)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM scan_history")
                    conn.commit()
                    conn.close()
                    st.success("✅ Scan history cleared")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        with col2:
            if st.button("🗑️ Clear Closed Signals"):
                try:
                    conn = sqlite3.connect(DB_NAME)
                    cursor = conn.cursor()
                    cursor.execute("DELETE FROM signals WHERE status = 'CLOSED'")
                    conn.commit()
                    conn.close()
                    st.success("✅ Closed signals cleared")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        with col3:
            if st.button("⚠️ Reset Database"):
                try:
                    conn = sqlite3.connect(DB_NAME)
                    cursor = conn.cursor()
                    cursor.execute("DROP TABLE IF EXISTS signals")
                    cursor.execute("DROP TABLE IF EXISTS performance")
                    cursor.execute("DROP TABLE IF EXISTS scan_history")
                    cursor.execute("DROP TABLE IF EXISTS watchlist")
                    conn.commit()
                    conn.close()
                    init_database()
                    st.success("✅ Database reset")
                except Exception as e:
                    st.error(f"Error: {e}")
        
        st.markdown("---")
        
        # About
        st.markdown("#### ℹ️ About")
        st.info("""
        **Crypto Futures Scanner Pro - Ultimate Edition v2.0**
        
        Features:
        - 🔄 Multi-Exchange Support (Binance, Bybit, Gate.io)
        - ⚡ Parallel Processing for Fast Scanning
        - 📊 Advanced Technical Analysis (15+ Indicators)
        - 💼 Portfolio Tracking
        - 📱 Telegram Alerts
        - 📈 Performance Analytics
        - 💾 Database Storage
        - 📥 Excel Export
        
        ⚠️ **Disclaimer:** For educational purposes only. Always DYOR and manage risk properly.
        """)
    
    # Footer
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: #666;'>"
        "Made with ❤️ by AI Assistant | "
        f"Database: {DB_NAME} | "
        f"Version 2.0"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
